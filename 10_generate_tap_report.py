#!/usr/bin/env python3
"""
10_generate_tap_report.py -- Generate a completed field Tap Report workbook.

This script combines two post-pipeline data sources to produce the Tap Report
used in the field during splicing and testing:

  1. HAF report       -- address-to-tap assignments exported from the GIS
                         system. The COMMENT column holds the tap name
                         (e.g. RC73E_FT_001); address columns provide the
                         civic address for each customer premise served.

  2. Asbuilt workbook -- the finished splice workbook produced by step 8.
                         Each FT sheet is scanned for PORT rows (identified
                         by the PORT label in column J, placed there by step
                         7). The buffer and fiber values on those rows form
                         the burn summary written to the Tap Report.

For each tap the script produces one row in the report containing:
  - Col A : tap name
  - Col B : port-by-port address list (unused ports labeled DARK)
  - Col C : installation type (UNDERGROUND or AERIAL)
  - Col D : total port count for the tap block (2 / 4 / 8 / 12)
  - Col E : number of active (live) ports
  - Col F : burn summary (e.g. BL / BL,OR) with directional background color

SE enclosure names are appended after the tap rows as reference rows.
Cell B3 of the Tap Report sheet is filled with the OLT site identifier.

Reads:
  data/Tap_Report_Template.xlsx
  data/*HAF*.xlsx               (auto-detected; or pass as first argument)
  output/Asbuilt_Workbook_post12.xlsx  (or pass as second argument)

Writes:
  output/{OLT Name} Tap Report.xlsx

Usage:
  python3 10_generate_tap_report.py                          # auto-detects files
  python3 10_generate_tap_report.py <haf.xlsx> <asbuilt.xlsx>
"""

import re
import sys
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from naming_utils import (
    classify_sheet_name,
    find_col_by_header,
    find_header_row,
    parse_location_id,
    safe_fill_hex,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
TEMPLATE   = DATA_DIR / "Tap_Report_Template.xlsx"
ASBUILT    = OUTPUT_DIR / "Asbuilt_Workbook_post12.xlsx"


def _find_haf() -> Path:
    """
    Locate the HAF report in the data directory by glob pattern.
    Raises FileNotFoundError if no matching file is present.
    """
    matches = sorted(DATA_DIR.glob("*HAF*"))
    if not matches:
        raise FileNotFoundError("No HAF file found in data/ — place it there and retry.")
    return matches[0]


# ---------------------------------------------------------------------------
# Formatting constants
#
# Background colors are matched to the live Tap Report standard.
# BG_A applies to the tap name column (col A); BG_BEG applies to all other
# data columns. Col F overrides BG_BEG with the directional color extracted
# from the PORT row background in the asbuilt workbook.
# ---------------------------------------------------------------------------

BG_A   = "BDD6EE"   # light steel blue  — tap name column (col A)
BG_BEG = "D9E2F3"   # pale blue         — address, hookup type, port count cols
FONT   = "Arial"


def _fill(hex6: str) -> PatternFill:
    """Return a solid PatternFill for the given 6-character RGB hex color."""
    return PatternFill(fill_type="solid", fgColor=hex6.lstrip("#"))


def _font(size: int = 11, bold: bool = False, color: str = "FF000000") -> Font:
    """Return an Arial Font object. color must be 8-character ARGB (default: opaque black)."""
    return Font(name=FONT, size=size, bold=bold, color=color)


def _align(horiz: str = "center", vert: str = "center", wrap: bool = False) -> Alignment:
    """Return an Alignment object with the given horizontal, vertical, and wrap settings."""
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)


def _sort_key(name: str) -> int:
    """Return the trailing numeric suffix of a location name for natural sort ordering."""
    m = re.search(r"\d+$", str(name))
    return int(m.group()) if m else 0


# ---------------------------------------------------------------------------
# STEP 1: Parse HAF file → tap address records
# ---------------------------------------------------------------------------

def parse_haf(haf_path: Path) -> tuple[str, dict]:
    """
    Read the HAF report and return (olt_name, records).

    olt_name is inferred from the first non-empty COMMENT value using
    naming_utils.parse_location_id (e.g. 'RC73E_FT_001' yields 'RC73E').

    records maps each tap name to a dict:
        {
          "addresses":   [{"text": str, "street_name": str, "house_num": int}, ...],
          "hookup_type": str  (e.g. 'UNDERGROUND')
        }

    Addresses within each tap are sorted by street name, then house number,
    so that port assignments are consistent across runs.
    """
    wb      = openpyxl.load_workbook(haf_path, read_only=True, data_only=True)
    ws      = wb.active
    headers = [str(c or "").strip().upper() for c in next(ws.iter_rows(values_only=True))]

    def idx(col_name: str) -> int:
        try:
            return headers.index(col_name)
        except ValueError:
            raise ValueError(f"Required column '{col_name}' not found in HAF")

    loc_idx    = idx("COMMENT")
    house_idx  = idx("HOUSE NUMBER")
    predir_idx = idx("PRE DIRECTION")
    street_idx = idx("STREET NAME")
    type_idx   = idx("STREET TYPE")
    hookup_idx = idx("HOOKUP TYPE")

    records: dict[str, dict] = {}
    olt_name: str | None     = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        location = str(row[loc_idx] or "").strip()
        if not location:
            continue

        if olt_name is None:
            parsed   = parse_location_id(location)
            olt_name = parsed.get("olt") or None

        if location not in records:
            records[location] = {
                "addresses":   [],
                "hookup_type": str(row[hookup_idx] or "").strip(),
            }

        pre_dir = str(row[predir_idx] or "").strip()
        parts   = [
            str(row[house_idx]  or "").strip(),
            pre_dir,
            str(row[street_idx] or "").strip(),
            str(row[type_idx]   or "").strip(),
        ]
        addr = " ".join(p for p in parts if p)
        records[location]["addresses"].append({
            "text":        addr,
            "street_name": str(row[street_idx] or "").strip(),
            "house_num":   int(row[house_idx]) if str(row[house_idx] or "").isdigit() else 0,
        })

    wb.close()

    # Sort addresses within each tap: by street name, then house number
    for entry in records.values():
        entry["addresses"].sort(key=lambda a: (a["street_name"], a["house_num"]))

    return olt_name or "UNKNOWN", records


# ---------------------------------------------------------------------------
# STEP 2: Parse asbuilt workbook → burn summaries per FT sheet
# ---------------------------------------------------------------------------

_PORT_SIZES = [2, 4, 8, 12]


def parse_burn_summaries(asbuilt_path: Path) -> dict:
    """
    Scan every FT sheet in the asbuilt workbook and extract the burn summary
    for each tap.

    A PORT row is identified by a 'PORTn' label in column J (col 10), which
    is placed there by step 7 (7_process_taps.py). For each PORT row the
    BUFFER and FIBER values are read from the SHEATHS header row columns.
    Fibers are grouped under their buffer tube, and the background color of
    the buffer cell is retained as the directional color for col F of the
    Tap Report.

    Returns a dict mapping tap sheet name (upper-cased) to:
        {
          "summary_text": str   (e.g. 'BL / BL,OR\\nOR / GR'),
          "total_fibers": int   (number of PORT rows = active fiber count),
          "block_size":   int   (2 / 4 / 8 / 12 — smallest block >= total),
          "bg_color":     str   (6-char RGB hex of the first buffer cell)
        }
    """
    wb       = openpyxl.load_workbook(asbuilt_path, data_only=True)
    burn_map = {}

    for sheet_name in wb.sheetnames:
        if classify_sheet_name(sheet_name) != "T":
            continue

        ws = wb[sheet_name]

        # Locate BUFFER and FIBER columns from the SHEATHS header row.
        hdr_row = find_header_row(ws, "CONNECTION")
        if hdr_row is None:
            continue

        conn_col   = find_col_by_header(ws, hdr_row, "CONNECTION")
        buffer_col = find_col_by_header(ws, hdr_row, "BUFFER",
                                        max_col=conn_col if conn_col else None)
        fiber_col  = find_col_by_header(ws, hdr_row, "FIBER",
                                        max_col=conn_col if conn_col else None)

        # Fallback to columns E and F if header detection fails.
        if buffer_col is None:
            buffer_col = 5
        if fiber_col is None:
            fiber_col = 6

        ports: list[dict] = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            port_label = str(ws.cell(r, 10).value or "").strip().upper()
            if not port_label.startswith("PORT"):
                continue

            buffer = str(ws.cell(r, buffer_col).value or "").strip().upper()
            fiber  = str(ws.cell(r, fiber_col).value  or "").strip().upper()
            if not buffer or not fiber:
                continue

            hex_color = safe_fill_hex(ws.cell(r, buffer_col)) or "FFFFFF"
            ports.append({"buffer": buffer, "fiber": fiber, "color": hex_color})

        if not ports:
            continue

        # Group unique fibers under each buffer; preserve insertion order.
        buffer_map:    dict[str, list[str]] = {}
        buffer_colors: dict[str, str]       = {}
        for p in ports:
            b = p["buffer"]
            if b not in buffer_map:
                buffer_map[b]    = []
                buffer_colors[b] = p["color"]
            if p["fiber"] not in buffer_map[b]:
                buffer_map[b].append(p["fiber"])

        total_fibers = len(ports)
        block_size   = next((x for x in _PORT_SIZES if x >= total_fibers), 12)
        summary_text = "\n".join(
            f"{buf} / {','.join(fibers)}"
            for buf, fibers in buffer_map.items()
        )
        # Background color comes from the first buffer encountered.
        bg_color = buffer_colors[next(iter(buffer_colors))]

        burn_map[sheet_name.upper()] = {
            "summary_text": summary_text,
            "total_fibers": total_fibers,
            "block_size":   block_size,
            "bg_color":     bg_color,
        }

    wb.close()
    return burn_map


# ---------------------------------------------------------------------------
# STEP 3: Collect SE sheet names for the trailing section
# ---------------------------------------------------------------------------

def collect_se_names(asbuilt_path: Path) -> list[str]:
    """
    Return all SE enclosure sheet names from the asbuilt workbook, sorted by
    numeric suffix. These are appended as reference rows at the bottom of the
    Tap Report after all tap rows have been written.
    """
    wb = openpyxl.load_workbook(asbuilt_path, read_only=True)
    names = [n for n in wb.sheetnames if classify_sheet_name(n) == "S"]
    wb.close()
    return sorted(names, key=_sort_key)


# ---------------------------------------------------------------------------
# STEP 4: Write the Tap Report
# ---------------------------------------------------------------------------

def write_tap_report(
    olt_name:        str,
    records:         dict,
    burn_summaries:  dict,
    se_names:        list[str],
    template_path:   Path,
    output_path:     Path,
) -> None:
    """
    Copy the Tap Report template and populate the 'Tap Report' sheet.

    Writes one row per tap starting at row 10 (the first data row in the
    template). Columns A-F are filled for each tap; formatting matches the
    established field report standard (Arial 11pt, blue background tones,
    directional color on col F). SE enclosure names are appended after the
    last tap row as reference-only rows (col A only).

    Cell B3 is written with the OLT site identifier (e.g. 'RC73E') to
    populate the Node Number field in the template header block.
    """
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Tap Report"]

    fill_a   = _fill(BG_A)
    fill_beg = _fill(BG_BEG)

    # Write OLT name into B3 (Node Number field in the template header)
    ws["B3"] = olt_name

    sorted_locs = sorted(records.keys(), key=_sort_key)
    row         = 10  # data starts at row 10 per template

    # --- Tap rows -----------------------------------------------------------
    for loc in sorted_locs:
        entry    = records[loc]
        addrs    = entry["addresses"]
        num_ports = next((p for p in _PORT_SIZES if p >= len(addrs)), 12)

        port_lines = [f"Port {i+1}: {a['text']}" for i, a in enumerate(addrs)]
        for i in range(len(addrs), num_ports):
            port_lines.append(f"Port {i+1}: DARK")
        address_text = "\n".join(port_lines)

        burn        = burn_summaries.get(loc.upper(), {})
        block_size  = burn.get("block_size",   num_ports)
        tot_fibers  = burn.get("total_fibers", len(addrs))
        burn_text   = burn.get("summary_text", "")
        burn_bg     = burn.get("bg_color",     BG_BEG)

        # Col A — Tap name
        c = ws.cell(row, 1, loc)
        c.font = _font(11, bold=True)
        c.alignment = _align("center", "center", wrap=True)
        c.fill = fill_a

        # Col B — Multiline port addresses
        c = ws.cell(row, 2, address_text)
        c.font = _font(11)
        c.alignment = _align("left", "center", wrap=True)
        c.fill = fill_beg

        # Col C — Aerial/Underground
        c = ws.cell(row, 3, entry["hookup_type"])
        c.font = _font(11)
        c.alignment = _align("center", "center")
        c.fill = fill_beg

        # Col D — Total port count (block size)
        c = ws.cell(row, 4, block_size)
        c.font = _font(11)
        c.alignment = _align("center", "center")
        c.fill = fill_beg

        # Col E — Active fibers (non-DARK ports spliced)
        c = ws.cell(row, 5, tot_fibers)
        c.font = _font(11)
        c.alignment = _align("center", "center")
        c.fill = fill_beg

        # Col F — Burn summary with directional background color; text always black
        c = ws.cell(row, 6, burn_text)
        c.font = _font(11, color="FF000000")
        c.alignment = _align("center", "center", wrap=True)
        c.fill = _fill(burn_bg)

        row += 1

    # --- SE / splitter enclosure rows (names only, no address data) ---------
    existing_upper = {loc.upper() for loc in sorted_locs}
    for se_name in se_names:
        if se_name.upper() in existing_upper:
            continue
        c = ws.cell(row, 1, se_name)
        c.font = _font(11, bold=True)
        c.alignment = _align("center", "center", wrap=True)
        c.fill = fill_a
        existing_upper.add(se_name.upper())
        row += 1

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    """
    Entry point. Resolves input paths, runs each processing stage in order,
    and writes the completed Tap Report to the output directory.
    """
    global DATA_DIR, OUTPUT_DIR, TEMPLATE, ASBUILT
    DATA_DIR   = Path(data_dir)
    OUTPUT_DIR = Path(output_dir)
    TEMPLATE   = DATA_DIR / "Tap_Report_Template.xlsx"
    ASBUILT    = OUTPUT_DIR / "Asbuilt_Workbook_post12.xlsx"
    haf_path    = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_haf()
    asbuilt     = Path(sys.argv[2]) if len(sys.argv) > 2 else ASBUILT

    print(f"HAF file : {haf_path.name}")
    print(f"Asbuilt  : {asbuilt.name}")

    olt_name, records = parse_haf(haf_path)
    print(f"OLT name : {olt_name}")
    print(f"Tap count: {len(records)}")

    burn_summaries = parse_burn_summaries(asbuilt)
    print(f"Burn summaries: {len(burn_summaries)} tap sheets found")

    se_names = collect_se_names(asbuilt)
    print(f"SE sheets to append: {len(se_names)}")

    output_path = OUTPUT_DIR / f"{olt_name} Tap Report.xlsx"
    write_tap_report(olt_name, records, burn_summaries, se_names, TEMPLATE, output_path)

    print(f"\nOutput   : {output_path}")


if __name__ == "__main__":
    main()

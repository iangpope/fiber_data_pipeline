
import re
from openpyxl import load_workbook

INPUT_FILE = "output/Combined_Final_Shifted.xlsx"
OUTPUT_FILE = "output/Combined_Final_Shifted_B3_Labeled.xlsx"

legacy_tap_rx = re.compile(r"^MIC[A-Z]{2}\d{2}\d{4}$", re.IGNORECASE)


def is_tap_sheet(sheet_name: str, ws) -> bool:
    name_up = sheet_name.upper()
    if "_FT_" in name_up or name_up.endswith("_FT"):
        return True
    if legacy_tap_rx.match(name_up):
        return True
    b3 = ws["B3"].value
    if isinstance(b3, str) and "PORT OTE" in b3.upper():
        return True
    # last-resort: presence of PORT NAME header
    for c in range(1, min(ws.max_column, 30) + 1):
        v = ws.cell(row=11, column=c).value
        if isinstance(v, str) and v.strip().upper() == "PORT NAME":
            return True
    return False

def detect_port_column(sheet):
    # After script 9, PORT labels are typically in J/K, but scan a small band.
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        for col in range(8, 16):  # H..O
            if len(row) >= col:
                v = row[col - 1].value
                if v and str(v).strip().upper().startswith("PORT"):
                    return col
    return None

def count_ports(sheet, port_col):
    count = 0
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        if len(row) >= port_col:
            val = row[port_col - 1].value
            if val and str(val).strip().upper().startswith("PORT"):
                count += 1
    return count

def label_port_ote(sheet, count):
    if count <= 2:
        sheet["B3"] = "2 PORT OTE"
    elif count <= 4:
        sheet["B3"] = "4 PORT OTE"
    elif count <= 8:
        sheet["B3"] = "8 PORT OTE"
    elif count <= 12:
        sheet["B3"] = "12 PORT OTE"

def label_b3_in_workbook(input_file, output_file):
    wb = load_workbook(input_file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if not is_tap_sheet(sheet_name, sheet):
            continue

        port_col = detect_port_column(sheet)
        if not port_col:
            continue

        port_count = count_ports(sheet, port_col)
        if port_count > 0:
            label_port_ote(sheet, port_count)
            # Taps are always 1 tray in your workflow
            sheet["B4"] = 1

    wb.save(output_file)

if __name__ == "__main__":
    label_b3_in_workbook(INPUT_FILE, OUTPUT_FILE)

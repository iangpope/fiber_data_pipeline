import openpyxl
import re

"""11_final.py

Final cleanups after labeling B3/B4 and shifting ports.

Change made (Jan 2026): remove hard-coded absolute macOS paths so the
pipeline can run from any folder.

Inputs/outputs are now relative to the project root.
"""

# Input and output filenames (relative)
INPUT_FILE = "output/Combined_Final_Shifted_B3_Labeled.xlsx"
OUTPUT_FILE = "output/Asbuilt_Workbook.xlsx"



# Compile regex for legacy naming pattern (e.g., MICMS02S007 or MICMS02D010)
legacy_name_rx = re.compile(r'^(?P<prefix>[A-Z]{2}[A-Z0-9]+?)(?P<type>S|D)?(?P<num>\d{3,4})$', re.IGNORECASE)

# Open the workbook
wb = openpyxl.load_workbook(INPUT_FILE)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    # Skip sheets that don't have the expected metadata structure
    # (We expect "Enclosure:" in A3 and "No. of Trays:" in A4 for sheets representing locations)
    if sheet.cell(row=3, column=1).value != "Enclosure:" or sheet.cell(row=4, column=1).value != "No. of Trays:":
        continue

    # Fill missing Enclosure and Trays based on naming conventions
    enclosure_cell = sheet.cell(row=3, column=2)  # B3
    trays_cell = sheet.cell(row=4, column=2)      # B4
    if (enclosure_cell.value is None or str(enclosure_cell.value).strip() == "") \
       or (trays_cell.value is None or str(trays_cell.value).strip() == ""):
        name_up = sheet_name.upper()
        enclosure_val = None
        trays_val = None
        if "_FT_" in name_up or name_up.endswith("_FT"):
            enclosure_val = "2 PORT OTE"
            trays_val = 1
        elif "_SE_" in name_up or name_up.endswith("_SE") or re.search(r'S\d+$', name_up):
            enclosure_val = "COMMSCOPE FOSC 450-D"
            trays_val = 2
        elif re.search(r'D\d+$', name_up):
            enclosure_val = "COMMSCOPE FOSC 450-B"
            trays_val = 1
        else:
            # Try legacy regex as fallback
            m = legacy_name_rx.match(name_up)
            if m:
                type_letter = (m.group("type") or "").upper()
                if type_letter == "S":
                    enclosure_val = "COMMSCOPE FOSC 450-D"
                    trays_val = 2
                elif type_letter == "D":
                    enclosure_val = "COMMSCOPE FOSC 450-B"
                    trays_val = 1
        # Only set if we determined values (otherwise leave as is)
        if enclosure_val:
            enclosure_cell.value = enclosure_val
        if trays_val is not None:
            trays_cell.value = trays_val

    # Identify the top metadata and content section boundaries
    # Meta data is rows 2-7 (labels in col A, values in col B), followed by a blank row.
    # Find the first blank row after metadata (should be row 8 if meta is present).
    # Then identify device lines vs connection lines.
    max_row = sheet.max_row
    meta_end_row = 7  # Based on known format (Splice ID through Longitude are rows 2-7)
    meta_blank_row = meta_end_row + 1  # expected blank row after metadata
    if meta_blank_row > max_row:
        # If no content beyond metadata, skip
        continue

    # Determine first connection line (where column B is empty while column A has a value)
    first_connection_row = None
    sheaths_row = None

    # Find the "SHEATHS" row which indicates the start of main splice data section
    for cell in sheet['A']:
        if cell.value == "SHEATHS":
            sheaths_row = cell.row
            break

    if sheaths_row is None:
        # If no SHEATHS label found, skip this sheet (no splice data)
        continue

    # Find first connection line by scanning from the end of metadata to just before SHEATHS
    for r in range(meta_blank_row + 1, sheaths_row):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        # If we hit an empty row (blank separator above SHEATHS), stop
        if (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
            break
        # The first row where colA has a value but colB is empty is a connection line
        if val_a is not None and (val_b is None or str(val_b).strip() == ""):
            first_connection_row = r
            break

    # If no connection lines were found (all lines have colB filled), then no direction lines to reorder or MST to add
    if first_connection_row is None:
        continue

    # Detect presence of 1x2 and 1x32 devices in device section
    found_1x2 = False
    found_1x32 = False
    found_demux = False
    # Device section spans from meta_blank_row+1 up to the row before first_connection_row
    device_start = meta_blank_row + 1
    device_end = first_connection_row - 1
    if device_end < device_start:
        device_end = device_start - 1  # no device lines
    for r in range(device_start, device_end + 1):
        cell_val = sheet.cell(row=r, column=1).value
        cell_val_b = sheet.cell(row=r, column=2).value
        if cell_val is None:
            continue
        text = str(cell_val).upper()
        text_b = str(cell_val_b).upper() if cell_val_b else ""
        if "1X32" in text or "1X32" in text_b:
            found_1x32 = True
        if "1X2" in text or "1X2" in text_b:
            found_1x2 = True
        if "DEMUX" in text:
            found_demux = True

    # Insert 'DEMUX 4CH' line if needed
    if found_1x2 and found_1x32 and not found_demux:
        # Insert a new row at the first connection line position to add a new device line before connections
        sheet.insert_rows(first_connection_row)
        # Populate the new row with "DEMUX 4CH"
        sheet.cell(row=first_connection_row, column=1).value = "DEMUX"
        sheet.cell(row=first_connection_row, column=2).value = "4CH"
        # Ensure other cells in that row (columns 3-5) are blank
        for col in range(3, 6):
            sheet.cell(row=first_connection_row, column=col).value = None
        # After insertion, adjust device_end and first_connection_row pointers
        device_end += 1
        first_connection_row += 1
        if sheaths_row:
            sheaths_row += 1  # "SHEATHS" shifts down by one
    # Recompute blank-above-sheaths row (now that we might have inserted a DEMUX row)
    blank_above_sheaths = sheaths_row - 1 if sheaths_row else None

    # Check for existing MST line in connection section
    has_mst_line = False
    connection_end = blank_above_sheaths if blank_above_sheaths else sheaths_row
    # (Connection lines run from first_connection_row through the row before the blank-above-sheaths)
    for r in range(first_connection_row, connection_end):
        val = sheet.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip().upper() == "MST":
            has_mst_line = True
            break

    # Insert 'MST 24CT' connection line if needed
    if found_1x2 and found_1x32 and not has_mst_line:
        # Determine the blank separator row above SHEATHS (should exist)
        if blank_above_sheaths is None:
            # Find blank row immediately above SHEATHS if not tracked
            blank_above_sheaths = sheaths_row - 1
            # If that row isn't actually blank, insert one to be safe
            val_a = sheet.cell(row=blank_above_sheaths, column=1).value
            val_b = sheet.cell(row=blank_above_sheaths, column=2).value
            if (val_a is not None and str(val_a).strip() != "") or (val_b is not None and str(val_b).strip() != ""):
                # Insert a blank row above SHEATHS
                sheet.insert_rows(sheaths_row)
                blank_above_sheaths = sheaths_row
                sheaths_row += 1
        # Insert MST line at the blank separator's position (this pushes the blank and SHEATHS down)
        sheet.insert_rows(blank_above_sheaths)
        # Place MST 24CT in the new row
        sheet.cell(row=blank_above_sheaths, column=1).value = "MST"
        sheet.cell(row=blank_above_sheaths, column=2).value = None
        sheet.cell(row=blank_above_sheaths, column=3).value = None
        sheet.cell(row=blank_above_sheaths, column=4).value = "24CT"
        sheet.cell(row=blank_above_sheaths, column=5).value = None
        # After insertion, adjust first_connection_row, sheaths_row
        if sheaths_row:
            sheaths_row += 1
        # Recompute blank_above_sheaths (now this will be the original blank, moved down one)
        blank_above_sheaths = blank_above_sheaths + 1
    # Reorder connection direction lines (DISABLED)
    #
    # Step 5 (5_format_top_section.py) is responsible for creating the
    # direction color bars in the correct order with the correct fills and
    # labels (N, E, W, S, Splitters, Muxing, etc.).
    #
    # This Step 11 script used to reorder ONLY the VALUES in the connection
    # section. Because Excel styles/fills are NOT moved when only values are
    # reassigned, that logic can desynchronize labels from bar colors (e.g.,
    # a green bar labeled 'West').
    #
    # So we intentionally do NOT reorder this section here. If ordering needs
    # to change, do it in Step 5 where styles and values are written together.


# Save the modified workbook to a new file
wb.save(OUTPUT_FILE)

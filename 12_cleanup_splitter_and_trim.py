#!/usr/bin/env python3
"""Step 12: Cleanup / trim columns, and fix splitter sheets.

Goals (as implemented):

1) Identify the main SHEATHS table by finding the header row that contains
   the header value "CONNECTION" (case-insensitive).

2) For any sheet, delete the *right-side* "SHEATH UUID" column (the one whose
   header is "SHEATH UUID" and whose column index is **to the right of** the
   CONNECTION column) and everything to the right of it.

3) On splitter sheets (detected by the presence of an "OPTICAL SPLITTERS"
   section): before deleting columns, shift the per-splitter metadata that
   currently sits to the far right of CONNECTION (e.g., PORT NAME / PORT
   WAVELENGTH / DEVICE NAME) leftward so that it starts immediately to the
   right of the BUFFER column *that is to the right of CONNECTION*.
   - We only shift rows where DEVICE NAME contains "1x2" or "1x32".
   - We never move/copy DEVICE UUID.

4) In the OPTICAL SPLITTERS section itself, clear the "SHEATH UUID" column
   (header + values + fill) without deleting the whole worksheet column.

This script operates on output/Asbuilt_Workbook.xlsx and writes:
output/Asbuilt_Workbook_post12.xlsx
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill


INPUT_XLSX = Path("output") / "Asbuilt_Workbook.xlsx"
OUTPUT_XLSX = Path("output") / "Asbuilt_Workbook_post12.xlsx"


def norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().upper()


def is_location_sheet(name: str) -> bool:
    n = name.strip().lower()
    if n in {"index", "legend", "notes", "sheet1"}:
        return False
    if n.startswith("tap report"):
        return False
    return True


def find_header_row(ws, max_scan_rows: int = 120) -> Optional[int]:
    """Find the header row of the main table by locating 'CONNECTION'."""
    max_r = min(ws.max_row, max_scan_rows)
    max_c = min(ws.max_column, 120)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if norm(ws.cell(r, c).value) == "CONNECTION":
                return r
    return None


def find_col_in_row(ws, row: int, label: str, *, min_col: int = 1, max_col: Optional[int] = None, gt_col: Optional[int] = None) -> Optional[int]:
    """Find a column in a row matching label. If gt_col is set, only consider cols > gt_col."""
    target = label.strip().upper()
    if max_col is None:
        max_col = ws.max_column
    for c in range(min_col, max_col + 1):
        if gt_col is not None and c <= gt_col:
            continue
        if norm(ws.cell(row, c).value) == target:
            return c
    return None


def find_all_cols_in_row(ws, row: int, label: str, *, max_col: Optional[int] = None) -> list[int]:
    target = label.strip().upper()
    if max_col is None:
        max_col = ws.max_column
    cols = []
    for c in range(1, max_col + 1):
        if norm(ws.cell(row, c).value) == target:
            cols.append(c)
    return cols


def find_cell_contains(ws, text: str, max_scan_rows: int = 400) -> Optional[Tuple[int, int]]:
    target = text.strip().upper()
    max_r = min(ws.max_row, max_scan_rows)
    max_c = min(ws.max_column, 120)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and target in v.strip().upper():
                return (r, c)
    return None


def top_table_last_row(ws, header_row: int, primary_col: int) -> int:
    """Find last row of the main table by scanning down primary_col until blank."""
    r = header_row + 1
    # stop at first long blank run
    blank_run = 0
    last_nonblank = header_row
    while r <= ws.max_row:
        v = ws.cell(r, primary_col).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            blank_run += 1
            if blank_run >= 5:
                break
        else:
            blank_run = 0
            last_nonblank = r
        r += 1
    return last_nonblank


def clear_range_values_and_fills(ws, row_min: int, row_max: int, col_min: int, col_max: int) -> None:
    empty_fill = PatternFill()  # default / no fill
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = empty_fill


def clear_optical_splitter_sheath_uuid(ws) -> bool:
    """Clear SHEATH UUID column in the OPTICAL SPLITTERS section (header + values + fill)."""
    found = find_cell_contains(ws, "OPTICAL SPLITTERS")
    if not found:
        return False

    opt_row, _ = found

    # Find the splitter sub-table header row that contains 'SHEATH UUID'
    sheath_uuid_pos = None
    scan_to = min(ws.max_row, opt_row + 40)
    max_c = min(ws.max_column, 160)
    for r in range(opt_row, scan_to + 1):
        for c in range(1, max_c + 1):
            if norm(ws.cell(r, c).value) == "SHEATH UUID":
                sheath_uuid_pos = (r, c)
                break
        if sheath_uuid_pos:
            break

    if not sheath_uuid_pos:
        return False

    hdr_row, sheath_uuid_col = sheath_uuid_pos

    # Clear SHEATH UUID col and any contiguous blank-header cols to its right
    last_clear_col = sheath_uuid_col
    for c in range(sheath_uuid_col + 1, max_c + 1):
        v = ws.cell(hdr_row, c).value
        if v is not None and str(v).strip() != "":
            break
        last_clear_col = c

    clear_range_values_and_fills(ws, hdr_row, ws.max_row, sheath_uuid_col, last_clear_col)
    return True


def sheet_has_splitter(ws) -> bool:
    return find_cell_contains(ws, "OPTICAL SPLITTERS") is not None


def process_sheet(ws, name: str) -> Tuple[bool, bool, bool]:
    """Returns (shifted_any, trimmed_any, cleared_optical_uuid)."""

    header_row = find_header_row(ws)
    if header_row is None:
        return (False, False, False)

    connection_col = find_col_in_row(ws, header_row, "CONNECTION")
    if connection_col is None:
        return (False, False, False)

    # Find BUFFER column to the right of CONNECTION (destination anchor)
    buffer_cols = find_all_cols_in_row(ws, header_row, "BUFFER")
    buffer_after_conn = None
    for c in buffer_cols:
        if c > connection_col:
            buffer_after_conn = c
            break
    if buffer_after_conn is None:
        # Can't safely shift without destination anchor, but we can still trim.
        buffer_after_conn = connection_col

    dest_start_col = buffer_after_conn + 1

    # Determine the *right-side* SHEATH UUID column to trim from: must be > CONNECTION
    sheath_uuid_cols = find_all_cols_in_row(ws, header_row, "SHEATH UUID")
    trim_col = None
    for c in sheath_uuid_cols:
        if c > connection_col:
            trim_col = c
            break

    # Identify the right-side DEVICE NAME / DEVICE UUID columns (must be > CONNECTION)
    device_name_col = find_col_in_row(ws, header_row, "DEVICE NAME", gt_col=connection_col)
    device_uuid_col = find_col_in_row(ws, header_row, "DEVICE UUID", gt_col=connection_col)

    # Build a candidate list of columns to shift (those between CONNECTION and DEVICE UUID, excluding base columns)
    shifted_any = False
    if sheet_has_splitter(ws) and device_name_col and device_uuid_col and (trim_col is None or dest_start_col < trim_col):
        keep_headers = {
            "FIBER",
            "BUFFER",
            "END ENCLOSURE",
            "END ENCLOSURE ",
            "START ENCLOSURE",
            "SHEATH NAME",
            "SHEATH UUID",
        }

        right_start_cols = []
        for c in range(connection_col + 1, device_uuid_col):
            h = norm(ws.cell(header_row, c).value)
            if h and h not in keep_headers:
                right_start_cols.append(c)

        # Determine primary column for the main table end (prefer a SHEATH UUID-like UUID column at left; fallback col 1)
        primary_col = 1
        # If there is a SHEATH UUID col on the left side (<= CONNECTION), use that for end detection
        for c in sheath_uuid_cols:
            if c <= connection_col:
                primary_col = c
                break

        last_row = top_table_last_row(ws, header_row, primary_col)

        # Shift only rows where DEVICE NAME includes 1X2 / 1X32
        for r in range(header_row + 1, last_row + 1):
            dname = ws.cell(r, device_name_col).value
            if not isinstance(dname, str):
                continue
            dname_u = dname.upper()
            if "1X2" not in dname_u and "1X32" not in dname_u:
                continue

            # Find first non-empty cell among right_start_cols
            src_start = None
            for c in right_start_cols:
                v = ws.cell(r, c).value
                if v is not None and (not isinstance(v, str) or v.strip() != ""):
                    src_start = c
                    break
            if src_start is None:
                continue

            src_end = device_uuid_col - 1
            width = src_end - src_start + 1

            # Copy block to destination, same spacing
            for i in range(width):
                src_c = src_start + i
                dst_c = dest_start_col + i
                # Avoid writing beyond trim_col if trimming will occur
                if trim_col and dst_c >= trim_col:
                    break
                ws.cell(r, dst_c).value = ws.cell(r, src_c).value

            # Clear original block
            for c in range(src_start, src_end + 1):
                ws.cell(r, c).value = None

            shifted_any = True

    # Clear optical splitter section SHEATH UUID (header + values + fill) on splitter sheets
    cleared_optical = False
    if sheet_has_splitter(ws):
        cleared_optical = clear_optical_splitter_sheath_uuid(ws)

    # Trim worksheet columns from the right-side SHEATH UUID onward
    trimmed_any = False
    if trim_col:
        # delete from trim_col to end
        ws.delete_cols(trim_col, ws.max_column - trim_col + 1)
        trimmed_any = True

    return (shifted_any, trimmed_any, cleared_optical)


def main() -> None:
    if not INPUT_XLSX.exists():
        raise SystemExit(f"Missing input file: {INPUT_XLSX}")

    wb = openpyxl.load_workbook(INPUT_XLSX)

    total_shift = 0
    total_trim = 0
    total_clear = 0

    for name in wb.sheetnames:
        if not is_location_sheet(name):
            continue
        ws = wb[name]
        shifted, trimmed, cleared = process_sheet(ws, name)
        total_shift += int(shifted)
        total_trim += int(trimmed)
        total_clear += int(cleared)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    print(f"\n✅ Step 12 complete: {OUTPUT_XLSX}")
    print(f"   • sheets shifted: {total_shift}")
    print(f"   • sheets trimmed : {total_trim}")
    print(f"   • optical uuid cleared: {total_clear}")


if __name__ == "__main__":
    main()

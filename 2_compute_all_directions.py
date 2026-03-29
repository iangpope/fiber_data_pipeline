"""
2_compute_all_directions.py -- Compute cable bearing directions and assign colors.

Reads the Connections Table (built by step 1) and the KMZ geometry to determine
which compass direction (North/East/South/West) each fiber cable travels when
leaving a splice enclosure. The result is an Excel workbook where each cable's
cell is filled with a directional color. This file is the manual checkpoint:
review it to verify direction assignments before running steps 3-8.

Special cases:
  - OLT cables (one endpoint is the bare site token, e.g. RC73E) -> olive green
  - MST tap cables (single-connection FT tap nearest to its SE enclosure) -> red

Reads:  data/Connections_Table.xlsx
        data/<project>.kmz
Writes: output/Colored_Connections_Table.xlsx
"""

import os
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from geopy.distance import geodesic
from math import atan2, degrees, radians, sin, cos
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from naming_utils import parse_location_id


# ---------------------------------------------------------------------------
# Regex patterns for parsing cable names and location tokens.
#
# New-format cable names embed location tokens like "RC73E_FT_032" or
# "RC73E_SE_006". Legacy format uses "MIC..." prefixed tokens. Both styles
# appear in older projects so both patterns are needed.
# ---------------------------------------------------------------------------

# Matches a location token in new naming convention (e.g. RC73E_FT_001)
NEW_LOC_TOKEN_RX    = re.compile(r"[A-Z0-9]+E_(?:FT|SE)_\d{3}")

# Matches a location token in legacy naming convention (e.g. MICRCTS001)
LEGACY_LOC_TOKEN_RX = re.compile(r"\bMIC\w+\b")

# Matches a bare OLT site token (e.g. RC73E, MS90E) -- 2 letters, 2-3 digits, E
OLT_TOKEN_RX        = re.compile(r"^[A-Z]{2}\d{2,3}E$")

# Matches a fiber count suffix in a cable name (e.g. _48CT, _096CT)
FIBER_SUFFIX_RX     = re.compile(r"_(\d{2,3}CT)\b", flags=re.IGNORECASE)


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------


def detect_kmz_file(data_dir: str = "data") -> str:
    """Return the path to the one KMZ file in the data folder, or raise."""
    for f in os.listdir(data_dir):
        if f.lower().endswith(".kmz"):
            return os.path.join(data_dir, f)
    raise FileNotFoundError(f"No .kmz file found in '{data_dir}' folder.")


# ---------------------------------------------------------------------------
# Cable name parsing helpers
# ---------------------------------------------------------------------------

def parse_endpoints(cable_name: str):
    """
    Extract (endpointA, endpointB, fiber_count) from a cable name string.

    Supports two naming formats:
      - New format: "RC73E_FT_001_TO_RC73E_SE_001_48CT"  (underscore-delimited)
      - Old format: "48CT RC73E_FT_001 TO RC73E_SE_001"  (space-delimited with fiber first)

    Returns (None, None, None) if the name does not match either pattern.
    """
    t = str(cable_name or "").strip()
    if not t:
        return None, None, None

    # New format: split on "_TO_"
    if "_TO_" in t:
        a, b = t.split("_TO_", 1)
        fiber = None
        m = re.search(r"_(\d{2,3}CT)\b", b, flags=re.IGNORECASE)
        if m:
            fiber = m.group(1).upper()
            b = re.sub(r"_(\d{2,3}CT)\b", "", b, flags=re.IGNORECASE)
        return a.strip(), b.strip(), fiber

    # Old format: "<fiber> <endA> TO <endB>"
    m = re.match(r"^\s*(\d{2,3}CT)\s+(\S+)\s+TO\s+(\S+)\s*$", t, flags=re.IGNORECASE)
    if m:
        return m.group(2).strip(), m.group(3).strip(), m.group(1).upper()

    return None, None, None


def is_olt_cable(cable_name: str, olt_token: str) -> bool:
    """
    Return True if either endpoint of the cable is the bare OLT site token.

    The OLT token is typically the short site identifier (e.g. "RC73E") that
    identifies the head-end optical line terminal rack. Cables connected to
    the OLT receive a distinct color to make the OLT connection obvious.
    """
    a, b, _ = parse_endpoints(cable_name)
    if not olt_token:
        return False
    return (a == olt_token) or (b == olt_token)


# ---------------------------------------------------------------------------
# KMZ parsing: extract all fiber cable line segments
# ---------------------------------------------------------------------------

def extract_fiber_segments(kmz_path: str) -> list:
    """
    Parse the KMZ and return a list of (cable_name, [(lat, lon), ...]) tuples.

    Each fiber cable in the KMZ is a LineString Placemark. This function
    iterates all placemarks in the document, skipping point placemarks (which
    represent splice enclosure locations, not cables), and collects the name
    and coordinate sequence of every line placemark.

    The KML namespace is stripped so the parser works regardless of whether
    the KML was exported with or without the opengis namespace prefix.
    """
    def strip_ns(tag: str) -> str:
        """Remove the XML namespace prefix from a tag string."""
        return tag.split("}")[-1] if "}" in tag else tag

    with zipfile.ZipFile(kmz_path, "r") as zf:
        kml_file = next((f for f in zf.namelist() if f.endswith(".kml")), None)
        root = ET.fromstring(zf.read(kml_file))

    def recursive_find(el) -> list:
        """Walk the element tree and collect all LineString placemarks."""
        segments = []
        for child in el.iter():
            tag = strip_ns(child.tag.lower())
            if tag != "placemark":
                continue
            coords_elem = child.find(".//{*}LineString/{*}coordinates")
            if coords_elem is None:
                continue   # skip point placemarks (splice enclosure pins)
            # Parse coordinate text: space-separated "lon,lat[,alt]" tokens
            coords  = coords_elem.text.strip().split()
            latlons = [
                (float(c.split(",")[1]), float(c.split(",")[0]))
                for c in coords if "," in c
            ]
            name_elem = next(
                (c for c in child if strip_ns(c.tag) == "name"), None
            )
            name = (
                name_elem.text.strip()
                if (name_elem is not None and name_elem.text)
                else None
            )
            segments.append((name, latlons))
        return segments

    segments = recursive_find(root)
    print(f"Extracted {len(segments)} line segments from KMZ")
    return segments


def merge_named_segments(segments: list) -> list:
    """
    Merge multi-part cable segments that share the same name into single chains.

    GIS exports can split a single cable into multiple LineString segments
    (e.g. when it crosses a tile boundary). This function groups all segments
    by cable name and chains them together end-to-end using a deque. If
    segments cannot be joined geometrically (no shared endpoints), they are
    appended in order to prevent data loss.

    Returns a list of (cable_name, [lat_lon, ...]) with one entry per cable.
    """
    from collections import defaultdict, deque

    # Group all coordinate sequences by cable name.
    grouped = defaultdict(list)
    for name, coords in segments:
        if name:
            grouped[name].append(coords)

    merged = []
    for name, paths in grouped.items():
        if len(paths) == 1:
            # Only one segment for this cable; no merging needed.
            merged.append((name, paths[0]))
            continue

        # Build a chain by connecting segments end-to-end.
        chain = deque(paths[0])
        remaining = list(paths[1:])

        while remaining:
            progress = False
            still_remaining = []
            for p in remaining:
                if chain[-1] == p[0]:           # tail of chain matches head of segment
                    chain.extend(p[1:])
                    progress = True
                elif chain[-1] == p[-1]:         # tail matches tail; append reversed
                    chain.extend(reversed(p[:-1]))
                    progress = True
                elif chain[0] == p[-1]:          # head of chain matches tail; prepend
                    chain.extendleft(reversed(p[:-1]))
                    progress = True
                elif chain[0] == p[0]:           # head matches head; prepend reversed
                    chain.extendleft(p[1:])
                    progress = True
                else:
                    still_remaining.append(p)
            if not progress:
                # No geometric connection found; append remaining segments anyway.
                for p in still_remaining:
                    chain.extend(p)
                break
            remaining = still_remaining

        merged.append((name, list(chain)))

    return merged


# ---------------------------------------------------------------------------
# Bearing calculation and color assignment
# ---------------------------------------------------------------------------

def bearing(p1: tuple, p2: tuple) -> float:
    """
    Compute the geodesic forward bearing from p1 to p2 in degrees (0-360).

    Uses the standard spherical haversine bearing formula. Both points are
    (lat, lon) tuples in decimal degrees. 0 = North, 90 = East, etc.
    """
    lat1, lon1 = radians(p1[0]), radians(p1[1])
    lat2, lon2 = radians(p2[0]), radians(p2[1])
    dlon = lon2 - lon1
    x = sin(dlon) * cos(lat2)
    y = cos(lat1) * sin(lat2) - sin(lat1) * cos(lat2) * cos(dlon)
    return (degrees(atan2(x, y)) + 360) % 360


def bearing_to_color(b: float) -> str:
    """
    Map a bearing angle (degrees) to one of four directional color names.

    Quadrant boundaries are set at 45-degree diagonals:
      North (orange) : 337.5 - 360 or 0 - 22.5
      East  (green)  : 67.5 - 112.5
      South (brown)  : 157.5 - 202.5
      West  (slate)  : 247.5 - 292.5

    Bearings that fall in the diagonal zones (between quadrants) are assigned
    to the nearest principal direction to avoid ambiguity.
    """
    if b < 22.5 or b >= 337.5:
        return "orange"   # North
    elif 67.5 <= b < 112.5:
        return "green"    # East
    elif 157.5 <= b < 202.5:
        return "brown"    # South
    elif 247.5 <= b < 292.5:
        return "slate"    # West

    # Diagonal zone: pick nearest cardinal direction by angular distance.
    closest = min(
        {"orange": 0, "green": 90, "brown": 180, "slate": 270}.items(),
        key=lambda item: abs((b - item[1] + 180) % 360 - 180),
    )
    return closest[0]


# ---------------------------------------------------------------------------
# Main processing: build and color the connections table
# ---------------------------------------------------------------------------

def is_diagonal_bearing(b: float) -> bool:
    """
    Return True if bearing b falls in a diagonal zone (between cardinal quadrants).

    The four clean quadrants are:
      North: [337.5, 360) or [0, 22.5)
      East:  [67.5, 112.5)
      South: [157.5, 202.5)
      West:  [247.5, 292.5)

    Any bearing outside these ranges is in a diagonal zone and the direction
    assignment is less reliable.
    """
    return not (
        b < 22.5 or b >= 337.5
        or 67.5 <= b < 112.5
        or 157.5 <= b < 202.5
        or 247.5 <= b < 292.5
    )


def _optimal_direction_assignment(cable_bearings: list) -> dict:
    """
    Assign cardinal direction colors to regular cables at one location so that
    the total angular deviation from assigned directions is minimised, with no
    two cables sharing the same color when four or fewer cables are present.

    cable_bearings : list of (col, bearing_degrees) tuples (OLT/MST excluded).
    Returns        : dict { col -> color_name }.
    """
    from itertools import combinations, permutations as _perms

    DIRS     = {"orange": 0, "green": 90, "brown": 180, "slate": 270}
    dir_list = list(DIRS.keys())

    def dev(b, color):
        angle = DIRS[color]
        return abs((b - angle + 180) % 360 - 180)

    n = len(cable_bearings)
    if n == 0:
        return {}
    if n == 1:
        col, b = cable_bearings[0]
        return {col: min(dir_list, key=lambda c: dev(b, c))}

    if n <= 4:
        # Exhaustive search over all ways to assign n distinct colors to n cables.
        # Max iterations: C(4,4)*4! = 24 — negligible cost.
        best_cost, best = float("inf"), {}
        for color_subset in combinations(dir_list, n):
            for perm in _perms(color_subset):
                cost = sum(dev(cable_bearings[i][1], perm[i]) for i in range(n))
                if cost < best_cost:
                    best_cost = cost
                    best = {cable_bearings[i][0]: perm[i] for i in range(n)}
        return best

    # More than 4 cables: assign one cable per direction first (minimum-deviation
    # greedy pairing), then assign extras to their nearest direction (repeats allowed).
    result, used_dirs = {}, set()
    for _ in range(4):
        best_cost, best_col, best_color = float("inf"), None, None
        for col, b in cable_bearings:
            if col in result:
                continue
            for color in dir_list:
                if color in used_dirs:
                    continue
                c = dev(b, color)
                if c < best_cost:
                    best_cost = c; best_col = col; best_color = color
        if best_col:
            result[best_col] = best_color
            used_dirs.add(best_color)
    for col, b in cable_bearings:
        if col not in result:
            result[col] = min(dir_list, key=lambda c: dev(b, c))
    return result


def generate_colored_table(named_edges: list, location_coords: dict,
                           connections_path: str, output_path: str) -> None:
    """
    Build the Colored Connections Table Excel file.

    For each location row in the Connections Table, each cable is assigned a
    directional color based on the bearing measured 5 m along the cable from
    the splice point. The 5 m walk avoids the ambiguity caused by short jogs
    near enclosures, giving a stable bearing representative of the cable's
    general routing direction.

    Special override rules (applied before bearing calculation):
      - OLT cables are always colored olive green.
      - MST tap cables (single-connection FT nearest its paired SE) are red.
      - Within one location, each directional color may only appear once;
        a fallback to the next-nearest color is used if a direction repeats.
    """
    df_main = pd.read_excel(connections_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Colored Connections"
    ws.append(df_main.columns.tolist())   # write header row

    # PatternFill objects for each directional and special-case color.
    fills = {
        "orange": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "brown":  PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid"),
        "green":  PatternFill(start_color="008000", end_color="008000", fill_type="solid"),
        "slate":  PatternFill(start_color="708090", end_color="708090", fill_type="solid"),
        "red":    PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "olt":    PatternFill(start_color="C5D9B5", end_color="C5D9B5", fill_type="solid"),
    }

    # Build a lookup from cable name to its coordinate list.
    name_to_coords = {name: coords for name, coords in named_edges}
    red_cables     = set()   # cable names confirmed as MST (colored red)
    all_colored_rows = []    # collect (entries, conn_cols, raw_colors) for two-pass output
    cable_confidence: dict[str, dict] = {}  # cable -> {bearing, confidence} for map overlay

    # Debug/confidence tracking: populated during the main loop and printed at the end.
    dbg_missing_cables  = []   # (location, cable_name) -- cable not found in KMZ
    dbg_short_cables    = []   # (location, cable_name, length_m) -- cable shorter than 5 m
    dbg_diagonal_cables = []   # (location, cable_name, bearing_deg, assigned_color)

    # ------------------------------------------------------------------
    # Pre-scan pass: identify MST tap locations.
    #
    # An MST tap is an FT enclosure whose only cable runs to a nearby SE
    # enclosure (< 50 m away). When multiple FT taps are close to the same
    # SE (which happens frequently in dense areas), only the single nearest
    # FT is the true MST tap. The rest are normal taps colored by direction.
    #
    # This pre-scan builds a candidate list per SE and picks the winner
    # before the main coloring loop runs, so the decision is consistent.
    # ------------------------------------------------------------------
    mst_winner_locs: set[str] = set()    # location names confirmed as MST taps
    se_candidates:  dict[str, list] = {} # SE enclosure -> [(distance_m, ft_location)]

    for _, row in df_main.iterrows():
        loc        = str(row["Location"]).strip()
        latlon     = (row["Latitude"], row["Longitude"])
        conn_cols  = [c for c in row.index if str(c).startswith("Connection")]
        valid_conn = [row[c] for c in conn_cols if not pd.isna(row[c])]

        # Only single-connection locations can be MST taps.
        if len(valid_conn) != 1:
            continue

        cable   = valid_conn[0]
        matches = NEW_LOC_TOKEN_RX.findall(cable) or LEGACY_LOC_TOKEN_RX.findall(cable)
        if len(matches) != 2:
            continue    # need exactly two endpoints to know which side is SE

        end1, end2 = matches
        other = end1 if loc != end1 else end2

        # The "other" endpoint must be an SE or SD type enclosure.
        is_se_type = re.search(r"[SD]\d{3}$", other) or re.search(r"_SE_\d{3}$", other)
        if not is_se_type:
            continue

        other_coord = location_coords.get(other)
        if not other_coord:
            continue

        dist_m = geodesic(latlon, other_coord).meters
        if dist_m < 50:
            # Collect this FT tap as a candidate for the SE enclosure.
            se_candidates.setdefault(other, []).append((dist_m, loc))

    # For each SE, the closest FT tap wins the MST designation.
    for se, candidates in se_candidates.items():
        winner = min(candidates, key=lambda t: t[0])[1]
        mst_winner_locs.add(winner)

    # ------------------------------------------------------------------
    # Identify the OLT location and its primary outbound cable.
    # The OLT is the bare site token row (e.g. "RC73E") with no type suffix.
    # ------------------------------------------------------------------
    olt_pattern = re.compile(r"^[A-Z]{2}\d{2,3}E$")
    olt_loc  = None
    olt_conn = None

    for _, row in df_main.iterrows():
        loc = str(row["Location"]).strip()
        if olt_pattern.match(loc):
            olt_loc   = loc
            conn_cols = [c for c in row.index if str(c).startswith("Connection")]
            for c in conn_cols:
                if pd.notna(row[c]):
                    olt_conn = row[c].strip()
                    break
            break

    # ------------------------------------------------------------------
    # Main coloring loop: determine the direction color for each cable
    # at each location.
    # ------------------------------------------------------------------
    for _, row in df_main.iterrows():
        loc       = row["Location"]
        latlon    = (row["Latitude"], row["Longitude"])
        conn_cols = [c for c in row.index if str(c).startswith("Connection")]
        entries   = [loc, latlon[0], latlon[1]]   # start each row with location + coords
        raw       = []                            # (col, color_name, bearing) for this row
        valid_conn = [row[c] for c in conn_cols if not pd.isna(row[c])]

        for col in conn_cols:
            cable = row[col]
            if pd.isna(cable):
                entries.append("")
                continue

            # -- OLT override: color any OLT cable olive green --
            if olt_loc and is_olt_cable(cable, olt_loc):
                raw.append((col, "olt", 0))
                entries.append(cable)
                cable_confidence.setdefault(str(cable), {"bearing": None, "confidence": "ok"})
                continue

            # -- Look up the cable's coordinate geometry from the KMZ --
            coords = name_to_coords.get(cable)
            if not coords:
                dbg_missing_cables.append((str(loc), str(cable)))
                entries.append(cable)
                cable_confidence[str(cable)] = {"bearing": None, "confidence": "missing"}
                continue

            # Ensure coordinate sequence starts at the current splice location.
            # Use geodesic distance instead of exact float equality to handle
            # floating point imprecision in KMZ coordinates.
            start_dist = geodesic(coords[0], latlon).meters
            end_dist   = geodesic(coords[-1], latlon).meters
            if end_dist < start_dist:
                coords = list(reversed(coords))

            # -- MST override: color confirmed MST tap cables red --
            matches = NEW_LOC_TOKEN_RX.findall(cable) or LEGACY_LOC_TOKEN_RX.findall(cable)
            if len(valid_conn) == 1 and len(matches) == 2:
                end1, end2 = matches
                other = end1 if loc != end1 else end2
                is_se_type = (
                    re.search(r"[SD]\d{3}$", other) or
                    re.search(r"_SE_\d{3}$", other)
                )
                if is_se_type and loc in mst_winner_locs:
                    raw.append((col, "red", 0))
                    red_cables.add(cable)
                    entries.append(cable)
                    cable_confidence.setdefault(str(cable), {"bearing": None, "confidence": "ok"})
                    continue

            # -- Compute bearing via KMZ geometry walk --
            # Walk 25 m along the cable from this enclosure to determine the
            # direction it exits.  This is more accurate than a straight
            # splice-to-splice bearing because it follows the physical route
            # (including initial bends near the enclosure) rather than the
            # straight line between endpoints.
            _WALK_DIST = 25  # metres — long enough to clear GPS jitter

            is_short = False
            dist     = 0
            walked   = latlon
            for i in range(len(coords) - 1):
                seg_len = geodesic(coords[i], coords[i + 1]).meters
                if dist + seg_len >= _WALK_DIST:
                    ratio  = (_WALK_DIST - dist) / seg_len
                    lat_w  = coords[i][0] + (coords[i + 1][0] - coords[i][0]) * ratio
                    lon_w  = coords[i][1] + (coords[i + 1][1] - coords[i][1]) * ratio
                    walked = (lat_w, lon_w)
                    break
                dist += seg_len

            if walked == latlon:
                # Cable shorter than walk distance — use the far endpoint instead.
                is_short  = True
                cable_len = sum(
                    geodesic(coords[i], coords[i + 1]).meters
                    for i in range(len(coords) - 1)
                )
                dbg_short_cables.append((str(loc), str(cable), round(cable_len, 2)))
                walked = coords[-1]

            b = bearing(latlon, walked)
            color = bearing_to_color(b)
            conf  = "short" if is_short else ("diagonal" if is_diagonal_bearing(b) else "ok")
            if is_diagonal_bearing(b):
                dbg_diagonal_cables.append((str(loc), str(cable), round(b, 1), color))
            raw.append((col, color, b))
            entries.append(cable)
            _prio = {"missing": 3, "short": 2, "diagonal": 1, "ok": 0}
            _prev = cable_confidence.get(str(cable), {})
            if _prio.get(conf, 0) >= _prio.get(_prev.get("confidence", "ok"), 0):
                cable_confidence[str(cable)] = {"bearing": round(b, 1), "confidence": conf}

        all_colored_rows.append((entries, conn_cols, raw))

    # ------------------------------------------------------------------
    # Output pass: write rows to the workbook and apply cell fills.
    #
    # OLT (olive) and MST (red) cables are given fixed colors. The remaining
    # cables at each location are assigned cardinal directions using
    # _optimal_direction_assignment, which minimises total angular deviation
    # and guarantees each cable gets a unique direction when ≤ 4 are present.
    # ------------------------------------------------------------------
    for entries, conn_cols, raw in all_colored_rows:
        ws.append(entries)
        assigned = {}   # col -> color_name for final fill application

        # OLT and MST/red cables have fixed colors — assign them first.
        # Collect the remaining cables for optimal cardinal-direction assignment.
        regular = []
        for col, colr, b in raw:
            if colr == "olt":
                assigned[col] = "olt"
            elif entries[3 + conn_cols.index(col)] in red_cables:
                assigned[col] = "red"
            else:
                regular.append((col, b))

        # Optimally assign N / E / S / W to minimise total angular deviation.
        # When ≤ 4 regular cables are present each gets a unique direction;
        # beyond 4 the optimal 4 are assigned first, extras get their nearest.
        assigned.update(_optimal_direction_assignment(regular))

        # Apply fills to the connection columns (columns 4 onward in the sheet).
        loc_current = entries[0]
        for idx, col in enumerate(conn_cols):
            cable = entries[3 + idx]
            if col in assigned and assigned[col] == "olt":
                ws.cell(row=ws.max_row, column=4 + idx).fill = fills["olt"]
            elif olt_conn and cable == olt_conn:
                # Secondary OLT cable match (cable string matches but wasn't
                # caught by is_olt_cable due to name variation).
                ws.cell(row=ws.max_row, column=4 + idx).fill = fills["olt"]
            elif col in assigned:
                ws.cell(row=ws.max_row, column=4 + idx).fill = fills[assigned[col]]

    # ------------------------------------------------------------------
    # Debug / confidence log: summarise rows that need manual scrutiny.
    # ------------------------------------------------------------------
    print("\n=== DEBUG / CONFIDENCE LOG ===")

    if dbg_missing_cables:
        print(f"\n[MISSING FROM KMZ] {len(dbg_missing_cables)} cable(s) not found "
              f"— no geometry, no color assigned:")
        for loc_name, cbl in dbg_missing_cables:
            print(f"  {loc_name:30s}  {cbl}")
    else:
        print("\n[MISSING FROM KMZ] none")

    if dbg_short_cables:
        print(f"\n[SHORT CABLES < 5 m] {len(dbg_short_cables)} cable(s) — "
              f"bearing computed from far endpoint, may be unreliable:")
        for loc_name, cbl, length in dbg_short_cables:
            print(f"  {loc_name:30s}  {cbl}  ({length} m)")
    else:
        print("\n[SHORT CABLES < 5 m] none")

    if dbg_diagonal_cables:
        print(f"\n[DIAGONAL BEARING] {len(dbg_diagonal_cables)} cable(s) — "
              f"bearing falls between cardinal quadrants, direction less certain:")
        for loc_name, cbl, b_val, color in dbg_diagonal_cables:
            print(f"  {loc_name:30s}  {cbl}  bearing={b_val}°  assigned={color}")
    else:
        print("\n[DIAGONAL BEARING] none")

    print("=== END DEBUG LOG ===\n")

    # Save confidence data for the web UI map overlay (harmless if web UI not used).
    import json as _json
    _conf_path = os.path.join(os.path.dirname(os.path.abspath(output_path)), "direction_confidence.json")
    with open(_conf_path, "w") as _f:
        _json.dump({
            "cables":  cable_confidence,
            "summary": {
                "diagonal": len(dbg_diagonal_cables),
                "short":    len(dbg_short_cables),
                "missing":  len(dbg_missing_cables),
            },
        }, _f)

    wb.save(output_path)
    print(f"Colored table saved to {output_path}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    """
    Run step 2: compute cable bearing directions and write the Colored
    Connections Table.

    Parameters
    ----------
    data_dir : str
        Directory containing Connections_Table.xlsx and the KMZ file.
    output_dir : str
        Directory for the Colored_Connections_Table.xlsx output.
    """
    import os
    os.makedirs(output_dir, exist_ok=True)

    connections_path = os.path.join(data_dir,   "Connections_Table.xlsx")
    output_path      = os.path.join(output_dir, "Colored_Connections_Table.xlsx")
    kmz_path         = detect_kmz_file(data_dir)

    # Extract and merge KMZ cable geometry.
    fiber_segments = extract_fiber_segments(kmz_path)
    fiber_segments = merge_named_segments(fiber_segments)

    # Build location coordinate lookup from the Connections Table.
    df_conn = pd.read_excel(connections_path)
    location_coords = {}
    for _, row in df_conn.iterrows():
        try:
            name = str(row["Location"]).strip()
            lat  = float(row["Latitude"])
            lon  = float(row["Longitude"])
            if name and not pd.isna(lat) and not pd.isna(lon):
                location_coords[name] = (lat, lon)
        except Exception:
            continue

    generate_colored_table(fiber_segments, location_coords,
                           connections_path, output_path)


if __name__ == "__main__":
    main()
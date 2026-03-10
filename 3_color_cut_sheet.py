import os
import openpyxl
from openpyxl.styles import PatternFill
import re

DATA_DIR = "data"
OUTPUT_DIR = "output"
COLORED_CONNECTIONS_PATH = os.path.join(OUTPUT_DIR, "Colored_Connections_Table.xlsx")
CUT_SHEET_PATH = os.path.join(DATA_DIR, "cut_sheet.xlsx")
COLORIZED_CUT_SHEET_PATH = os.path.join(OUTPUT_DIR, "Colorized_Cut_Sheet.xlsx")

# Define fill objects
FILLS = {
    "FFA500": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # orange
    "8B4513": PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid"),  # corrected brown (south)
    "008000": PatternFill(start_color="008000", end_color="008000", fill_type="solid"),  # green
    "708090": PatternFill(start_color="708090", end_color="708090", fill_type="solid"),  # slate
    "FF0000": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # red
    "C5D9B5": PatternFill(start_color="C5D9B5", end_color="C5D9B5", fill_type="solid"),  # OLT (olive green)
    "7FFF00": PatternFill(start_color="7FFF00", end_color="7FFF00", fill_type="solid"),  # puke
    "FFFF00": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # yellow
    "ADD8E6": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # light blue
    "FFB6C1": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # light pink
    "DB7093": PatternFill(start_color="DB7093", end_color="DB7093", fill_type="solid"),  # dark pink
    "FFDAB9": PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid"),  # pale orange
    "FFA07A": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),  # darker orange
}

def load_cable_colors():
    wb = openpyxl.load_workbook(COLORED_CONNECTIONS_PATH)
    ws = wb.active
    cable_map = {}
    for row in ws.iter_rows(min_row=2):
        loc = row[0].value
        for cell in row[3:]:
            if cell.value and cell.fill:
                rgb = cell.fill.start_color.rgb[-6:].upper()
                cable_map.setdefault(loc, {})[cell.value.strip()] = rgb
    return cable_map

def apply_colors_to_cut_sheet():
    color_map = load_cable_colors()
    wb = openpyxl.load_workbook(CUT_SHEET_PATH)

    for sheetname in wb.sheetnames:
        if sheetname == "Index":
            continue

        ws = wb[sheetname]
        conn_dict = color_map.get(sheetname, {})

        rows_to_delete = []
        for r in range(ws.max_row, 6, -1):
            row = [ws.cell(row=r, column=c).value for c in range(1, 21)]
            g_to_i = row[6:9]
            non_g_to_i = row[:6] + row[9:]
            if not any(row) or (any(g_to_i) and not any(non_g_to_i)):
                rows_to_delete.append(r)
        for r in rows_to_delete:
            ws.delete_rows(r)

        # Detect splitter sheets by content: does col A contain "OPTICAL SPLITTERS"?
        is_splitter_or_node = any(
            isinstance(ws.cell(row=r2, column=1).value, str)
            and "OPTICAL SPLITTERS" in ws.cell(row=r2, column=1).value.upper()
            for r2 in range(1, ws.max_row + 1)
        )

        seen_left = set()
        for r in range(7, ws.max_row + 1):
            val_left = ws.cell(row=r, column=2).value
            if isinstance(val_left, str):
                val_left = val_left.strip()
                fill = conn_dict.get(val_left)
                if not fill and re.match(r"\d{3}CT_", val_left):
                    fill = "FF0000"
                if fill and fill in FILLS:
                    for c in range(1, 10):
                        ws.cell(row=r, column=c).fill = FILLS[fill]
                    if val_left not in seen_left:
                        ws.cell(row=r, column=1).fill = FILLS["FFFF00"]
                        ws.cell(row=r, column=2).fill = FILLS["FFFF00"]
                        seen_left.add(val_left)

            val_right = ws.cell(row=r, column=15).value
            if isinstance(val_right, str):
                val_right = val_right.strip()
                fill = conn_dict.get(val_right)
                if not fill and re.match(r"\d{3}CT_", val_right):
                    fill = "FF0000"
                if fill and fill in FILLS:
                    for c in range(11, 21):
                        ws.cell(row=r, column=c).fill = FILLS[fill]

            if str(ws.cell(row=r, column=10).value).strip() == "<- FUSION ->":
                ws.cell(row=r, column=10).fill = FILLS["FFFF00"]

            for c in range(11, 21):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and "PORT" in val.upper() and val.upper() != "PORT NAME":
                    for cc in range(11, 21):
                        ws.cell(row=r, column=cc).fill = FILLS["FFFF00"]

            q_val = str(ws.cell(row=r, column=17).value).upper().strip() if ws.cell(row=r, column=17).value else ""
            s_val = str(ws.cell(row=r, column=19).value).upper().strip() if ws.cell(row=r, column=19).value else ""

            if is_splitter_or_node:
                if "COMMON" in q_val:
                    for c in range(11, 21):
                        ws.cell(row=r, column=c).fill = FILLS["ADD8E6"]
                elif "OUT" in q_val:
                    if "1X2" in s_val:
                        for c in range(11, 21):
                            ws.cell(row=r, column=c).fill = FILLS["DB7093"]
                    else:
                        for c in range(11, 21):
                            ws.cell(row=r, column=c).fill = FILLS["FFB6C1"]
                elif "CH" in q_val and "MUX" in s_val and "DEMUX" not in s_val:
                    for c in range(11, 21):
                        ws.cell(row=r, column=c).fill = FILLS["FFDAB9"]
                elif "CH" in q_val and "DEMUX" in s_val:
                    for c in range(11, 21):
                        ws.cell(row=r, column=c).fill = FILLS["FFA07A"]

    os.makedirs(os.path.dirname(COLORIZED_CUT_SHEET_PATH), exist_ok=True)
    wb.save(COLORIZED_CUT_SHEET_PATH)
    print(f"✅ Colorized cut sheet saved to {COLORIZED_CUT_SHEET_PATH}")

if __name__ == "__main__":
    apply_colors_to_cut_sheet()

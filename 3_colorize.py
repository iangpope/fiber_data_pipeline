"""
3_colorize.py -- Colorize the cut sheet workbook using direction colors from
the Colored Connections Table, then highlight the OPTICAL SPLITTERS sub-table.

This step combines what were previously two separate scripts into a single
workbook load/save cycle:

  Pass 1 -- Main SHEATHS section coloring:
      Each sheath row is colored by the direction color assigned to its cable
      in step 2. PORT rows in tap sheets are highlighted yellow. Splitter port
      rows (COMMON, OUT 1x32, OUT 1x2, MUX, DEMUX) receive distinct colors
      to distinguish port types at a glance.

  Pass 2 -- OPTICAL SPLITTERS sub-table coloring:
      Each splitter row is colored by device type (COMMON, 1X32, 1X2, MUX,
      DEMUX) using columns B, C, G, and H as the source of truth. The cable
      connection color from step 2 takes priority over the device-type color
      on the right-side columns (F-O).

  Pass 3 -- Device group highlighting:
      The first row of each device group in the OPTICAL SPLITTERS section has
      columns A and B highlighted yellow so it is easy to see where one device
      ends and the next begins.

Reads:  data/cut_sheet.xlsx
        output/Colored_Connections_Table.xlsx
Writes: output/Colorized_Cut_Sheet_Final_v7_highlighted.xlsx
"""

import os
import re
import openpyxl

from config import DATA_DIR, OUTPUT_DIR, COLOR, FILLS, get_fill, CONN_RAW_FUSION
from naming_utils import (
    is_location_sheet,
    sheet_has_optical_splitters,
    optical_splitters_row,
    safe_fill_hex,
    find_header_row,
    find_col_by_header,
)


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
COLORED_CONNECTIONS_PATH = os.path.join(OUTPUT_DIR, "Colored_Connections_Table.xlsx")
CUT_SHEET_PATH           = os.path.join(DATA_DIR,   "cut_sheet.xlsx")
OUTPUT_PATH              = os.path.join(OUTPUT_DIR, "Colorized_Cut_Sheet_Final_v7_highlighted.xlsx")

# Matches a fiber-count prefix at the start of a cable name in legacy MST format
# (e.g. "096CT_RC73E_FT_032_TO_..."). Used to detect MST cables by naming
# pattern when they don't appear in the connections table.
MST_CABLE_RX = re.compile(r"\d{3}CT_")

YELLOW_FILL = get_fill("FFFF00")


# ---------------------------------------------------------------------------
# Load cable-to-color mapping from the Colored Connections Table
# ---------------------------------------------------------------------------

def load_cable_colors() -> dict:
    """
    Build a nested dict mapping  location -> {cable_name -> hex_color}  from
    the Colored Connections Table produced by step 2.

    Each cell in the connections table (columns D onward) contains a cable name
    and may have a directional fill color. safe_fill_hex() is used instead of
    reading the fill directly because openpyxl can return indexed or theme-based
    color objects that do not have a usable RGB string; safe_fill_hex normalizes
    these to a plain 6-character hex string or returns None if the cell is unfilled.

    The resulting dict is used by the colorize functions to look up which color
    to apply to each cable row in the cut sheet.
    """
    wb = openpyxl.load_workbook(COLORED_CONNECTIONS_PATH)
    ws = wb.active
    cable_map: dict = {}

    for row in ws.iter_rows(min_row=2):
        loc = row[0].value
        if not loc:
            continue
        for cell in row[3:]:       # connection columns start at index 3 (column D)
            if not cell.value:
                continue
            rgb = safe_fill_hex(cell)
            if rgb:
                # Store as cable_map[location][cable_name] = hex_color
                cable_map.setdefault(loc, {})[cell.value.strip()] = rgb

    return cable_map


# ---------------------------------------------------------------------------
# Pass 1 -- Color the main SHEATHS section
# ---------------------------------------------------------------------------

def _apply_main_section_colors(ws, conn_dict: dict, splitter_row: int | None) -> None:
    """
    Apply directional and special-purpose colors to all rows in the main
    SHEATHS section (from the row after the header through the row before
    OPTICAL SPLITTERS, or to the last row if there is no OPTICAL SPLITTERS
    section).

    Left-side columns (col 1 through CONNECTION-1):
      The cable name is read from the SHEATH NAME column. If a color exists in
      conn_dict for that cable, the entire left side is filled with that color.
      The first occurrence of each cable name also gets the UUID and SHEATH NAME
      columns filled yellow to mark the "splice start" of that sheath block.

    Right-side columns (CONNECTION+1 through sheet end):
      The right-side SHEATH NAME identifies the far-end cable. PORT rows receive
      a yellow fill. For splitter sheets, port type colors (COMMON, 1x32, 1x2,
      MUX, DEMUX) are applied based on the PORT NAME and DEVICE NAME columns.
    """
    # Detect column positions from the SHEATHS header row so the function is
    # resilient to column additions or reordering in future Magellan exports.
    hdr_row      = find_header_row(ws, "SHEATH UUID") or 6
    sheath_col   = find_col_by_header(ws, hdr_row, "SHEATH NAME")                      or 2
    conn_col     = find_col_by_header(ws, hdr_row, "CONNECTION")                        or 10
    r_sheath_col = find_col_by_header(ws, hdr_row, "SHEATH NAME", after_col=conn_col)   or 15
    port_col     = find_col_by_header(ws, hdr_row, "PORT NAME")                         or 17
    device_col   = find_col_by_header(ws, hdr_row, "DEVICE NAME")                       or 19
    right_end    = ws.max_column

    is_splitter = splitter_row is not None
    seen_left: set = set()   # tracks cables already colored on the left side

    end_row = (splitter_row - 1) if splitter_row else ws.max_row

    for r in range(hdr_row + 1, end_row + 1):
        # -- Left-side (incoming sheath) --
        val_left = ws.cell(row=r, column=sheath_col).value
        if isinstance(val_left, str):
            val_left = val_left.strip()
            fill = conn_dict.get(val_left)

            # Fallback: recognize legacy MST cable naming by pattern
            if not fill and MST_CABLE_RX.match(val_left):
                fill = COLOR["MST"]

            if fill and fill in FILLS:
                # Color the full left side (UUID column through CONNECTION-1).
                for c in range(1, conn_col):
                    ws.cell(row=r, column=c).fill = FILLS[fill]

                # On the first row for this cable, mark the UUID and SHEATH NAME
                # columns yellow to indicate the start of this sheath's splice block.
                if val_left not in seen_left:
                    ws.cell(row=r, column=1).fill          = FILLS[COLOR["FUSION"]]
                    ws.cell(row=r, column=sheath_col).fill = FILLS[COLOR["FUSION"]]
                    seen_left.add(val_left)

        # -- CONNECTION column -- fusion marker --
        if str(ws.cell(row=r, column=conn_col).value or "").strip() == CONN_RAW_FUSION:
            ws.cell(row=r, column=conn_col).fill = FILLS[COLOR["FUSION"]]

        # -- Right-side (outgoing sheath) --
        val_right = ws.cell(row=r, column=r_sheath_col).value
        if isinstance(val_right, str):
            val_right = val_right.strip()
            fill = conn_dict.get(val_right)
            if not fill and MST_CABLE_RX.match(val_right):
                fill = COLOR["MST"]
            if fill and fill in FILLS:
                for c in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=c).fill = FILLS[fill]

        # -- PORT rows: override right-side with yellow --
        # PORT rows mark the boundary of a fiber port block in tap sheets.
        # They must be yellow so step 7 can detect and reorder them.
        for c in range(conn_col + 1, right_end + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "PORT" in val.upper() and val.upper() != "PORT NAME":
                for cc in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=cc).fill = FILLS[COLOR["FUSION"]]
                break   # stop checking once PORT is found on this row

        # -- Splitter port type colors (right side, splitter sheets only) --
        # PORT NAME column contains the port direction: COMMON, OUT, CH
        # DEVICE NAME column contains the device type: 1X2, 1X32, MUX, DEMUX
        if is_splitter:
            q_val = str(ws.cell(row=r, column=port_col).value   or "").upper().strip()
            s_val = str(ws.cell(row=r, column=device_col).value or "").upper().strip()

            if "COMMON" in q_val:
                for c in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=c).fill = FILLS[COLOR["COMMON"]]
            elif "OUT" in q_val:
                # Distinguish 1x2 (dark pink) from 1x32 (light pink) splitter outputs.
                key = "1X2" if "1X2" in s_val else "1X32"
                for c in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=c).fill = FILLS[COLOR[key]]
            elif "CH" in q_val and "MUX" in s_val and "DEMUX" not in s_val:
                for c in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=c).fill = FILLS[COLOR["MUX"]]
            elif "CH" in q_val and "DEMUX" in s_val:
                for c in range(conn_col + 1, right_end + 1):
                    ws.cell(row=r, column=c).fill = FILLS[COLOR["DEMUX"]]


# ---------------------------------------------------------------------------
# Pass 2 -- Color the OPTICAL SPLITTERS sub-table
# ---------------------------------------------------------------------------

def _apply_optical_splitter_colors(ws, conn_dict: dict, splitter_row: int) -> None:
    """
    Apply device-type colors to the OPTICAL SPLITTERS sub-table.

    Left columns (up to CONNECTION) are colored by the device type in the
    PORT NAME column (COMMON, 1X32, 1X2, MUX, DEMUX). Right columns (after
    CONNECTION) are colored by the connected device type from the right-side
    PORT NAME and DEVICE NAME columns. If the far-end SHEATH NAME matches an
    entry in conn_dict, the directional color overrides the device-type color.
    """
    # Locate the header row of the OPTICAL SPLITTERS sub-table by scanning for
    # a "CONNECTION" cell in the rows immediately following the section label.
    spl_hdr_row = None
    for r in range(splitter_row + 1, min(ws.max_row, splitter_row + 5) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "CONNECTION":
                spl_hdr_row = r
                break
        if spl_hdr_row:
            break

    if spl_hdr_row is None:
        return   # no header row found; skip this section

    # Detect column positions from the splitter sub-table header row.
    l_device_col    = find_col_by_header(ws, spl_hdr_row, "DEVICE NAME")                          or 2
    l_port_col      = find_col_by_header(ws, spl_hdr_row, "PORT NAME")                             or 3
    s_conn_col      = find_col_by_header(ws, spl_hdr_row, "CONNECTION")                            or 5
    r_port_col      = find_col_by_header(ws, spl_hdr_row, "PORT NAME",   after_col=s_conn_col)     or 7
    r_device_col    = find_col_by_header(ws, spl_hdr_row, "DEVICE NAME", after_col=s_conn_col)     or 8
    sheath_name_col = find_col_by_header(ws, spl_hdr_row, "SHEATH NAME", after_col=s_conn_col)     or 13
    right_end       = ws.max_column

    for r in range(spl_hdr_row + 1, ws.max_row + 1):
        val_b = str(ws.cell(row=r, column=l_device_col).value   or "")
        val_c = str(ws.cell(row=r, column=l_port_col).value     or "")
        val_e = str(ws.cell(row=r, column=s_conn_col).value     or "").strip().upper()
        val_g = str(ws.cell(row=r, column=r_port_col).value     or "")
        val_h = str(ws.cell(row=r, column=r_device_col).value   or "")
        val_n = str(ws.cell(row=r, column=sheath_name_col).value or "").strip()

        # -- Left-side colors: device type of the left port --
        if "COMMON" in val_c.upper():
            for c in range(1, s_conn_col):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["COMMON"])
        elif "1X32" in val_b.upper():
            for c in range(1, s_conn_col):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["1X32"])
        elif "1X2" in val_b.upper():
            for c in range(1, s_conn_col):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["1X2"])
        elif "MUX" in val_b.upper() and "DEMUX" not in val_b.upper():
            for c in range(1, s_conn_col):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["MUX"])
        elif "DEMUX" in val_b.upper():
            for c in range(1, s_conn_col):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["DEMUX"])

        # -- Right-side colors: device type of the connected port --
        if "COMMON" in val_g.upper():
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["COMMON"])
        elif "1X32" in val_h.upper():
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["1X32"])
        elif "1X2" in val_h.upper():
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["1X2"])
        elif "MUX" in val_h.upper() and "DEMUX" not in val_h.upper():
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["MUX"])
        elif "DEMUX" in val_h.upper():
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(COLOR["DEMUX"])

        # Override right-side with connection color if available in the color table.
        if val_n in conn_dict:
            for c in range(s_conn_col + 1, right_end + 1):
                ws.cell(row=r, column=c).fill = get_fill(conn_dict[val_n])

        # Mark fusion splices in the CONNECTION column yellow.
        if val_e == CONN_RAW_FUSION:
            ws.cell(row=r, column=s_conn_col).fill = get_fill(COLOR["FUSION"])


# ---------------------------------------------------------------------------
# Pass 3 -- Highlight the first row of each device group
# ---------------------------------------------------------------------------

def _highlight_device_group_starts(ws, splitter_row: int) -> None:
    """
    Yellow-highlight columns A and B on the first row of each device group
    in the OPTICAL SPLITTERS section.

    Device groups are identified by changes in the column A cell value (which
    holds the device UUID or identifier). When the value in column A changes
    relative to the previous row, that row is the start of a new device group.
    Highlighting the group-start row makes it easy to see at a glance where
    one splitter or MUX ends and the next begins.
    """
    data_start = splitter_row + 2   # row after the OPTICAL SPLITTERS header row
    prev_val = None

    for r in range(data_start, ws.max_row + 1):
        col_a_val = ws.cell(row=r, column=1).value
        # Highlight this row if it is the first data row or the device UUID changed.
        if r == data_start or col_a_val != prev_val:
            ws.cell(row=r, column=1).fill = YELLOW_FILL
            ws.cell(row=r, column=2).fill = YELLOW_FILL
        prev_val = col_a_val


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    global COLORED_CONNECTIONS_PATH, CUT_SHEET_PATH, OUTPUT_PATH
    COLORED_CONNECTIONS_PATH = os.path.join(output_dir, "Colored_Connections_Table.xlsx")
    CUT_SHEET_PATH           = os.path.join(data_dir,   "cut_sheet.xlsx")
    OUTPUT_PATH              = os.path.join(output_dir, "Colorized_Cut_Sheet_Final_v7_highlighted.xlsx")
    cable_colors = load_cable_colors()
    wb = openpyxl.load_workbook(CUT_SHEET_PATH)

    for sheetname in wb.sheetnames:
        # Skip non-location sheets (Index, Legend, etc.)
        if not is_location_sheet(sheetname):
            continue

        ws         = wb[sheetname]
        conn_dict  = cable_colors.get(sheetname, {})

        # -- Pre-pass: remove sub-circuit rows injected by Magellan and blank rows.
        #
        # Magellan inserts annotation rows between fiber rows that contain only
        # wavelength/circuit data (the sub-circuit columns). These rows have no
        # SHEATH UUID in column A. Every real data row has a UUID in column A;
        # section header rows ("SHEATHS", "OPTICAL SPLITTERS") also have text
        # in column A. Rows with a None in column A are safe to delete.
        #
        # Scanning from the bottom prevents row-index shifting during deletion.
        hdr_row = find_header_row(ws, "SHEATH UUID") or 6
        rows_to_delete = []
        for r in range(ws.max_row, hdr_row, -1):
            if ws.cell(row=r, column=1).value is None:
                rows_to_delete.append(r)
        for r in rows_to_delete:
            ws.delete_rows(r)

        # Detect the OPTICAL SPLITTERS section row (or None if this sheet has none).
        srow = optical_splitters_row(ws)

        # Run the three color passes.
        _apply_main_section_colors(ws, conn_dict, srow)

        if srow is not None:
            _apply_optical_splitter_colors(ws, conn_dict, srow)
            _highlight_device_group_starts(ws, srow)

    os.makedirs(output_dir, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Colorized + highlighted cut sheet saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

"""
4_format_top_section.py -- Insert a metadata header and direction bar legend
into each location sheet of the colorized cut sheet workbook.

For each location sheet that exists in the Colored Connections Table, this
script:

  1. Removes any previously inserted top section (rows above SHEATHS) so it
     is safe to re-run without duplicating the header.

  2. Inserts a metadata block (rows 2-7):
       A2:A7  -- label column: Splice ID, Enclosure, No. of Trays, Location,
                  Latitude, Longitude
       B2:B7  -- value column filled with pale yellow

  3. Builds a set of colored direction bars (rows 9 onward, one per cable
     direction/type). Each bar spans columns A-E with the direction label,
     cable identifier, and fiber count in specific cells.

  Bar types and their order:
       OLT   -- olive green  (connection to the head-end OLT rack)
       MST   -- red          (main sheath terminal tap)
       North -- orange
       East  -- green
       West  -- slate
       South -- brown
       Splitter -- pink/dark-pink (1x32 / 1x2 splitter devices)
       MUX/DEMUX -- peach/salmon

A debug CSV is written to output/step5_bar_debug.csv for validating color
classification during development. Set DEBUG_ENABLED = False to suppress it.

Reads:  output/Colored_Connections_Table.xlsx
        output/Colorized_Cut_Sheet_Final_v7_highlighted.xlsx
Writes: output/Combined_Formatted_Output.xlsx
        output/step5_bar_debug.csv  (if DEBUG_ENABLED)
"""

import re
import csv
import os
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

from config import COLOR
from naming_utils import safe_fill_hex


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
DEBUG_ENABLED   = True   # write classification details to CSV for review


# ---------------------------------------------------------------------------
# Constants and color mappings
# ---------------------------------------------------------------------------

# Regex to recognize a bare OLT site token (e.g. RC73E, MS90E).
OLT_TOKEN_RX = re.compile(r"^[A-Z]{2}\d{2,3}E$", flags=re.IGNORECASE)

# Hex values for special bar types (pulled from the central config).
COLOR_SPLIT_1X2  = COLOR["1X2"]    # dark pink  -- 1x2 splitter bar
COLOR_SPLIT_1X32 = COLOR["1X32"]   # light pink -- 1x32 splitter bar
COLOR_MUX        = COLOR["MUX"]    # peach      -- MUX bar
COLOR_DEMUX      = COLOR["DEMUX"]  # salmon     -- DEMUX bar
LABEL_FILL       = "FFF9DB"        # pale yellow -- metadata value cell background
OLT_BAR_COLOR    = COLOR["OLT"]    # olive green -- OLT connection bar

# Map from fill hex back to a direction label used when classifying cable colors.
# Only the four cardinal directions, OLT, and MST are mapped.
COLOR_TO_DIRECTION = {
    COLOR["NORTH"]: "North",
    COLOR["SOUTH"]: "South",
    COLOR["EAST"]:  "East",
    COLOR["WEST"]:  "West",
    OLT_BAR_COLOR:  "OLT",
    COLOR["MST"]:   "MST",
}

# Display order for the direction bars in each sheet's top section.
# Lower numbers appear first. Directions not listed default to order 99.
BAR_ORDER = {
    "OLT":      0,
    "MST":      1,
    "North":    2,
    "East":     3,
    "West":     4,
    "South":    5,
    "Splitter": 6,
    "MUX":      7,
    "DEMUX":    8,
}

# Reference palette for nearest-color classification of cable fill colors.
# Stored as 6-char hex RGB for direct comparison.
PALETTE_RGB = {
    "North": COLOR["NORTH"],
    "South": COLOR["SOUTH"],
    "East":  COLOR["EAST"],
    "West":  COLOR["WEST"],
    "MST":   COLOR["MST"],
}


# ---------------------------------------------------------------------------
# Color utility functions
# ---------------------------------------------------------------------------

def _hex_to_rgb(h: str):
    """
    Convert a 6-character hex string to an (R, G, B) integer tuple.
    Returns None if the string is not a valid 6-character hex color.
    """
    h = (h or "").strip().upper()
    h = h[-6:] if len(h) >= 6 else h
    if len(h) != 6 or not all(ch in "0123456789ABCDEF" for ch in h):
        return None
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))


def classify_direction_from_color(rgb6: str):
    """
    Map a 6-character hex color to the nearest direction label in PALETTE_RGB.

    Uses squared Euclidean distance in RGB space. A conservative threshold
    (120 per channel) is applied so that unrecognized theme or accent colors
    are not mislabeled -- they return None instead.

    Returns 'North', 'East', 'South', 'West', 'MST', or None.
    """
    rgb = _hex_to_rgb(rgb6)
    if rgb is None:
        return None
    best = None
    for name, pal_hex in PALETTE_RGB.items():
        prgb = _hex_to_rgb(pal_hex)
        if prgb is None:
            continue
        # Squared Euclidean distance in RGB space.
        d = (rgb[0] - prgb[0])**2 + (rgb[1] - prgb[1])**2 + (rgb[2] - prgb[2])**2
        if best is None or d < best[0]:
            best = (d, name)
    # Only accept the match if it is within the tolerance threshold.
    if best and best[0] <= 120**2:
        return best[1]
    return None


def palette_distance_report(rgb6: str):
    """
    Return (best_name, best_dist_sq, dist_sq_map) for debugging.

    Computes the squared RGB distance from rgb6 to every palette color and
    returns the closest match along with the full distance map. Used only
    when DEBUG_ENABLED is True to populate the debug CSV.
    """
    rgb = _hex_to_rgb(rgb6)
    if rgb is None:
        return (None, None, {})
    dist_map  = {}
    best_name = None
    best_d    = None
    for name, pal_hex in PALETTE_RGB.items():
        prgb = _hex_to_rgb(pal_hex)
        if prgb is None:
            continue
        d = (rgb[0] - prgb[0])**2 + (rgb[1] - prgb[1])**2 + (rgb[2] - prgb[2])**2
        dist_map[name] = d
        if best_d is None or d < best_d:
            best_d    = d
            best_name = name
    return (best_name, best_d, dist_map)


def to_argb(rgb6: str) -> str:
    """
    Convert a 6-character RGB hex string to an 8-character ARGB string with
    full opacity (FF alpha prefix). openpyxl's PatternFill requires ARGB format.
    """
    rgb6 = (rgb6 or "").strip().upper()[-6:]
    return "FF" + rgb6


# ---------------------------------------------------------------------------
# Worksheet helper functions
# ---------------------------------------------------------------------------

def parse_connection_value(val: str):
    """
    Parse a cable connection value string into (fiber_count, endpoint_A, endpoint_B).

    Supports both naming conventions:
      Old: '48CT RC73E_FT_001 TO RC73E_SE_001'
      New: 'RC73E_FT_001_TO_RC73E_SE_001_48CT'

    Returns (fiber, locA, locB) on success, or (None, None, None) if unparseable.
    """
    t = str(val or "").strip()
    if not t:
        return None, None, None

    # Old format: fiber count comes first, separated by spaces.
    m = re.match(r"^(\d{2,3}CT)\s+(\S+)\s+TO\s+(\S+)$", t, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper(), m.group(2).strip(), m.group(3).strip()

    # New format: "_TO_" delimiter; fiber count is a suffix on the second endpoint.
    if "_TO_" in t:
        a, b = t.split("_TO_", 1)
        fiber = None
        m2 = re.search(r"_(\d{2,3}CT)\b", b, flags=re.IGNORECASE)
        if m2:
            fiber = m2.group(1).upper()
            b = re.sub(r"_(\d{2,3}CT)\b", "", b, flags=re.IGNORECASE)
        return fiber, a.strip(), b.strip()

    return None, None, None


def cell_fill_hex(cell) -> str | None:
    """
    Return the last-6-character RGB hex for a cell with a solid fill.
    Returns None for cells with no fill or non-solid fill patterns.

    This is a local duplicate of safe_fill_hex for cells read from the
    connections table workbook where the fill was set programmatically.
    """
    f = getattr(cell, "fill", None)
    if not f or getattr(f, "patternType", None) != "solid":
        return None
    sc = getattr(f, "start_color", None)
    if sc is not None:
        rgb = getattr(sc, "rgb", None) or getattr(sc, "index", None)
    if isinstance(rgb, str):
        rgb = rgb.strip().upper()
        return rgb[-6:]   # strip FF alpha prefix if present
    return None


def find_row_in_colA(ws, text: str) -> int | None:
    """
    Return the row number of the first cell in column A whose string value
    exactly matches text, or None if not found.
    """
    for cell in ws["A"]:
        if str(cell.value).strip() == text:
            return cell.row
    return None


def guess_olt_token(sheet_name: str) -> str:
    """
    Derive the OLT site token from a sheet name.

    For bare OLT sheets (e.g. 'RC73E'), the name itself is the token.
    For location sheets (e.g. 'RC73E_FT_001'), the prefix before the first
    underscore is the token. Returns empty string if no token can be found.
    """
    s = str(sheet_name).strip()
    if OLT_TOKEN_RX.match(s):
        return s
    prefix = s.split("_", 1)[0].strip()
    if OLT_TOKEN_RX.match(prefix):
        return prefix
    return ""


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    """
    Run step 4: insert metadata headers and direction bar legends into each
    location sheet of the colorized cut sheet workbook.

    Parameters
    ----------
    data_dir : str
        Unused by this step (all inputs come from output_dir).
    output_dir : str
        Directory containing the step 2 and step 3 output workbooks and
        where the step 4 output will be written.
    """
    conn_table_path = os.path.join(output_dir, "Colored_Connections_Table.xlsx")
    input_wb_path   = os.path.join(output_dir, "Colorized_Cut_Sheet_Final_v7_highlighted.xlsx")
    output_wb_path  = os.path.join(output_dir, "Combined_Formatted_Output.xlsx")
    debug_dump_path = os.path.join(output_dir, "step5_bar_debug.csv")

    # ------------------------------------------------------------------
    # Load connection table data
    # ------------------------------------------------------------------
    conn_wb    = load_workbook(conn_table_path)
    conn_sheet = conn_wb["Colored Connections"]

    # conn_dict: location -> list of (fiber, locA, locB, fill_hex) per cable
    # coords:    location -> (latitude, longitude)
    conn_dict: dict = {}
    coords:    dict = {}

    for row in conn_sheet.iter_rows(min_row=2):
        loc = row[0].value
        if loc is None:
            continue

        lat = row[1].value
        lon = row[2].value
        coords[str(loc).strip()] = (lat, lon)

        connections = []
        for cell in row[3:]:
            val = cell.value
            if not val:
                continue
            fiber, locA, locB = parse_connection_value(val)
            if not fiber:
                continue
            fill_color = safe_fill_hex(cell)
            connections.append((fiber, str(locA), str(locB), (fill_color or "").upper()))

        conn_dict[str(loc).strip()] = connections

    # ------------------------------------------------------------------
    # Load the colorized cut sheet workbook
    # ------------------------------------------------------------------
    wb = load_workbook(input_wb_path)

    yellow_fill = PatternFill(start_color=LABEL_FILL, end_color=LABEL_FILL, fill_type="solid")
    white_fill  = PatternFill(start_color="FFFFFF",   end_color="FFFFFF",   fill_type="solid")

    thin_side  = Side(border_style="thin",  color="000000")
    thick_side = Side(border_style="thick", color="000000")

    debug_rows = []
    run_id     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ------------------------------------------------------------------
    # Process each sheet
    # ------------------------------------------------------------------
    for sheet_name in wb.sheetnames:
        loc_key = str(sheet_name).strip()
        if loc_key not in conn_dict:
            continue

        ws = wb[sheet_name]

        sheaths_row = find_row_in_colA(ws, "SHEATHS")
        if sheaths_row and sheaths_row > 1:
            ws.delete_rows(1, sheaths_row - 1)

        olt_id = guess_olt_token(loc_key)

        lat_val, lon_val = coords.get(loc_key, (None, None))
        lat_str = f"{lat_val:.5f}" if isinstance(lat_val, (float, int)) else (lat_val or "")
        lon_str = f"{lon_val:.5f}" if isinstance(lon_val, (float, int)) else (lon_val or "")

        labels = ["Splice ID:", "Enclosure:", "No. of Trays:", "Location:", "Latitude:", "Longitude:"]
        values = [loc_key, "", "", "", lat_str, lon_str]

        current_conns = conn_dict[loc_key]
        grouped = {}

        for fiber, locA, locB, color in current_conns:
            color = (color or "").upper()
            if not color:
                continue

            if olt_id and (
                str(locA).upper() == str(olt_id).upper() or
                str(locB).upper() == str(olt_id).upper()
            ):
                label     = "OLT"
                fill_rgb6 = OLT_BAR_COLOR
            else:
                label = classify_direction_from_color(color)
                if not label:
                    continue
                fill_rgb6 = color[-6:]

            if DEBUG_ENABLED:
                best_name, best_d, dist_map = palette_distance_report(color)
                debug_rows.append({
                    "run_id":           run_id,
                    "sheet":            loc_key,
                    "olt_id":           olt_id,
                    "fiber":            fiber,
                    "locA":             str(locA),
                    "locB":             str(locB),
                    "raw_fill":         color.upper(),
                    "olt_override":     "1" if (
                        olt_id and (
                            str(locA).upper() == str(olt_id).upper() or
                            str(locB).upper() == str(olt_id).upper()
                        )
                    ) else "0",
                    "classified_label": label,
                    "bar_fill":         fill_rgb6,
                    "best_palette":     best_name,
                    "best_dist_sq":     best_d,
                    "dist_North":       dist_map.get("North"),
                    "dist_South":       dist_map.get("South"),
                    "dist_East":        dist_map.get("East"),
                    "dist_West":        dist_map.get("West"),
                    "dist_MST":         dist_map.get("MST"),
                })

            key          = (label, fiber, fill_rgb6)
            grouped[key] = grouped.get(key, 0) + 1

        connection_bars = []
        for (label, fiber, fill_rgb6), n in grouped.items():
            textB = olt_id if label == "OLT" else None
            connection_bars.append((label, n, textB, fiber, fill_rgb6))

        splitter_bars = []
        mux_bars      = []
        found_1x32 = found_1x2 = Found_mux = found_demux = False

        opt_row = find_row_in_colA(ws, "OPTICAL SPLITTERS")
        if opt_row:
            r = opt_row + 1
            while True:
                valA = ws.cell(row=r, column=1).value
                valB = ws.cell(row=r, column=2).value
                if not valA and not valB:
                    break
                if isinstance(valB, str):
                    text = valB.upper()
                    if (not found_1x32) and ("_1X32" in text):
                        splitter_bars.append(("Splitter", 1, None, "1X32", COLOR_SPLIT_1X32))
                        found_1x32 = True
                    if (not found_1x2) and ("_1X2" in text):
                        splitter_bars.append(("Splitter", 1, None, "1X2", COLOR_SPLIT_1X2))
                        found_1x2 = True
                    if (not found_demux) and ("DEMUX" in text):
                        subtype = "40CH" if "40CH" in text else "4CH"
                        mux_bars.append(("DEMUX", 1, None, subtype, COLOR_DEMUX))
                        found_demux = True
                    if (not Found_mux) and ("MUX" in text):
                        subtype = "40CH" if "40CH" in text else "4CH"
                        mux_bars.append(("MUX", 1, None, subtype, COLOR_MUX))
                        Found_mux = True
                r += 1

        connection_bars.sort(
            key=lambda b: (BAR_ORDER.get(b[0], 99), str(b[3] or ""), str(b[2] or ""))
        )
        bars = connection_bars + splitter_bars + mux_bars

        if DEBUG_ENABLED:
            for i, (lbl, cnt, textB, fiber_or_kind, fill_hex) in enumerate(bars):
                debug_rows.append({
                    "run_id":       run_id,
                    "sheet":        loc_key,
                    "record_type":  "bar",
                    "bar_index":    i,
                    "bar_label":    lbl,
                    "bar_count":    cnt,
                    "bar_textB":    textB or "",
                    "bar_fiber":    fiber_or_kind or "",
                    "bar_fill":     fill_hex,
                })

        total_new_rows = 1 + 6 + 1 + len(bars) + 1
        ws.insert_rows(1, total_new_rows)

        meta_start = 2
        for i, (lab, val) in enumerate(zip(labels, values)):
            r  = meta_start + i
            cA = ws.cell(row=r, column=1)
            cB = ws.cell(row=r, column=2)
            cA.value = lab
            cB.value = val
            cA.font      = Font(name="Calibri", bold=True)
            cA.alignment = Alignment(horizontal="left")
            cA.fill      = white_fill
            cB.alignment = Alignment(horizontal="center")
            cB.fill      = yellow_fill

        meta_end = meta_start + len(labels) - 1

        for r in range(meta_start, meta_end + 1):
            for c in (1, 2):
                left   = thick_side if c == 1 else thin_side
                right  = thick_side if c == 2 else thin_side
                top    = thick_side if r == meta_start else thin_side
                bottom = thick_side if r == meta_end   else thin_side
                if c == 1:
                    right = thin_side
                else:
                    left = thin_side
                ws.cell(row=r, column=c).border = Border(
                    left=left, right=right, top=top, bottom=bottom
                )

        start_bar_row = meta_end + 2

        def set_bar_cell(cell, fill_hex, value=None) -> None:
            cell.value     = value
            cell.fill      = PatternFill(
                start_color=to_argb(fill_hex),
                end_color=to_argb(fill_hex),
                fill_type="solid",
            )
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for offset, (label, count, textB, fiber_or_kind, fill_hex) in enumerate(bars):
            textA = (
                f"{label} ({count})"
                if count > 1 and label not in ("OLT",)
                else label
            )
            textD = fiber_or_kind
            r     = start_bar_row + offset
            for c in range(1, 6):
                val = None
                if c == 1:
                    val = textA
                elif c == 2 and textB:
                    val = textB
                elif c == 4 and textD:
                    val = textD
                set_bar_cell(ws.cell(row=r, column=c), fill_hex, val)

    # ------------------------------------------------------------------
    # Save output workbook
    # ------------------------------------------------------------------
    wb.save(output_wb_path)

    # ------------------------------------------------------------------
    # Write debug CSV
    # ------------------------------------------------------------------
    if DEBUG_ENABLED:
        try:
            os.makedirs(os.path.dirname(debug_dump_path), exist_ok=True)
            fieldnames = [
                "run_id", "record_type", "sheet", "olt_id", "fiber", "locA", "locB",
                "raw_fill", "olt_override", "classified_label", "bar_fill",
                "best_palette", "best_dist_sq",
                "dist_North", "dist_South", "dist_East", "dist_West", "dist_MST",
                "bar_index", "bar_label", "bar_count", "bar_textB", "bar_fiber",
            ]
            with open(debug_dump_path, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fieldnames)
                w.writeheader()
                for row in debug_rows:
                    if "record_type" not in row:
                        row["record_type"] = "conn"
                    w.writerow(row)
            print(f"Debug dump: {debug_dump_path}")
        except Exception as e:
            print(f"Debug dump failed: {e}")

    print(f"Wrote: {output_wb_path}")


if __name__ == "__main__":
    main()

import os
import openpyxl
from openpyxl.styles import PatternFill

CUT_SHEET_PATH = "output/Colorized_Cut_Sheet.xlsx"
CONNECTIONS_PATH = "output/Colored_Connections_Table.xlsx"
OUTPUT_PATH = "output/Colorized_Cut_Sheet_Final_v7.xlsx"

def get_fill(hex_code):
    return PatternFill(start_color=hex_code, end_color=hex_code, fill_type="solid")

COLORS = {
    "FUSION": "FFFF00",
    "COMMON": "ADD8E6",
    "1X2": "DB7093",
    "1X32": "FFB6C1",
    "MUX": "FFDAB9",
    "DEMUX": "FFA07A",
    "YELLOW": "FFFF00"
}

def load_connection_colors():
    wb = openpyxl.load_workbook(CONNECTIONS_PATH)
    ws = wb.active
    color_map = {}
    for row in ws.iter_rows(min_row=2):
        location = str(row[0].value)
        if not location:
            continue
        color_map.setdefault(location, {})
        for cell in row[3:]:
            if cell.value and cell.fill and cell.fill.start_color and cell.fill.start_color.type == "rgb":
                color_map[location][cell.value.strip()] = cell.fill.start_color.rgb[-6:]
    return color_map

def apply_splitter_coloring():
    conn_colors = load_connection_colors()
    wb = openpyxl.load_workbook(CUT_SHEET_PATH)

    for sheetname in wb.sheetnames:
        if sheetname == "Index" or sheetname not in conn_colors:
            continue

        ws = wb[sheetname]
        colors_for_sheet = conn_colors[sheetname]
        splitter_mode = False
        header_row = None
        prev_device_uuid = ""
        prev_fill_code_ad = ""

        for r in range(7, ws.max_row + 1):
            val_a = str(ws.cell(row=r, column=1).value or "")
            val_b = str(ws.cell(row=r, column=2).value or "")
            val_c = str(ws.cell(row=r, column=3).value or "")
            val_g = str(ws.cell(row=r, column=7).value or "")
            val_h = str(ws.cell(row=r, column=8).value or "")
            val_e = str(ws.cell(row=r, column=5).value or "").strip().upper()
            val_n = str(ws.cell(row=r, column=14).value or "").strip()

            if "OPTICAL SPLITTERS" in val_a.upper():
                splitter_mode = True
                continue

            if splitter_mode and val_e == "CONNECTION":
                header_row = r
                continue

            if splitter_mode and header_row and r > header_row:
                current_device_uuid = val_a

                # Determine intended A–D color
                new_fill_code_ad = ""
                if "COMMON" in val_c.upper():
                    new_fill_code_ad = COLORS["COMMON"]
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = get_fill(new_fill_code_ad)
                elif "1X32" in val_b.upper():
                    new_fill_code_ad = COLORS["1X32"]
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = get_fill(new_fill_code_ad)
                elif "1X2" in val_b.upper():
                    new_fill_code_ad = COLORS["1X2"]
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = get_fill(new_fill_code_ad)
                elif "MUX" in val_b.upper() and "DEMUX" not in val_b.upper():
                    new_fill_code_ad = COLORS["MUX"]
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = get_fill(new_fill_code_ad)
                elif "DEMUX" in val_b.upper():
                    new_fill_code_ad = COLORS["DEMUX"]
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = get_fill(new_fill_code_ad)

                # Column G = COMMON → F–O light blue
                if "COMMON" in val_g.upper():
                    for c in range(6, 16):
                        ws.cell(row=r, column=c).fill = get_fill(COLORS["COMMON"])
                else:
                    if "1X32" in val_h.upper():
                        for c in range(6, 16):
                            ws.cell(row=r, column=c).fill = get_fill(COLORS["1X32"])
                    elif "1X2" in val_h.upper():
                        for c in range(6, 16):
                            ws.cell(row=r, column=c).fill = get_fill(COLORS["1X2"])
                    elif "MUX" in val_h.upper() and "DEMUX" not in val_h.upper():
                        for c in range(6, 16):
                            ws.cell(row=r, column=c).fill = get_fill(COLORS["MUX"])
                    elif "DEMUX" in val_h.upper():
                        for c in range(6, 16):
                            ws.cell(row=r, column=c).fill = get_fill(COLORS["DEMUX"])

                if val_n in colors_for_sheet:
                    hex_color = colors_for_sheet[val_n]
                    for c in range(6, 16):
                        ws.cell(row=r, column=c).fill = get_fill(hex_color)

                if val_e == "<- FUSION ->":
                    ws.cell(row=r, column=5).fill = get_fill(COLORS["FUSION"])

                prev_device_uuid = current_device_uuid

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✅ Saved updated cut sheet to {OUTPUT_PATH}")

if __name__ == "__main__":
    apply_splitter_coloring()


# === Highlighting logic merged from highlight_splitter_section_final.py ===

import openpyxl
from openpyxl.styles import PatternFill

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

INPUT_PATH = "output/Colorized_Cut_Sheet_Final_v7.xlsx"
OUTPUT_PATH = "output/Colorized_Cut_Sheet_Final_v7_highlighted.xlsx"

wb = openpyxl.load_workbook(INPUT_PATH)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    optical_splitter_row = None

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if isinstance(cell_val, str) and "OPTICAL SPLITTERS" in cell_val.upper():
            optical_splitter_row = row
            break

    if optical_splitter_row is None:
        continue

    header_row = optical_splitter_row + 1
    data_start_row = header_row + 1

    prev_val = None
    for r in range(data_start_row, ws.max_row + 1):
        col_a_val = ws.cell(row=r, column=1).value

        # Always highlight the first row after the header
        if r == data_start_row:
            ws.cell(row=r, column=1).fill = YELLOW_FILL
            ws.cell(row=r, column=2).fill = YELLOW_FILL
        elif col_a_val != prev_val:
            ws.cell(row=r, column=1).fill = YELLOW_FILL
            ws.cell(row=r, column=2).fill = YELLOW_FILL

        prev_val = col_a_val

wb.save(OUTPUT_PATH)
print(f"✅ Highlighted sheet saved to {OUTPUT_PATH}")

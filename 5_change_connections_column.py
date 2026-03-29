"""
5_change_connections_column.py -- Normalize the CONNECTION column and clean up
the OPTICAL SPLITTERS sub-table in each sheet of the formatted output workbook.

This step performs three structural adjustments that were left over from the raw
color-formatted output:

  1. Main sheath section: shifts columns left to remove padding columns (G, H, I)
     so that the CONNECTION marker lands in the correct output column.
  2. Optical splitters section: removes duplicate or padding columns (D and F)
     so the splitter sub-table columns align with the rest of the sheet.
  3. Cleans up CONNECTION cell values:
       - "<- CONTINUOUS ->" and "<- FUSION ->" are normalized to "<--->"
       - Yellow fill is cleared from non-PORT rows (yellow is reserved for
         fusion splice markers; clearing it here avoids leakage into output)

Reads:  output/Combined_Formatted_Output.xlsx
Writes: output/Combined_Formatted_Output_processed.xlsx
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import re
import os

from naming_utils import is_location_sheet
from config import CONN_RAW_FUSION, CONN_RAW_CONTINUOUS, CONN_FUSED, COLOR


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
input_file  = os.path.join("output", "Combined_Formatted_Output.xlsx")
output_file = os.path.join("output", "Combined_Formatted_Output_processed.xlsx")


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def find_row(ws, col_letter: str, text: str):
    """
    Return the row number of the first cell in the given column that exactly
    matches the given text, or None if not found.

    Used to locate anchor rows like 'SHEATHS' and 'OPTICAL SPLITTERS' which
    mark the boundary between different sections of a sheet.
    """
    for cell in ws[col_letter]:
        if cell.value and str(cell.value).strip() == text:
            return cell.row
    return None


def shift_row_left(ws, row_idx: int, start_col: int, num_cols: int) -> None:
    """
    Shift all cells in the given row leftward by num_cols positions, starting
    from start_col. Cells that shift past the end of the row are set to None.

    This is used to remove structural padding columns inserted during earlier
    pipeline steps without deleting an entire column (which would affect all
    sheets globally).
    """
    last_col = ws.max_column
    for col in range(start_col, last_col + 1):
        src_col = col + num_cols
        # Copy value from the source column (or None if past the end of data).
        ws.cell(row=row_idx, column=col).value = (
            ws.cell(row=row_idx, column=src_col).value
            if src_col <= last_col else None
        )
        # Copy style from the source column (or keep existing if no source).
        ws.cell(row=row_idx, column=col)._style = (
            ws.cell(row=row_idx, column=src_col)._style
            if src_col <= last_col
            else ws.cell(row=row_idx, column=col)._style
        )


# ---------------------------------------------------------------------------
# Section processors
# ---------------------------------------------------------------------------

def process_main_section(ws, sheet_type: str) -> None:
    """
    Process the main SHEATHS section of a sheet.

    Shifts all data rows left by 3 columns to remove the G/H/I padding columns
    added by earlier steps. Then normalizes the CONNECTION column and clears
    any stray yellow fill from non-PORT rows.

    The yellow fill is intentionally preserved on PORT rows in tap sheets
    because those rows are the port boundary markers used later in step 7.
    """
    sheaths_row = find_row(ws, "A", "SHEATHS")
    if not sheaths_row:
        return   # sheet does not have a SHEATHS section; nothing to do

    header_row   = sheaths_row + 1
    splitter_row = find_row(ws, "A", "OPTICAL SPLITTERS")

    # Process rows up to the start of the splitter section, or to the sheet end.
    end_row = (
        splitter_row - 1
        if splitter_row and splitter_row > header_row
        else ws.max_row
    )

    # Shift columns G-end left by 3 for every data row in this section.
    for row in range(header_row, end_row + 1):
        shift_row_left(ws, row, 7, 3)

    connection_col = 7    # column G holds the CONNECTION marker after the shift
    port_col       = 14   # column N holds the PORT NAME value

    for row in range(header_row + 1, end_row + 1):
        cell = ws.cell(row=row, column=connection_col)
        val  = str(cell.value).strip() if cell.value else ""

        # Normalize verbose connection markers to the compact arrow symbol.
        if val in (CONN_RAW_CONTINUOUS, CONN_RAW_FUSION):
            cell.value     = CONN_FUSED
            cell.alignment = Alignment(horizontal="center")
        elif val.upper() == "X":
            # X marks a splice count or break; center it for readability.
            cell.alignment = Alignment(horizontal="center")

        # Clear yellow fill from non-PORT rows. Yellow is the fusion/port color;
        # keeping it on non-port rows would cause misinterpretation in later steps.
        if val.upper() != "X":
            port_val = (
                str(ws.cell(row=row, column=port_col).value).upper()
                if ws.cell(row=row, column=port_col).value
                else ""
            )
            # Preserve yellow on tap PORT rows; clear it everywhere else.
            if not (sheet_type == "tap" and port_val.startswith("PORT")):
                fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
                if fg and COLOR["FUSION"] in fg:
                    cell.fill = PatternFill(fill_type=None)


def process_optical_section(ws) -> None:
    """
    Process the OPTICAL SPLITTERS sub-table of a sheet.

    Removes two surplus padding columns (D and then F) from every row in
    this section, shifting the remaining data left. After the shift, the
    CONNECTION column lands in column D. The same connection value normalization
    and yellow-fill cleanup logic from the main section is applied here.

    If column G contains 'DEVICE UUID' in the header row, that column is
    also removed (it is metadata that belongs in an earlier section, not here).
    """
    splitter_row = find_row(ws, "A", "OPTICAL SPLITTERS")
    if not splitter_row:
        return   # no optical splitters section on this sheet

    header_row = splitter_row + 1
    end_row    = ws.max_row

    # Remove the padding columns from every row in the splitter section.
    for row in range(header_row, end_row + 1):
        shift_row_left(ws, row, 4, 1)   # remove old column D
        shift_row_left(ws, row, 5, 1)   # remove old column F (now E after previous shift)

    connection_col = 4   # column D holds the CONNECTION marker for the splitter section

    for row in range(header_row + 1, end_row + 1):
        cell = ws.cell(row=row, column=connection_col)
        val  = str(cell.value).strip() if cell.value else ""

        # Normalize connection markers.
        if val in (CONN_RAW_CONTINUOUS, CONN_RAW_FUSION):
            cell.value     = CONN_FUSED
            cell.alignment = Alignment(horizontal="center")
        elif val.upper() == "X":
            cell.alignment = Alignment(horizontal="center")

        # Clear stray yellow fill.
        if val.upper() != "X":
            fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
            if fg and COLOR["FUSION"] in fg:
                cell.fill = PatternFill(fill_type=None)

    # If the header row has DEVICE UUID in column G, remove that column from
    # all splitter rows. DEVICE UUID is internal GIS metadata and should not
    # appear in the final splice workbook.
    header_cell = ws.cell(row=header_row, column=7)
    if str(header_cell.value).strip().upper() == "DEVICE UUID":
        for row in range(header_row, end_row + 1):
            shift_row_left(ws, row, 7, 1)


def auto_adjust_column_widths(ws) -> None:
    """
    Set each column's width to fit its longest cell value plus a small margin.

    openpyxl does not auto-fit column widths automatically; this function
    scans every populated cell and uses the character count of the cell value
    as a proxy for width. A 2-character padding is added to prevent text
    from being clipped at the column edge.
    """
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                val        = str(cell.value)
                col_letter = cell.column_letter
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), len(val))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width + 2


# ---------------------------------------------------------------------------
# Sheet type classification
# ---------------------------------------------------------------------------

def classify_sheet(sheet_name: str) -> str:
    """
    Classify a sheet by its location type: 'tap', 'splitter', 'distribution',
    'olt', or 'unknown'.

    The classification drives formatting decisions in process_main_section:
    for example, tap sheets need their PORT rows preserved, while splitter
    sheets do not. New naming convention (RC73E_FT_001 / RC73E_SE_001) is
    checked first; legacy naming (MIC...) is checked as a fallback.
    """
    s = sheet_name.strip()

    # New naming convention checks.
    if re.search(r"_FT_\d+$", s, re.IGNORECASE):
        return "tap"
    if re.search(r"_SE_\d+$", s, re.IGNORECASE):
        return "splitter"

    # Legacy MIC... naming convention checks.
    if re.match(r"MIC[A-Z]{4}S\d{3}$", s):
        return "splitter"
    if re.match(r"MIC[A-Z]{4}D\d{3}$", s):
        return "distribution"
    if re.match(r"MIC[A-Z]{4}\d{4}$", s):
        return "tap"

    # Bare OLT site token (e.g. RC73E, MS90E) -- 2-10 alphanumeric chars ending in E.
    if re.match(r"^[A-Z0-9]{2,10}E$", s, re.IGNORECASE):
        return "olt"

    return "unknown"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    global input_file, output_file
    input_file  = os.path.join(output_dir, "Combined_Formatted_Output.xlsx")
    output_file = os.path.join(output_dir, "Combined_Formatted_Output_processed.xlsx")
    wb     = load_workbook(input_file)
    counts = {}   # tally processed sheets by type for the summary line

    for ws in wb.worksheets:
        sheet_name = ws.title

        # Skip non-location sheets (Index, Legend, Notes, etc.).
        if not is_location_sheet(sheet_name):
            continue

        sheet_type = classify_sheet(sheet_name)
        counts[sheet_type] = counts.get(sheet_type, 0) + 1

        process_main_section(ws, sheet_type)
        process_optical_section(ws)
        auto_adjust_column_widths(ws)

    # Sort sheets alphabetically so the output workbook has a predictable order.
    wb._sheets.sort(key=lambda ws: ws.title)
    wb.save(output_file)

    summary = ", ".join(f"{v} {k}" for k, v in sorted(counts.items()))
    print(f"Processed file saved to: {output_file} ({summary})")


if __name__ == "__main__":
    main()

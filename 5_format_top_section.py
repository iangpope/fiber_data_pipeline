import re
import csv
import os
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

"""
Step 5: Format top section + add direction/special bars.

- Reads output/Colored_Connections_Table.xlsx ("Colored Connections")
- Loads output/Colorized_Cut_Sheet_Final_v7_highlighted.xlsx
- For each sheet that exists in the connections table:
    * Removes any previously inserted top section (everything above the "SHEATHS" row)
    * Inserts a fresh metadata table (A2:B7) with thick outer border
    * Inserts color bars (A:E) ordered: OLT, MST, North, East, West, South, Splitters, Mux/Demux
- Saves to output/Combined_Formatted_Output.xlsx
"""

# -----------------------------
# Constants
# -----------------------------
CONN_TABLE_PATH = "output/Colored_Connections_Table.xlsx"
INPUT_WB_PATH   = "output/Colorized_Cut_Sheet_Final_v7_highlighted.xlsx"
OUTPUT_WB_PATH  = "output/Combined_Formatted_Output.xlsx"

DEBUG_DUMP_PATH = "output/step5_bar_debug.csv"
DEBUG_ENABLED = True

OLT_TOKEN_RX = re.compile(r"^[A-Z]{2}\d{2,3}E$", flags=re.IGNORECASE)

COLOR_SPLIT_1X2  = "DB7093"  # 1x2 splitter bar
COLOR_SPLIT_1X32 = "FFB6C1"  # 1x32 splitter bar
COLOR_MUX        = "FFDAB9"  # MUX bar
COLOR_DEMUX      = "FFA07A"  # DEMUX bar
LABEL_FILL       = "FFF9DB"  # pale yellow for metadata value background
OLT_BAR_COLOR    = "C5D9B5"  # pale green (OLT links)

# Direction mapping by fill color (RGB only, last 6 hex digits)
COLOR_TO_DIRECTION = {
    "FFA500": "North",  # Orange
    "8B4513": "South",  # Brown
    "008000": "East",   # Green
    "708090": "West",   # Slate
    OLT_BAR_COLOR: "OLT",
    "FF0000": "MST",    # Red
}

BAR_ORDER = {
    "OLT": 0,
    "MST": 1,
    "North": 2,
    "East": 3,
    "West": 4,
    "South": 5,
    "Splitter": 6,
    "MUX": 7,
    "DEMUX": 8,
}


# Canonical palette (RGB, no alpha). Used to classify slightly-off shades reliably.
PALETTE_RGB = {
    "North": "FFA500",  # Orange
    "South": "8B4513",  # Brown
    "East":  "008000",  # Green
    "West":  "708090",  # Slate
    "MST":   "FF0000",  # Red
}

def _hex_to_rgb(h: str):
    h = (h or "").strip().upper()
    h = h[-6:] if len(h) >= 6 else h
    if len(h) != 6 or not all(ch in "0123456789ABCDEF" for ch in h):
        return None
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))

def classify_direction_from_color(rgb6: str):
    """Return 'North/East/West/South/MST' based on nearest palette color. Returns None if unknown."""
    rgb = _hex_to_rgb(rgb6)
    if rgb is None:
        return None
    best = None
    for name, pal_hex in PALETTE_RGB.items():
        prgb = _hex_to_rgb(pal_hex)
        if prgb is None:
            continue
        d = (rgb[0]-prgb[0])**2 + (rgb[1]-prgb[1])**2 + (rgb[2]-prgb[2])**2
        if best is None or d < best[0]:
            best = (d, name)
    # Conservative threshold: if it's wildly different, don't guess.
    # (This prevents random theme colors from being mislabeled.)
    if best and best[0] <= 120**2:
        return best[1]
    return None



def palette_distance_report(rgb6: str):
    """Return (best_name, best_dist_sq, dist_sq_map) for nearest palette color."""
    rgb = _hex_to_rgb(rgb6)
    if rgb is None:
        return (None, None, {})
    dist_map = {}
    best_name = None
    best_d = None
    for name, pal_hex in PALETTE_RGB.items():
        prgb = _hex_to_rgb(pal_hex)
        if prgb is None:
            continue
        d = (rgb[0]-prgb[0])**2 + (rgb[1]-prgb[1])**2 + (rgb[2]-prgb[2])**2
        dist_map[name] = d
        if best_d is None or d < best_d:
            best_d = d
            best_name = name
    return (best_name, best_d, dist_map)

def to_argb(rgb6: str):
    """openpyxl wants ARGB; force opaque."""
    rgb6 = (rgb6 or "").strip().upper()
    rgb6 = rgb6[-6:]
    return "FF" + rgb6

# -----------------------------
# Helpers
# -----------------------------
def parse_connection_value(val: str):
    """
    Parse either:
      - '48CT LOC_A TO LOC_B'
      - 'LOC_A_TO_LOC_B_48CT' (RC077E style)
    Returns (fiber, locA, locB) or (None,None,None).
    """
    t = str(val or "").strip()
    if not t:
        return None, None, None

    m = re.match(r"^(\d{2,3}CT)\s+(\S+)\s+TO\s+(\S+)$", t, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper(), m.group(2).strip(), m.group(3).strip()

    if "_TO_" in t:
        a, b = t.split("_TO_", 1)
        fiber = None
        m2 = re.search(r"_(\d{2,3}CT)\b", b, flags=re.IGNORECASE)
        if m2:
            fiber = m2.group(1).upper()
            b = re.sub(r"_(\d{2,3}CT)\b", "", b, flags=re.IGNORECASE)
        return fiber, a.strip(), b.strip()

    return None, None, None


def cell_fill_hex(cell):
    """Return last-6 RGB hex for a solid fill, or None."""
    f = getattr(cell, "fill", None)
    if not f or getattr(f, "patternType", None) != "solid":
        return None
    rgb = None
    sc = getattr(f, "start_color", None)
    if sc is not None:
        rgb = getattr(sc, "rgb", None) or getattr(sc, "index", None)
    if isinstance(rgb, str):
        rgb = rgb.strip().upper()
        # openpyxl may give ARGB like '00RRGGBB' or 'FFRRGGBB' — take last 6.
        return rgb[-6:]
    return None


def find_row_in_colA(ws, text):
    for cell in ws["A"]:
        if str(cell.value).strip() == text:
            return cell.row
    return None


def guess_olt_token(sheet_name: str) -> str:
    s = str(sheet_name).strip()
    if OLT_TOKEN_RX.match(s):
        return s
    # token before first underscore
    prefix = s.split("_", 1)[0].strip()
    if OLT_TOKEN_RX.match(prefix):
        return prefix
    return ""


# -----------------------------
# Load connections table
# -----------------------------
conn_wb = load_workbook(CONN_TABLE_PATH)
conn_sheet = conn_wb["Colored Connections"]

conn_dict = {}  # loc -> list[(fiber, locA, locB, fill_hex)]
coords = {}     # loc -> (lat, lon)

for row in conn_sheet.iter_rows(min_row=2):  # skip header
    loc = row[0].value
    if loc is None:
        continue

    lat = row[1].value
    lon = row[2].value
    coords[str(loc).strip()] = (lat, lon)

    connections = []
    for cell in row[3:]:
        val = cell.value
        if not val:
            continue
        fiber, locA, locB = parse_connection_value(val)
        if not fiber:
            continue
        fill_color = cell_fill_hex(cell)  # already last-6 RGB
        connections.append((fiber, str(locA), str(locB), (fill_color or "").upper()))

    conn_dict[str(loc).strip()] = connections

# -----------------------------
# Load main workbook
# -----------------------------
wb = load_workbook(INPUT_WB_PATH)

# Styles
yellow_fill = PatternFill(start_color=LABEL_FILL, end_color=LABEL_FILL, fill_type="solid")
white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_side  = Side(border_style="thin", color="000000")
thick_side = Side(border_style="thick", color="000000")


# Debug accumulator
debug_rows = []
run_id = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# -----------------------------
# Process sheets
# -----------------------------
for sheet_name in wb.sheetnames:
    loc_key = str(sheet_name).strip()
    if loc_key not in conn_dict:
        continue

    ws = wb[sheet_name]

    # Remove any previously inserted top section (everything above SHEATHS)
    sheaths_row = find_row_in_colA(ws, "SHEATHS")
    if sheaths_row and sheaths_row > 1:
        ws.delete_rows(1, sheaths_row - 1)

    olt_id = guess_olt_token(loc_key)

    # Metadata values
    lat_val, lon_val = coords.get(loc_key, (None, None))
    lat_str = f"{lat_val:.5f}" if isinstance(lat_val, (float, int)) else (lat_val or "")
    lon_str = f"{lon_val:.5f}" if isinstance(lon_val, (float, int)) else (lon_val or "")

    labels = ["Splice ID:", "Enclosure:", "No. of Trays:", "Location:", "Latitude:", "Longitude:"]
    values = [loc_key, "", "", "", lat_str, lon_str]

    # Build connection bars grouped by direction+fiber+color (NOT by endpoints)
    current_conns = conn_dict[loc_key]
    grouped = {}  # (label, fiber, fill_rgb6) -> count
    for fiber, locA, locB, color in current_conns:
        color = (color or "").upper()
        if not color:
            continue

        # OLT override: ONLY when one endpoint is the bare OLT token.
        if olt_id and (str(locA).upper() == str(olt_id).upper() or str(locB).upper() == str(olt_id).upper()):
            label = "OLT"
            fill_rgb6 = OLT_BAR_COLOR
        else:
            label = classify_direction_from_color(color)
            if not label:
                # Unknown color: skip rather than guess
                continue
            fill_rgb6 = color[-6:]

        # --- DEBUG: record classification details ---
        if DEBUG_ENABLED:
            best_name, best_d, dist_map = palette_distance_report(color)
            debug_rows.append({
                "run_id": run_id,
                "sheet": loc_key,
                "olt_id": olt_id,
                "fiber": fiber,
                "locA": str(locA),
                "locB": str(locB),
                "raw_fill": (color or "").upper(),
                "olt_override": "1" if (olt_id and (str(locA).upper() == str(olt_id).upper() or str(locB).upper() == str(olt_id).upper())) else "0",
                "classified_label": label,
                "bar_fill": fill_rgb6,
                "best_palette": best_name,
                "best_dist_sq": best_d,
                "dist_North": dist_map.get("North"),
                "dist_South": dist_map.get("South"),
                "dist_East": dist_map.get("East"),
                "dist_West": dist_map.get("West"),
                "dist_MST": dist_map.get("MST"),
            })
        # --- END DEBUG ---
        key = (label, fiber, fill_rgb6)
        grouped[key] = grouped.get(key, 0) + 1

    connection_bars = []
    # Store count separately so label text can't drift from the sort label/fill.
    # tuple: (label, count, textB, fiber, fill_rgb6)
    for (label, fiber, fill_rgb6), n in grouped.items():
        textB = olt_id if label == "OLT" else None
        connection_bars.append((label, n, textB, fiber, fill_rgb6))

    # Build splitter + mux/demux bars by scanning OPTICAL SPLITTERS section
    splitter_bars = []
    mux_bars = []

    found_1x32 = found_1x2 = False
    found_mux = found_demux = False

    opt_row = find_row_in_colA(ws, "OPTICAL SPLITTERS")
    if opt_row:
        r = opt_row + 1
        while True:
            valA = ws.cell(row=r, column=1).value
            valB = ws.cell(row=r, column=2).value
            if not valA and not valB:
                break

            if isinstance(valB, str):
                text = valB.upper()
                if (not found_1x32) and ("_1X32" in text):
                    splitter_bars.append(("Splitter", 1, None, "1X32", COLOR_SPLIT_1X32))
                    found_1x32 = True
                if (not found_1x2) and ("_1X2" in text):
                    splitter_bars.append(("Splitter", 1, None, "1X2", COLOR_SPLIT_1X2))
                    found_1x2 = True
                if (not found_demux) and ("DEMUX" in text):
                    subtype = "40CH" if "40CH" in text else "4CH"
                    mux_bars.append(("DEMUX", 1, None, subtype, COLOR_DEMUX))
                    found_demux = True
                if (not found_mux) and ("MUX" in text):
                    subtype = "40CH" if "40CH" in text else "4CH"
                    mux_bars.append(("MUX", 1, None, subtype, COLOR_MUX))
                    found_mux = True
            r += 1

    # Sort connection bars in desired order
    connection_bars.sort(key=lambda b: (BAR_ORDER.get(b[0], 99), str(b[3] or ''), str(b[2] or '')))

    # Final bars list in order
    bars = connection_bars + splitter_bars + mux_bars
    if DEBUG_ENABLED:
        for i, (lbl, cnt, textB, fiber_or_kind, fill_hex) in enumerate(bars):
            debug_rows.append({
                "run_id": run_id,
                "sheet": loc_key,
                "record_type": "bar",
                "bar_index": i,
                "bar_label": lbl,
                "bar_count": cnt,
                "bar_textB": textB or "",
                "bar_fiber": fiber_or_kind or "",
                "bar_fill": fill_hex,
            })

    # Insert new rows:
    # row 1 blank + 6 metadata rows + row 8 blank + bars + final blank
    total_new_rows = 1 + 6 + 1 + len(bars) + 1
    ws.insert_rows(1, total_new_rows)

    # Write metadata (A2:B7)
    meta_start = 2
    for i, (lab, val) in enumerate(zip(labels, values)):
        r = meta_start + i
        cA = ws.cell(row=r, column=1)
        cB = ws.cell(row=r, column=2)
        cA.value = lab
        cB.value = val

        # style
        cA.font = cA.font.copy(bold=True)
        # openpyxl style objects are immutable; use .copy(...) to create a new Alignment
        cA.alignment = cA.alignment.copy(horizontal="left")
        cA.fill = white_fill

        cB.alignment = cB.alignment.copy(horizontal="center")
        cB.fill = yellow_fill

    meta_end = meta_start + len(labels) - 1  # row 7

    # Thick outer border around A2:B7, thin inner
    for r in range(meta_start, meta_end + 1):
        for c in (1, 2):
            left = thick_side if c == 1 else thin_side
            right = thick_side if c == 2 else thin_side
            top = thick_side if r == meta_start else thin_side
            bottom = thick_side if r == meta_end else thin_side

            # Inner seam between A and B should be thin
            if c == 1:
                right = thin_side
            else:
                left = thin_side

            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    # Bar rows start after meta + blank row
    start_bar_row = meta_end + 2  # row 9

    def set_bar_cell(cell, fill_hex, value=None):
        cell.value = value
        cell.fill = PatternFill(start_color=to_argb(fill_hex), end_color=to_argb(fill_hex), fill_type="solid")
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
        cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrap_text=True)

    for offset, (label, count, textB, fiber_or_kind, fill_hex) in enumerate(bars):
        # Build display label from the canonical label + count
        textA = f"{label} ({count})" if count > 1 and label not in ("OLT",) else label
        textD = fiber_or_kind
        r = start_bar_row + offset
        # Fill A:E only (avoid wiping gridlines elsewhere)
        for c in range(1, 6):
            val = None
            if c == 1:
                val = textA
            elif c == 2 and textB:
                val = textB
            elif c == 4 and textD:
                val = textD
            set_bar_cell(ws.cell(row=r, column=c), fill_hex, val)

# Save
wb.save(OUTPUT_WB_PATH)

# Write debug CSV
if DEBUG_ENABLED:
    try:
        os.makedirs(os.path.dirname(DEBUG_DUMP_PATH), exist_ok=True)
        fieldnames = [
            "run_id","record_type","sheet","olt_id","fiber","locA","locB","raw_fill","olt_override",
            "classified_label","bar_fill","best_palette","best_dist_sq",
            "dist_North","dist_South","dist_East","dist_West","dist_MST",
            "bar_index","bar_label","bar_count","bar_textB","bar_fiber"
        ]
        with open(DEBUG_DUMP_PATH, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for row in debug_rows:
                if "record_type" not in row:
                    row["record_type"] = "conn"
                w.writerow(row)
        print(f"🧾 Debug dump: {DEBUG_DUMP_PATH}")
    except Exception as e:
        print(f"⚠️ Debug dump failed: {e}")

print(f"✅ Wrote: {OUTPUT_WB_PATH}")
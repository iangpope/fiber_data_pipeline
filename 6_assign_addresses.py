"""
6_assign_addresses.py -- Match each splice location to a street address using KML
leader lines for precise tap-to-address assignment.

STRATEGY
--------
The KMZ contains two types of address geometry:

  1. Leader lines (featureClass = "leaderLine"):  2-point LineString placemarks
     drawn from each service address to the tap enclosure that serves it.  One
     endpoint is near the tap; the other is near the address pin.

  2. Address points (featureClassGroup = "address", Point geometry):  GPS pins
     for each street address, inside KML Folders whose <name> is the full
     address string.

For FT tap enclosures, we build a tap->address list map by matching each
leader line's two endpoints against known tap GPS coordinates (one end) and
address GPS pins (other end).  Each tap sheet then gets the nearest address
from its own assigned list rather than the global nearest address.

For SE and OLT sheets, the leader line map will be empty, so we fall back to
the original global nearest-neighbour behaviour so nothing breaks.

Reads:  data/<project>.kmz
        data/Connections_Table.xlsx        (location -> GPS lookup)
        output/Combined_Formatted_Output_processed.xlsx
Writes: output/Combined_Formatted_Output_with_Addresses.xlsx
"""

import os
import re
import zipfile
from xml.etree import ElementTree as ET
from math import radians, cos, sin, sqrt, atan2

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Directory and file paths
# ---------------------------------------------------------------------------
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

kmz_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith(".kmz")]
if len(kmz_files) != 1:
    raise FileNotFoundError(
        f"Expected 1 KMZ file in {DATA_DIR}, found {len(kmz_files)}"
    )
KMZ_PATH = os.path.join(DATA_DIR, kmz_files[0])
print(f"Using KMZ file: {kmz_files[0]}")

CONNECTIONS_PATH    = os.path.join(DATA_DIR,   "Connections_Table.xlsx")
PROCESSED_XLSX_PATH = os.path.join(OUTPUT_DIR, "Combined_Formatted_Output_processed.xlsx")
FINAL_OUTPUT_PATH   = os.path.join(OUTPUT_DIR, "Combined_Formatted_Output_with_Addresses.xlsx")

# Maximum distance (metres) for snapping a leader line endpoint to a tap or
# address point.  Leader lines validated at 7–28 m in RC73E; 75 m is generous.
SNAP_THRESHOLD_M = 75


# ---------------------------------------------------------------------------
# Distance calculation
# ---------------------------------------------------------------------------

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Return the great-circle distance in metres between two GPS coordinates.

    Uses the Haversine formula, which is accurate over the short distances
    (< 1 km) typical within a single FTTH project area.
    """
    R     = 6_371_000
    phi1, phi2 = radians(lat1), radians(lat2)
    d_phi      = radians(lat2 - lat1)
    d_lambda   = radians(lon2 - lon1)
    a = (
        sin(d_phi / 2) ** 2
        + cos(phi1) * cos(phi2) * sin(d_lambda / 2) ** 2
    )
    return 2 * R * atan2(sqrt(a), sqrt(1 - a))


def nearest(lat: float, lon: float, candidates: list) -> tuple:
    """
    Return the (item, distance_m) from `candidates` closest to (lat, lon).

    Each item in `candidates` must have 'Latitude' and 'Longitude' keys.
    Returns (None, inf) if the list is empty.
    """
    best_item = None
    best_dist = float("inf")
    for item in candidates:
        d = haversine(lat, lon, item["Latitude"], item["Longitude"])
        if d < best_dist:
            best_dist = d
            best_item = item
    return best_item, best_dist


# ---------------------------------------------------------------------------
# KMZ parsing helpers
# ---------------------------------------------------------------------------

def _read_kml(kmz_path: str) -> str:
    """Extract and return the raw KML text from the KMZ archive."""
    with zipfile.ZipFile(kmz_path, "r") as kmz:
        kml_file = next(
            (f for f in kmz.namelist() if f.endswith(".kml")), None
        )
        if not kml_file:
            raise FileNotFoundError("No KML file found inside the KMZ.")
        return kmz.read(kml_file).decode("utf-8", errors="ignore")


def extract_leader_lines(kmz_path: str) -> list:
    """
    Parse all leader line placemarks from the KMZ.

    Leader lines are LineString elements whose ExtendedData contains:
        <Data name="featureClass"><value>leaderLine</value></Data>

    Each leader line is a 2-point segment.  We return a list of dicts:
        {"p1": (lat, lon), "p2": (lat, lon)}
    where p1 is the first coordinate and p2 is the last (they may be in
    either order relative to tap vs. address end).

    We use a regex scan on the raw KML text for speed (the file is ~6 MB).
    """
    kml = _read_kml(kmz_path)

    # Match each block that declares featureClass=leaderLine and then contains
    # a LineString coordinates block somewhere after it.
    # KML structure (with tabs/newlines):
    #   <Data name="featureClass">
    #       <value>leaderLine</value>
    #   </Data>
    #   </ExtendedData>
    #   <LineString>
    #       <coordinates>lon,lat,alt lon,lat,alt</coordinates>
    #   </LineString>
    pattern = re.compile(
        r'<value>leaderLine</value>\s*</Data>\s*</ExtendedData>\s*'
        r'<LineString>\s*<coordinates>(.*?)</coordinates>',
        re.DOTALL,
    )

    leader_lines = []
    for m in pattern.finditer(kml):
        coords_text = m.group(1).strip().replace("\n", " ").split()
        if len(coords_text) < 2:
            continue
        try:
            lon1, lat1 = map(float, coords_text[0].split(",")[:2])
            lon2, lat2 = map(float, coords_text[-1].split(",")[:2])
            leader_lines.append({
                "p1": (lat1, lon1),
                "p2": (lat2, lon2),
            })
        except (ValueError, IndexError):
            continue

    print(f"Extracted {len(leader_lines)} leader lines from KMZ.")
    return leader_lines


def extract_address_points(kmz_path: str) -> list:
    """
    Extract address placemarks from the KMZ.

    Returns a list of dicts:
        {"Address": str, "Latitude": float, "Longitude": float}

    Address folders are <Folder> elements whose <name> contains "MI" and a
    comma (e.g. "3294 104TH, GRANT, MI 49327").  The GPS pin for each address
    comes from the first <Point><coordinates> element inside that folder.
    """
    kml_bytes = None
    with zipfile.ZipFile(kmz_path, "r") as kmz:
        kml_file = next(
            (f for f in kmz.namelist() if f.endswith(".kml")), None
        )
        kml_bytes = kmz.read(kml_file)

    root = ET.fromstring(kml_bytes)
    ns   = {"kml": "http://www.opengis.net/kml/2.2"}

    addresses = []
    for folder in root.findall(".//kml:Folder", ns):
        name_elem = folder.find("kml:name", ns)
        if name_elem is None or not name_elem.text:
            continue
        folder_name = name_elem.text.strip()

        # Address folders always contain "MI" and a comma (state abbreviation).
        if "MI" not in folder_name or "," not in folder_name:
            continue

        # The address pin is in the first Point placemark inside this folder.
        for pm in folder.findall(".//kml:Placemark", ns):
            coords_elem = pm.find(".//kml:Point/kml:coordinates", ns)
            if coords_elem is None or not coords_elem.text:
                continue
            parts = coords_elem.text.strip().split(",")
            if len(parts) < 2:
                continue
            try:
                lon, lat = float(parts[0]), float(parts[1])
                addresses.append({
                    "Address":   folder_name,
                    "Latitude":  lat,
                    "Longitude": lon,
                })
                break  # one pin per address folder is enough
            except ValueError:
                continue

    print(f"Extracted {len(addresses)} address points from KMZ.")
    return addresses


# ---------------------------------------------------------------------------
# Build tap -> address list map from leader lines
# ---------------------------------------------------------------------------

def build_tap_address_map(
    leader_lines:   list,
    address_points: list,
    tap_coords:     dict,
    threshold_m:    float = SNAP_THRESHOLD_M,
) -> dict:
    """
    Match each leader line to a (tap, address) pair and return a mapping.

    Algorithm for each leader line:
      - Treat p1 and p2 as the two endpoints of the leader line.
      - Find the nearest tap to p1 and the nearest address to p1.
      - Find the nearest tap to p2 and the nearest address to p2.
      - Assign: the endpoint closest to a tap (within threshold_m) is the
        "tap end"; the other endpoint is the "address end".
      - Look up the nearest address point to the "address end" (within
        threshold_m) to get the address string.
      - If both sides resolve -> record address under that tap name.

    Returns: {tap_name: [address_str, ...]}
    """
    # Convert tap_coords dict to a searchable list for nearest()
    tap_list = [
        {"Name": name, "Latitude": lat, "Longitude": lon}
        for name, (lat, lon) in tap_coords.items()
    ]

    tap_address_map: dict[str, list] = {}
    unmatched = 0

    for ll in leader_lines:
        p1_lat, p1_lon = ll["p1"]
        p2_lat, p2_lon = ll["p2"]

        # Distance from each endpoint to the nearest tap
        nearest_tap_p1, d_tap_p1 = nearest(p1_lat, p1_lon, tap_list)
        nearest_tap_p2, d_tap_p2 = nearest(p2_lat, p2_lon, tap_list)

        # Determine which endpoint is the tap end
        if d_tap_p1 < d_tap_p2 and d_tap_p1 <= threshold_m:
            tap_name    = nearest_tap_p1["Name"]
            addr_lat, addr_lon = p2_lat, p2_lon
        elif d_tap_p2 <= threshold_m:
            tap_name    = nearest_tap_p2["Name"]
            addr_lat, addr_lon = p1_lat, p1_lon
        else:
            # Neither endpoint is close to a known tap (SE/OLT address, etc.)
            unmatched += 1
            continue

        # Find the address point closest to the address end of the leader line
        nearest_addr, d_addr = nearest(addr_lat, addr_lon, address_points)
        if nearest_addr is None or d_addr > threshold_m:
            unmatched += 1
            continue

        tap_address_map.setdefault(tap_name, [])
        addr_str = nearest_addr["Address"]
        if addr_str not in tap_address_map[tap_name]:
            tap_address_map[tap_name].append(addr_str)

    total_assignments = sum(len(v) for v in tap_address_map.values())
    print(
        f"Leader line matching: {len(leader_lines) - unmatched} matched, "
        f"{unmatched} unmatched (SE/OLT or out-of-threshold)."
    )
    print(
        f"Tap->address map: {len(tap_address_map)} taps, "
        f"{total_assignments} total address assignments."
    )
    return tap_address_map


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    try:
        # ------------------------------------------------------------------
        # Load tap GPS coordinates from Connections Table
        # ------------------------------------------------------------------
        print("Loading connection coordinates...")
        df_conn = pd.read_excel(CONNECTIONS_PATH)
        location_coords = {
            row["Location"]: (row["Latitude"], row["Longitude"])
            for _, row in df_conn.iterrows()
        }
        print(f"Loaded {len(location_coords)} locations.")

        # ------------------------------------------------------------------
        # Extract leader lines and address points from KMZ
        # ------------------------------------------------------------------
        leader_lines   = extract_leader_lines(kmz_path=KMZ_PATH)
        address_points = extract_address_points(kmz_path=KMZ_PATH)

        # Flat address list used as fallback for non-FT sheets
        all_addresses = address_points  # same objects, cleaner alias

        # ------------------------------------------------------------------
        # Build the tap -> address list map via leader line geometry
        # ------------------------------------------------------------------
        tap_address_map = build_tap_address_map(
            leader_lines=leader_lines,
            address_points=address_points,
            tap_coords=location_coords,
        )

        # Convert the full address list to a searchable form (with lat/lon)
        all_addr_searchable = address_points  # already has Latitude/Longitude

        # ------------------------------------------------------------------
        # Open the processed workbook from step 5
        # ------------------------------------------------------------------
        print("Loading output workbook...")
        wb = load_workbook(PROCESSED_XLSX_PATH)

        updated_count = 0
        for sheet_name in wb.sheetnames:
            if sheet_name not in location_coords:
                continue

            lat1, lon1 = location_coords[sheet_name]

            # --- Phase 1: use the leader-line-derived address list ----------
            assigned_addrs = tap_address_map.get(sheet_name, [])

            if assigned_addrs:
                # Build a searchable subset from the pre-matched address strings
                candidate_points = [
                    a for a in all_addr_searchable
                    if a["Address"] in assigned_addrs
                ]
                best_item, _ = nearest(lat1, lon1, candidate_points)
                best_addr = best_item["Address"] if best_item else None
            else:
                # --- Phase 2: fallback – global nearest neighbour -----------
                # (Applies to SE, OLT, and any tap that had no leader line match)
                best_item, _ = nearest(lat1, lon1, all_addr_searchable)
                best_addr = best_item["Address"] if best_item else None

            if best_addr:
                wb[sheet_name]["B5"] = best_addr
                updated_count += 1

        wb.save(FINAL_OUTPUT_PATH)
        print(f"Updated {updated_count} sheets.")
        print(f"Output saved to: {FINAL_OUTPUT_PATH}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error: {e}")


if __name__ == "__main__":
    main()

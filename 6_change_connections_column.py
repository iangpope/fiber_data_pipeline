from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import re
import os

# File paths
input_file = os.path.join("output", "Combined_Formatted_Output.xlsx")
output_file = os.path.join("output", "Combined_Formatted_Output_processed.xlsx")

def find_row(ws, col_letter, text):
    for cell in ws[col_letter]:
        if cell.value and str(cell.value).strip() == text:
            return cell.row
    return None

def shift_row_left(ws, row_idx, start_col, num_cols):
    last_col = ws.max_column
    for col in range(start_col, last_col + 1):
        src_col = col + num_cols
        ws.cell(row=row_idx, column=col).value = (
            ws.cell(row=row_idx, column=src_col).value if src_col <= last_col else None
        )
        ws.cell(row=row_idx, column=col)._style = (
            ws.cell(row=row_idx, column=src_col)._style if src_col <= last_col else ws.cell(row=row_idx, column=col)._style
        )

def process_main_section(ws, sheet_type):
    sheaths_row = find_row(ws, 'A', 'SHEATHS')
    if not sheaths_row:
        return
    header_row = sheaths_row + 1
    splitter_row = find_row(ws, 'A', 'OPTICAL SPLITTERS')
    end_row = splitter_row - 1 if splitter_row and splitter_row > header_row else ws.max_row

    for row in range(header_row, end_row + 1):
        shift_row_left(ws, row, 7, 3)

    connection_col = 7
    port_col = 14

    for row in range(header_row + 1, end_row + 1):
        cell = ws.cell(row=row, column=connection_col)
        val = str(cell.value).strip() if cell.value else ""

        if val in ["<- CONTINUOUS ->", "<- FUSION ->"]:
            cell.value = "< --- >"
            cell.alignment = Alignment(horizontal="center")
        elif val.upper() == "X":
            cell.alignment = Alignment(horizontal="center")

        if val.upper() != "X":
            port_val = str(ws.cell(row=row, column=port_col).value).upper() if ws.cell(row=row, column=port_col).value else ""
            if not (sheet_type == "tap" and port_val.startswith("PORT")):
                fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
                if fg and "FFFF00" in fg:
                    cell.fill = PatternFill(fill_type=None)

def process_optical_section(ws):
    splitter_row = find_row(ws, 'A', 'OPTICAL SPLITTERS')
    if not splitter_row:
        return
    header_row = splitter_row + 1
    end_row = ws.max_row

    for row in range(header_row, end_row + 1):
        shift_row_left(ws, row, 4, 1)  # Delete D
        shift_row_left(ws, row, 5, 1)  # Delete F (now E)

    connection_col = 4  # D

    for row in range(header_row + 1, end_row + 1):
        cell = ws.cell(row=row, column=connection_col)
        val = str(cell.value).strip() if cell.value else ""
        if val in ["<- CONTINUOUS ->", "<- FUSION ->"]:
            cell.value = "< --- >"
            cell.alignment = Alignment(horizontal="center")
        elif val.upper() == "X":
            cell.alignment = Alignment(horizontal="center")

        if val.upper() != "X":
            fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
            if fg and "FFFF00" in fg:
                cell.fill = PatternFill(fill_type=None)

    # Check for DEVICE UUID in column G of header row
    header_cell = ws.cell(row=header_row, column=7)  # Column G
    if str(header_cell.value).strip().upper() == "DEVICE UUID":
        for row in range(header_row, end_row + 1):
            shift_row_left(ws, row, 7, 1)  # Delete column G

def auto_adjust_column_widths(ws):
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                val = str(cell.value)
                col_letter = cell.column_letter
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), len(val))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width + 2

def classify_sheet(sheet_name):
    if sheet_name.endswith("E") and not sheet_name.startswith("MIC"):
        return 'olt'
    elif re.match(r"MIC[A-Z]{4}S\d{3}$", sheet_name):
        return 'splitter'
    elif re.match(r"MIC[A-Z]{4}D\d{3}$", sheet_name):
        return 'distribution'
    elif re.match(r"MIC[A-Z]{4}\d{4}$", sheet_name):
        return 'tap'
    else:
        return 'unknown'

def main():
    wb = load_workbook(input_file)
    for ws in wb.worksheets:
        sheet_name = ws.title
        sheet_type = classify_sheet(sheet_name)
        print(f"Processing {sheet_name} as {sheet_type}")
        process_main_section(ws, sheet_type)
        process_optical_section(ws)
        auto_adjust_column_widths(ws)

    wb._sheets.sort(key=lambda ws: ws.title)
    wb.save(output_file)
    print(f"✅ Processed file saved to: {output_file}")

if __name__ == "__main__":
    main()


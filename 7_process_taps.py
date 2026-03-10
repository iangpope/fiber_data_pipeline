"""
7_process_taps.py -- Process all tap (FT) sheets in three sequential stages.

This script handles all structural modifications specific to tap enclosure
sheets. Each stage operates on the same workbook; all three run before saving.

  Stage A -- Reorder sheath blocks:
      Within the SHEATHS section, PORT-containing sheath blocks must appear
      before non-PORT blocks (matching the physical splice order). This stage
      detects whether blocks are out of order and swaps them in place using
      a full row snapshot to preserve all cell styles and metadata.

  Stage B -- Shift PORT columns:
      PORT NAME, PORT WAVELENGTH, and DEVICE NAME values are moved to columns
      J, K, L (10, 11, 12) so they align consistently across all tap sheets.
      Any data that previously occupied columns M and beyond is cleared after
      the shift to avoid leftover values from the raw export.

  Stage C -- Label B3 with enclosure type:
      Cell B3 is written with the OTE (Optical Termination Enclosure) size
      string (e.g. "4 PORT OTE") based on how many PORT rows are present after
      the shift. Cell B4 is set to 1 (tap enclosures always have 1 tray).

Reads:  output/Combined_Formatted_Output_with_Addresses.xlsx
Writes: output/Combined_Reordered_With_OTE.xlsx
"""

from __future__ import annotations

import re
from copy import copy
from typing import List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import naming_utils


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
INPUT_FILE  = "output/Combined_Formatted_Output_with_Addresses.xlsx"
OUTPUT_FILE = "output/Combined_Reordered_With_OTE.xlsx"

# Regex to match "PORT 1", "PORT 12", etc. in PORT NAME cells.
PORT_RX = re.compile(r"^\s*PORT\s*\d+\s*$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Shared helpers used across all three stages
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    """Return a stripped string, or empty string if the value is None."""
    return str(s).strip() if s is not None else ""


def is_port_label(value) -> bool:
    """
    Return True if the cell value is a PORT label (e.g. "PORT 1", "PORT 12").

    Strips whitespace and checks that the value starts with "PORT" followed
    immediately by a digit, which distinguishes port number refs from the
    header cell "PORT NAME".
    """
    s = _norm(value).upper().replace(" ", "")
    return s.startswith("PORT") and len(s) > 4 and s[4].isdigit()


def find_header_row(ws, max_scan_rows: int = 60) -> int | None:
    """
    Scan the first max_scan_rows rows to find the row whose cells contain the
    text "CONNECTION" (the header row of the SHEATHS data table).

    Returns the row number, or None if not found.
    """
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "CONNECTION":
                return r
    return None


def detect_port_col(ws, header_row: int | None = None) -> int | None:
    """
    Find the column that contains PORT NAME values.

    First searches the header row for a cell with the exact text "PORT NAME".
    If that fails, falls back to scanning columns 13-17 for any cell starting
    with "PORT" to handle sheets where the header is missing or misaligned.
    """
    if header_row:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip().upper() == "PORT NAME":
                return c

    # Fallback scan across likely port columns.
    for c in [13, 14, 15, 16, 17]:
        if c > ws.max_column:
            continue
        for r in range((header_row or 1) + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper().startswith("PORT"):
                return c
    return None


def copy_cell(src, dst) -> None:
    """
    Copy all data and style attributes from one cell to another.

    openpyxl does not provide a built-in cell copy; this function replicates
    value, style, number format, alignment, border, fill, font, protection,
    and comment from src to dst.
    """
    dst.value         = src.value
    dst._style        = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment     = copy(src.alignment)
    dst.border        = copy(src.border)
    dst.fill          = copy(src.fill)
    dst.font          = copy(src.font)
    dst.protection    = copy(src.protection)
    dst.comment       = src.comment


# ---------------------------------------------------------------------------
# Stage A -- Reorder sheath blocks
# ---------------------------------------------------------------------------

def _build_sheath_blocks(ws, start_row: int) -> List[Tuple[int, int, str]]:
    """
    Identify contiguous sheath blocks in the SHEATHS section.

    A sheath block is a run of consecutive rows that all share the same cable
    name in column B (SHEATH_COL). Rows with None in column B are treated as
    continuation rows of the current block. When the cable name changes, the
    previous block is finalized.

    Returns a list of (start_row, end_row, cable_name) tuples.
    """
    SHEATH_COL = 2
    blocks: List[Tuple[int, int, str]] = []
    current_name = None
    block_start  = None

    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(r, SHEATH_COL).value
        name = name.strip() if isinstance(name, str) else None

        if name is None and current_name is None:
            continue   # leading blank rows before any sheath data
        if current_name is None and name is not None:
            current_name = name     # start of first block
            block_start  = r
            continue
        if name is None:
            continue   # continue within the current block
        if current_name is not None and name != current_name:
            # New cable name encountered; finalize previous block.
            blocks.append((block_start, r - 1, current_name))
            current_name = name
            block_start  = r

    # Close the final open block at the end of the sheet.
    if current_name is not None and block_start is not None:
        blocks.append((block_start, ws.max_row, current_name))

    return [b for b in blocks if b[1] >= b[0]]


def _block_has_ports(ws, block: Tuple[int, int, str], port_col: int) -> bool:
    """
    Return True if any row in the given block has a PORT value in port_col.

    This is used to classify blocks as "PORT blocks" (should appear first)
    vs. regular sheath blocks (appear after PORT blocks).
    """
    start, end, _ = block
    for r in range(start, end + 1):
        v = ws.cell(r, port_col).value
        if isinstance(v, str) and v.strip().upper().startswith("PORT"):
            return True
    return False


def _snapshot_rows(ws, start: int, end: int):
    """
    Take an in-memory snapshot of all cell data and row dimension metadata
    for rows start through end (inclusive).

    Returns (rows, row_dims) where:
      rows     -- list of lists of cell property dicts (one per row, per column)
      row_dims -- list of row dimension dicts (height, hidden, outlineLevel)

    This is needed because modifying row order in openpyxl requires explicitly
    reading and re-writing all cell data; there is no built-in row move or swap.
    """
    max_col  = ws.max_column
    rows     = []
    row_dims = []

    for r in range(start, end + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            row_cells.append({
                "value":        cell.value,
                "_style":       copy(cell._style),
                "number_format": cell.number_format,
                "protection":   copy(cell.protection),
                "alignment":    copy(cell.alignment),
                "comment":      cell.comment,
                "hyperlink":    cell.hyperlink,
            })
        rows.append(row_cells)
        rd = ws.row_dimensions[r]
        row_dims.append({
            "height":       rd.height,
            "hidden":       rd.hidden,
            "outlineLevel": rd.outlineLevel,
        })

    return rows, row_dims


def _write_rows(ws, start_row: int, rows, row_dims) -> None:
    """
    Write a previously snapshotted set of rows back into the worksheet
    starting at start_row, restoring all cell data and row dimension metadata.
    """
    max_col = ws.max_column
    for i, row_cells in enumerate(rows):
        r = start_row + i
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            snap = row_cells[c - 1]
            cell.value          = snap["value"]
            cell._style         = copy(snap["_style"])
            cell.number_format  = snap["number_format"]
            cell.protection     = copy(snap["protection"])
            cell.alignment      = copy(snap["alignment"])
            cell.comment        = snap["comment"]
            cell.hyperlink      = snap["hyperlink"]

        rd  = ws.row_dimensions[r]
        dim = row_dims[i]
        rd.height       = dim["height"]
        rd.hidden       = dim["hidden"]
        rd.outlineLevel = dim["outlineLevel"]


def reorder_sheath_blocks(ws) -> bool:
    """
    Move PORT-containing sheath blocks to appear before non-PORT blocks.

    Blocks are scored 0 (has ports) or 1 (no ports) and then sorted by score
    then by original position, preserving relative order within each group.
    If the blocks are already in the correct order, no changes are made.

    Returns True if any reordering was performed.
    """
    header_row = find_header_row(ws)
    port_col   = detect_port_col(ws, header_row)
    if port_col is None:
        return False

    blocks = _build_sheath_blocks(ws, (header_row or 10) + 1)
    if len(blocks) < 2:
        return False   # only one block; nothing to reorder

    # Sort blocks: PORT blocks (score 0) before non-PORT (score 1),
    # preserving original relative order within each group (stable sort).
    scored        = [(0 if _block_has_ports(ws, b, port_col) else 1, idx, b)
                     for idx, b in enumerate(blocks)]
    scored_sorted = sorted(scored, key=lambda t: (t[0], t[1]))

    if [t[2] for t in scored_sorted] == [t[2] for t in scored]:
        return False   # already in the correct order

    # Snapshot all blocks before writing (avoid overwriting source data).
    region_start = min(b[0] for b in blocks)
    block_snaps  = [(_snapshot_rows(ws, b[0], b[1]), b[1] - b[0] + 1)
                    for _, _, b in scored]

    write_ptr = region_start
    for _, orig_idx, _ in scored_sorted:
        (rows, dims), height = block_snaps[orig_idx]
        _write_rows(ws, write_ptr, rows, dims)
        write_ptr += height

    return True


# ---------------------------------------------------------------------------
# Stage B -- Shift PORT columns to J/K/L (10/11/12)
# ---------------------------------------------------------------------------

def shift_port_columns(ws) -> int:
    """
    Move PORT NAME, PORT WAVELENGTH, and DEVICE NAME values from wherever they
    currently appear in the source data to a fixed position at columns J/K/L
    (10/11/12). Clear any data in column M onward after the move.

    The fixed target position ensures that downstream steps (8_finalize.py and
    any further Excel work) can always find port metadata in the same columns.

    Returns the number of PORT rows that were shifted.
    """
    header_row    = find_header_row(ws)
    port_name_col = detect_port_col(ws, header_row=header_row)

    if not port_name_col:
        return 0   # no PORT NAME column found; nothing to do

    max_row        = ws.max_row
    max_col        = ws.max_column
    target_start_col = 10   # target: column J
    shifted_rows   = 0

    # Copy the three port data columns (PORT NAME, WAVELENGTH, DEVICE NAME)
    # to the fixed target position for each PORT-labelled row.
    for r in range((header_row or 1) + 1, max_row + 1):
        if is_port_label(ws.cell(r, port_name_col).value):
            shifted_rows += 1
            for i in range(3):
                copy_cell(
                    ws.cell(r, port_name_col + i),
                    ws.cell(r, target_start_col + i),
                )

    # After the shift, clear everything from column M (13) onward to remove
    # the original (now redundant) port data and any leftover junk columns.
    if shifted_rows:
        for r in range(1, max_row + 1):
            for c in range(13, max_col + 1):
                cell       = ws.cell(r, c)
                cell.value = None
                cell.fill  = PatternFill(fill_type=None)

    return shifted_rows


# ---------------------------------------------------------------------------
# Stage C -- Label B3 with OTE enclosure type
# ---------------------------------------------------------------------------

def _detect_port_col_after_shift(sheet) -> int | None:
    """
    Find the PORT NAME column in the sheet after the column shift has been
    applied. Scans rows 13 onward (skipping the metadata header block) in
    columns 8-15 for a cell value starting with "PORT".
    """
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        for col in range(8, 16):
            if len(row) >= col:
                v = row[col - 1].value
                if v and str(v).strip().upper().startswith("PORT"):
                    return col
    return None


def _count_ports(sheet, port_col: int) -> int:
    """
    Count the number of PORT rows in the sheet by scanning port_col from
    row 13 onward.
    """
    count = 0
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        if len(row) >= port_col:
            val = row[port_col - 1].value
            if val and str(val).strip().upper().startswith("PORT"):
                count += 1
    return count


def label_port_ote(sheet, count: int) -> None:
    """
    Write the OTE enclosure size label into cell B3 based on the number of
    PORT rows found, and set B4 to 1 (all tap enclosures use 1 splice tray).

    OTE sizes are:
      2 PORT OTE  -- up to 2 ports
      4 PORT OTE  -- up to 4 ports
      8 PORT OTE  -- up to 8 ports
      12 PORT OTE -- up to 12 ports
    """
    if count <= 2:
        sheet["B3"] = "2 PORT OTE"
    elif count <= 4:
        sheet["B3"] = "4 PORT OTE"
    elif count <= 8:
        sheet["B3"] = "8 PORT OTE"
    elif count <= 12:
        sheet["B3"] = "12 PORT OTE"
    sheet["B4"] = 1   # tap enclosures always have 1 splice tray


def label_enclosure(ws) -> bool:
    """
    Detect the port column, count port rows, and write the OTE label to B3.
    Returns True if the label was written, False if no PORT column was found.
    """
    port_col = _detect_port_col_after_shift(ws)
    if not port_col:
        return False
    count = _count_ports(ws, port_col)
    if count > 0:
        label_port_ote(ws, count)
        return True
    return False


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    wb = load_workbook(INPUT_FILE)
    reordered = shifted = labeled = 0

    for name in wb.sheetnames:
        ws = wb[name]

        # Only tap (FT) sheets are processed by this script; splitter, OLT,
        # and other sheet types are left untouched.
        if not naming_utils.is_tap(name, df=None):
            continue

        # Run the three stages in order. Each returns a bool/count indicating
        # whether any change was made, used only to produce the summary line.
        if reorder_sheath_blocks(ws):
            reordered += 1

        n = shift_port_columns(ws)
        if n:
            shifted += 1

        if label_enclosure(ws):
            labeled += 1

    wb.save(OUTPUT_FILE)
    print(f"Tap processing complete: {OUTPUT_FILE}")
    print(f"   sheets reordered : {reordered}")
    print(f"   sheets shifted   : {shifted}")
    print(f"   sheets labeled   : {labeled}")


if __name__ == "__main__":
    main()

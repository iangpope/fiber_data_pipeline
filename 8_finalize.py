"""
8_finalize.py -- Final cleanups: fill enclosure labels, insert DEMUX/MST rows,
shift splitter metadata, and trim excess columns.

This script performs two cleanup passes (formerly two separate scripts) before
producing the finished as-built workbook.

  Part A -- Enclosure labeling and DEMUX/MST row insertion:
      - Fills B3 (Enclosure type) and B4 (Tray count) on any sheet where those
        cells are still blank, using the sheet name to determine the correct
        enclosure model (e.g. COMMSCOPE FOSC 450-D for SE enclosures).
      - Inserts a "DEMUX 4CH" row and an "MST 24CT" row in sheets that have
        both 1x2 and 1x32 devices in the device list above the SHEATHS section,
        because such sheets represent an SE enclosure with a paired DEMUX/MST.

  Part B -- Column cleanup and trimming:
      - In splitter sheets, moves fiber/buffer metadata values that landed in
        the wrong columns (right of the CONNECTION column) to their correct
        destination columns immediately after BUFFER.
      - Clears the SHEATH UUID column in the OPTICAL SPLITTERS sub-table
        (the UUID is internal GIS metadata, not needed in the as-built output).
      - Deletes all columns from SHEATH UUID onward in the main SHEATHS section
        to trim the sheet down to the columns that belong in the final output.

Reads:  output/Combined_Reordered_With_OTE.xlsx
Writes: output/Asbuilt_Workbook_post12.xlsx
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill

from naming_utils import (
    is_location_sheet,
    find_header_row,
    find_col_by_header,
    find_all_cols_by_header,
    safe_fill_hex,
)


# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
INPUT_FILE  = Path("output") / "Combined_Reordered_With_OTE.xlsx"
OUTPUT_FILE = Path("output") / "Asbuilt_Workbook_post12.xlsx"

# Regex to parse legacy location names and extract the enclosure type suffix
# (S = splice/SE, D = distribution/DE) and the numeric identifier.
legacy_name_rx = re.compile(
    r'^(?P<prefix>[A-Z]{2}[A-Z0-9]+?)(?P<type>S|D)?(?P<num>\d{3,4})$',
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Part A -- Enclosure labels and DEMUX/MST row insertion
# ---------------------------------------------------------------------------

def _fill_enclosure_labels(sheet, sheet_name: str) -> None:
    """
    Fill B3 (enclosure model) and B4 (tray count) if they are currently blank.

    The enclosure model is determined from the sheet name:
      _FT_  (tap)         ->  "2 PORT OTE",          1 tray
      _SE_  (splitter)    ->  "COMMSCOPE FOSC 450-D", 2 trays
      _D    (distribution)->  "COMMSCOPE FOSC 450-B", 1 tray
      legacy S suffix     ->  "COMMSCOPE FOSC 450-D", 2 trays
      legacy D suffix     ->  "COMMSCOPE FOSC 450-B", 1 tray

    If both cells already have values, this function returns without changes
    to avoid overwriting anything written by step 7.
    """
    # Locate the enclosure and tray rows by scanning column A for the label
    # text rather than relying on fixed row numbers.
    enc_row = trays_row = None
    for r in range(1, min(sheet.max_row, 20) + 1):
        v = str(sheet.cell(r, 1).value or "").strip()
        if v == "Enclosure:":
            enc_row = r
        elif v == "No. of Trays:":
            trays_row = r
        if enc_row and trays_row:
            break

    if enc_row is None or trays_row is None:
        return   # sheet does not have standard metadata layout

    enc_cell   = sheet.cell(row=enc_row,   column=2)
    trays_cell = sheet.cell(row=trays_row, column=2)

    # Skip if both enclosure and tray values are already populated.
    if (enc_cell.value and str(enc_cell.value).strip() and
            trays_cell.value and str(trays_cell.value).strip()):
        return

    name_up = sheet_name.upper()
    enc_val = trays_val = None

    # New naming convention: check for _FT_ or _SE_ in the sheet name.
    if "_FT_" in name_up or name_up.endswith("_FT"):
        enc_val, trays_val = "2 PORT OTE", 1
    elif "_SE_" in name_up or name_up.endswith("_SE") or re.search(r'S\d+$', name_up):
        enc_val, trays_val = "COMMSCOPE FOSC 450-D", 2
    elif re.search(r'D\d+$', name_up):
        enc_val, trays_val = "COMMSCOPE FOSC 450-B", 1
    else:
        # Legacy naming: parse the type letter (S/D) from the end of the name.
        m = legacy_name_rx.match(name_up)
        if m:
            t = (m.group("type") or "").upper()
            if t == "S":
                enc_val, trays_val = "COMMSCOPE FOSC 450-D", 2
            elif t == "D":
                enc_val, trays_val = "COMMSCOPE FOSC 450-B", 1

    # Only write if we determined a value and the cell is still blank.
    if enc_val and not (enc_cell.value and str(enc_cell.value).strip()):
        enc_cell.value = enc_val
    if trays_val is not None and not (trays_cell.value and str(trays_cell.value).strip()):
        trays_cell.value = trays_val


def _insert_demux_mst_rows(sheet) -> None:
    """
    Insert DEMUX and MST rows in SE enclosure sheets that have both 1x2 and
    1x32 splitter devices.

    SE enclosures that contain a 1x2 coupler (for the passive MST tap) in
    addition to 1x32 output splitters require specific DEMUX and MST label
    rows in the device metadata block above SHEATHS. This function:

      1. Scans the device block (rows 9 to first connection row) for 1x32,
         1x2, and DEMUX references.
      2. If both 1x2 and 1x32 are present but DEMUX is missing, inserts a
         "DEMUX 4CH" row at the top of the connection block.
      3. If no MST row exists just before the SHEATHS row, inserts one.
    """
    # Locate the SHEATHS anchor row.
    sheaths_row = None
    for cell in sheet['A']:
        if cell.value == "SHEATHS":
            sheaths_row = cell.row
            break
    if sheaths_row is None:
        return   # no SHEATHS section on this sheet

    meta_blank_row = 8   # device metadata ends around row 8 in standard layout

    # Find the first row in the connection/cable block (where data transitions
    # from device metadata to cable/fiber entries).
    first_conn_row = None
    for r in range(meta_blank_row + 1, sheaths_row):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        # Stop at a row that is blank in both A and B (end of device block).
        if (val_a is None or not str(val_a).strip()) and \
           (val_b is None or not str(val_b).strip()):
            break
        # The first connection row has content in A but not in B.
        if val_a and not (val_b and str(val_b).strip()):
            first_conn_row = r
            break

    if first_conn_row is None:
        return

    # Scan the device block for presence of 1x32, 1x2, and DEMUX.
    found_1x2  = False
    found_1x32 = False
    found_demux = False
    for r in range(meta_blank_row + 1, first_conn_row):
        cell_a = str(sheet.cell(row=r, column=1).value or "").upper()
        cell_b = str(sheet.cell(row=r, column=2).value or "").upper()
        if "1X32" in cell_a or "1X32" in cell_b:
            found_1x32 = True
        if "1X2" in cell_a or "1X2" in cell_b:
            found_1x2 = True
        if "DEMUX" in cell_a:
            found_demux = True

    # Only apply to sheets that have both device types (SE with MST coupler).
    if not (found_1x2 and found_1x32):
        return

    # Insert DEMUX row if missing.
    if not found_demux:
        sheet.insert_rows(first_conn_row)
        sheet.cell(row=first_conn_row, column=1).value = "DEMUX"
        sheet.cell(row=first_conn_row, column=2).value = "4CH"
        first_conn_row += 1
        if sheaths_row:
            sheaths_row += 1   # shift down to account for the inserted row

    # Check whether an MST row already exists just before SHEATHS.
    blank_above = sheaths_row - 1 if sheaths_row else None
    has_mst = any(
        isinstance(sheet.cell(row=r, column=1).value, str) and
        sheet.cell(row=r, column=1).value.strip().upper() == "MST"
        for r in range(first_conn_row, blank_above or sheaths_row)
    )

    # Insert MST row if missing.
    if not has_mst and blank_above:
        sheet.insert_rows(blank_above)
        sheet.cell(row=blank_above, column=1).value = "MST"
        sheet.cell(row=blank_above, column=4).value = "24CT"


def run_part_a(wb) -> None:
    """
    Run Part A for every qualifying sheet: fill B3/B4 enclosure labels and
    insert DEMUX/MST rows where needed.

    A qualifying sheet is identified by having "Enclosure:" in A3 and
    "No. of Trays:" in A4, which is the standard header layout for all
    location sheets.
    """
    for name in wb.sheetnames:
        sheet = wb[name]
        # Only process sheets that follow the standard location sheet layout.
        # Qualify the sheet by scanning column A for the "Enclosure:" label
        # rather than assuming it always sits at a fixed row number.
        if not any(
            str(sheet.cell(r, 1).value or "").strip() == "Enclosure:"
            for r in range(1, min(sheet.max_row, 20) + 1)
        ):
            continue
        _fill_enclosure_labels(sheet, name)
        _insert_demux_mst_rows(sheet)


# ---------------------------------------------------------------------------
# Part B -- Column shifting and trimming
# ---------------------------------------------------------------------------

def _norm(v) -> str:
    """Return the normalized (stripped, uppercase) string value of a cell."""
    return str(v).strip().upper() if v is not None else ""


def _clear_range(ws, r_min, r_max, c_min, c_max) -> None:
    """
    Clear the values and fills of all cells in the specified rectangular range.
    Used to wipe data from junk columns without deleting the columns themselves.
    """
    empty = PatternFill()
    for r in range(r_min, r_max + 1):
        for c in range(c_min, c_max + 1):
            cell       = ws.cell(r, c)
            cell.value = None
            cell.fill  = empty


def _clear_optical_splitter_sheath_uuid(ws) -> bool:
    """
    Find and clear the SHEATH UUID column in the OPTICAL SPLITTERS sub-table.

    SHEATH UUID is an internal GIS identifier used during export. It is not
    useful in the final as-built workbook and its presence can cause confusion
    with the UUIDs in the main SHEATHS table above.

    Clears all cells in the SHEATH UUID column and any blank columns immediately
    following it (which may also be part of the same UUID block).

    Returns True if the column was found and cleared, False otherwise.
    """
    # Locate the OPTICAL SPLITTERS header row.
    opt_row = None
    for r in range(1, min(ws.max_row, 400) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "OPTICAL SPLITTERS" in v.upper():
            opt_row = r
            break
    if opt_row is None:
        return False

    # Find the SHEATH UUID header cell within the splitter sub-table.
    sheath_pos = None
    for r in range(opt_row, min(ws.max_row, opt_row + 40) + 1):
        for c in range(1, min(ws.max_column, 160) + 1):
            if _norm(ws.cell(r, c).value) == "SHEATH UUID":
                sheath_pos = (r, c)
                break
        if sheath_pos:
            break

    if not sheath_pos:
        return False

    hdr_row, sheath_uuid_col = sheath_pos

    # Determine the extent of the SHEATH UUID block (it may span multiple
    # blank columns to the right).
    last_col = sheath_uuid_col
    for c in range(sheath_uuid_col + 1, min(ws.max_column, 160) + 1):
        if ws.cell(hdr_row, c).value and str(ws.cell(hdr_row, c).value).strip():
            break   # hit a non-blank header; stop expanding the clear range
        last_col = c

    _clear_range(ws, hdr_row, ws.max_row, sheath_uuid_col, last_col)
    return True


def _process_sheet_cleanup(ws, name: str) -> Tuple[bool, bool, bool]:
    """
    Perform column shifting, SHEATH UUID clearing, and column trimming for one sheet.

    Shifting: In splitter sheets, certain fiber/buffer metadata values that were
    exported to the right of the CONNECTION column are moved to their intended
    destination (immediately after the BUFFER column). This aligns the metadata
    with the left-side column layout expected by the field technicians.

    Clearing: The SHEATH UUID column in the OPTICAL SPLITTERS section is cleared
    (see _clear_optical_splitter_sheath_uuid).

    Trimming: All columns from SHEATH UUID onward in the main table are deleted,
    removing UUID and internal metadata columns that belong to the export format
    but not to the as-built output.

    Returns (shifted, trimmed, cleared) booleans indicating which actions ran.
    """
    hdr_row = find_header_row(ws)
    if hdr_row is None:
        return (False, False, False)

    conn_col = find_col_by_header(ws, hdr_row, "CONNECTION")
    if conn_col is None:
        return (False, False, False)

    # Find the BUFFER column that immediately follows the CONNECTION column;
    # shifted splitter metadata will be placed starting one column after it.
    buffer_cols  = find_all_cols_by_header(ws, hdr_row, "BUFFER")
    buffer_after = next((c for c in buffer_cols if c > conn_col), conn_col)
    dest_start   = buffer_after + 1

    # Find the SHEATH UUID column to the right of CONNECTION; this determines
    # where to begin trimming.
    sheath_uuid_cols = find_all_cols_by_header(ws, hdr_row, "SHEATH UUID")
    trim_col = next((c for c in sheath_uuid_cols if c > conn_col), None)

    # Locate device-related columns that bracket the metadata to be shifted.
    device_name_col = find_col_by_header(ws, hdr_row, "DEVICE NAME", after_col=conn_col)
    device_uuid_col = find_col_by_header(ws, hdr_row, "DEVICE UUID", after_col=conn_col)

    # Determine whether this sheet has an OPTICAL SPLITTERS section (used to
    # decide whether to run the shifting logic, which is only for splitter sheets).
    has_splitter = any(
        isinstance(ws.cell(r, 1).value, str) and
        "OPTICAL SPLITTERS" in ws.cell(r, 1).value.upper()
        for r in range(1, min(ws.max_row, 400) + 1)
    )

    # -- Shift splitter metadata columns --
    shifted = False
    if has_splitter and device_name_col and device_uuid_col:
        # Columns between conn_col and device_uuid_col that are not part of
        # the standard left-side schema are candidates for relocation.
        keep = {
            "FIBER", "BUFFER", "END ENCLOSURE", "END ENCLOSURE ",
            "START ENCLOSURE", "SHEATH NAME", "SHEATH UUID",
        }
        right_cols = [
            c for c in range(conn_col + 1, device_uuid_col)
            if _norm(ws.cell(hdr_row, c).value) and
            _norm(ws.cell(hdr_row, c).value) not in keep
        ]

        # Determine the active data range using the primary SHEATH UUID column.
        primary   = next((c for c in sheath_uuid_cols if c <= conn_col), 1)
        last_row  = hdr_row
        blank_run = 0
        for r in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(r, primary).value
            if v is None or (isinstance(v, str) and not v.strip()):
                blank_run += 1
                if blank_run >= 5:
                    break   # 5 consecutive blank rows = end of data
            else:
                blank_run = 0
                last_row  = r

        # Shift rows where the device name indicates a splitter type (1x2, 1x32).
        for r in range(hdr_row + 1, last_row + 1):
            dname = ws.cell(r, device_name_col).value
            if not isinstance(dname, str):
                continue
            if "1X2" not in dname.upper() and "1X32" not in dname.upper():
                continue

            # Find the first non-empty cell in the right_cols for this row.
            src_start = next(
                (c for c in right_cols
                 if ws.cell(r, c).value is not None and
                 (not isinstance(ws.cell(r, c).value, str) or
                  ws.cell(r, c).value.strip())),
                None,
            )
            if src_start is None:
                continue

            src_end = device_uuid_col - 1
            for i in range(src_end - src_start + 1):
                dst_c = dest_start + i
                # Stop if we would write into or past the trim column.
                if trim_col and dst_c >= trim_col:
                    break
                ws.cell(r, dst_c).value = ws.cell(r, src_start + i).value
            # Clear the source cells after copying.
            for c in range(src_start, src_end + 1):
                ws.cell(r, c).value = None
            shifted = True

    # -- Clear SHEATH UUID in the splitter sub-table --
    cleared = _clear_optical_splitter_sheath_uuid(ws) if has_splitter else False

    # -- Trim columns from SHEATH UUID onward --
    trimmed = False
    if trim_col:
        ws.delete_cols(trim_col, ws.max_column - trim_col + 1)
        trimmed = True

    return (shifted, trimmed, cleared)


def run_part_b(wb) -> None:
    """
    Run Part B cleanup for every qualifying location sheet in the workbook.
    """
    for name in wb.sheetnames:
        if not is_location_sheet(name):
            continue
        ws = wb[name]
        _process_sheet_cleanup(ws, name)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if not INPUT_FILE.exists():
        raise SystemExit(f"Missing input: {INPUT_FILE}")

    wb = openpyxl.load_workbook(str(INPUT_FILE))

    # Part A: enclosure labels and special-case row insertions.
    run_part_a(wb)

    # Part B: column shifting and trimming.
    run_part_b(wb)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    print(f"Step 8 complete: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()

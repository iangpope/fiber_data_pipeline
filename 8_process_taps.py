"""Step 8: Process all tap sheets — reorder sheath blocks, shift port columns,
and label the enclosure type in cell B3.

Replaces three separate scripts:
  - 8_reorder_sheaths_detect_M_or_N.py
  - 9_shift_ports_preserve_all_ports.py
  - 10_label_b3_final_shifted_fixed_cols.py

Reads:  output/Combined_Formatted_Output_with_Addresses.xlsx
Writes: output/Combined_Reordered_With_OTE.xlsx
"""

from __future__ import annotations

import re
from copy import copy
from typing import List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import naming_utils

INPUT_FILE  = "output/Combined_Formatted_Output_with_Addresses.xlsx"
OUTPUT_FILE = "output/Combined_Reordered_With_OTE.xlsx"

PORT_RX = re.compile(r"^\s*PORT\s*\d+\s*$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    return str(s).strip() if s is not None else ""


def is_port_label(value) -> bool:
    s = _norm(value).upper().replace(" ", "")
    return s.startswith("PORT") and len(s) > 4 and s[4].isdigit()


def find_header_row(ws, max_scan_rows: int = 60) -> int | None:
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "CONNECTION":
                return r
    return None


def detect_port_col(ws, header_row: int | None = None) -> int | None:
    """Find the PORT NAME column, searching the header row first then fallback scan."""
    if header_row:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip().upper() == "PORT NAME":
                return c
    # Fallback: scan likely columns
    for c in [13, 14, 15, 16, 17]:
        if c > ws.max_column:
            continue
        for r in range((header_row or 1) + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper().startswith("PORT"):
                return c
    return None


def copy_cell(src, dst) -> None:
    dst.value = src.value
    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.font = copy(src.font)
    dst.protection = copy(src.protection)
    dst.comment = src.comment


# ---------------------------------------------------------------------------
# Stage A: Reorder sheath blocks so PORT blocks come first
# ---------------------------------------------------------------------------

def _build_sheath_blocks(ws, start_row: int) -> List[Tuple[int, int, str]]:
    SHEATH_COL = 2
    blocks: List[Tuple[int, int, str]] = []
    current_name = None
    block_start = None

    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(r, SHEATH_COL).value
        name = name.strip() if isinstance(name, str) else None

        if name is None and current_name is None:
            continue
        if current_name is None and name is not None:
            current_name = name
            block_start = r
            continue
        if name is None:
            continue
        if current_name is not None and name != current_name:
            blocks.append((block_start, r - 1, current_name))
            current_name = name
            block_start = r

    if current_name is not None and block_start is not None:
        blocks.append((block_start, ws.max_row, current_name))

    return [b for b in blocks if b[1] >= b[0]]


def _block_has_ports(ws, block: Tuple[int, int, str], port_col: int) -> bool:
    start, end, _ = block
    for r in range(start, end + 1):
        v = ws.cell(r, port_col).value
        if isinstance(v, str) and v.strip().upper().startswith("PORT"):
            return True
    return False


def _snapshot_rows(ws, start: int, end: int):
    max_col = ws.max_column
    rows = []
    row_dims = []
    for r in range(start, end + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            row_cells.append({
                "value": cell.value,
                "_style": copy(cell._style),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment),
                "comment": cell.comment,
                "hyperlink": cell.hyperlink,
            })
        rows.append(row_cells)
        rd = ws.row_dimensions[r]
        row_dims.append({"height": rd.height, "hidden": rd.hidden, "outlineLevel": rd.outlineLevel})
    return rows, row_dims


def _write_rows(ws, start_row: int, rows, row_dims) -> None:
    max_col = ws.max_column
    for i, row_cells in enumerate(rows):
        r = start_row + i
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            snap = row_cells[c - 1]
            cell.value = snap["value"]
            cell._style = copy(snap["_style"])
            cell.number_format = snap["number_format"]
            cell.protection = copy(snap["protection"])
            cell.alignment = copy(snap["alignment"])
            cell.comment = snap["comment"]
            cell.hyperlink = snap["hyperlink"]
        rd = ws.row_dimensions[r]
        dim = row_dims[i]
        rd.height = dim["height"]
        rd.hidden = dim["hidden"]
        rd.outlineLevel = dim["outlineLevel"]


def reorder_sheath_blocks(ws) -> bool:
    """Move PORT-containing sheath blocks to appear before non-PORT blocks."""
    header_row = find_header_row(ws)
    port_col = detect_port_col(ws, header_row)
    if port_col is None:
        return False

    blocks = _build_sheath_blocks(ws, (header_row or 10) + 1)
    if len(blocks) < 2:
        return False

    scored = [(0 if _block_has_ports(ws, b, port_col) else 1, idx, b) for idx, b in enumerate(blocks)]
    scored_sorted = sorted(scored, key=lambda t: (t[0], t[1]))
    if [t[2] for t in scored_sorted] == [t[2] for t in scored]:
        return False  # already in order

    region_start = min(b[0] for b in blocks)
    block_snaps = [(_snapshot_rows(ws, b[0], b[1]), b[1] - b[0] + 1) for _, _, b in scored]

    write_ptr = region_start
    for _, orig_idx, _ in scored_sorted:
        (rows, dims), height = block_snaps[orig_idx]
        _write_rows(ws, write_ptr, rows, dims)
        write_ptr += height
    return True


# ---------------------------------------------------------------------------
# Stage B: Shift PORT NAME/WAVELENGTH/DEVICE NAME columns to J/K/L
# ---------------------------------------------------------------------------

def shift_port_columns(ws) -> int:
    """Move PORT rows to cols J/K/L (10/11/12) and clear col M onward."""
    header_row = find_header_row(ws)
    port_name_col = detect_port_col(ws, header_row=header_row)

    if not port_name_col:
        return 0

    max_row = ws.max_row
    max_col = ws.max_column
    target_start_col = 10  # J
    shifted_rows = 0

    for r in range((header_row or 1) + 1, max_row + 1):
        if is_port_label(ws.cell(r, port_name_col).value):
            shifted_rows += 1
            for i in range(3):  # PORT NAME, PORT WAVELENGTH, DEVICE NAME
                copy_cell(ws.cell(r, port_name_col + i), ws.cell(r, target_start_col + i))

    if shifted_rows:
        for r in range(1, max_row + 1):
            for c in range(13, max_col + 1):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)

    return shifted_rows


# ---------------------------------------------------------------------------
# Stage C: Label B3 with PORT OTE enclosure type
# ---------------------------------------------------------------------------

def _detect_port_col_after_shift(sheet) -> int | None:
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        for col in range(8, 16):
            if len(row) >= col:
                v = row[col - 1].value
                if v and str(v).strip().upper().startswith("PORT"):
                    return col
    return None


def _count_ports(sheet, port_col: int) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row):
        if len(row) >= port_col:
            val = row[port_col - 1].value
            if val and str(val).strip().upper().startswith("PORT"):
                count += 1
    return count


def label_port_ote(sheet, count: int) -> None:
    if count <= 2:
        sheet["B3"] = "2 PORT OTE"
    elif count <= 4:
        sheet["B3"] = "4 PORT OTE"
    elif count <= 8:
        sheet["B3"] = "8 PORT OTE"
    elif count <= 12:
        sheet["B3"] = "12 PORT OTE"
    sheet["B4"] = 1  # taps always 1 tray


def label_enclosure(ws) -> bool:
    port_col = _detect_port_col_after_shift(ws)
    if not port_col:
        return False
    count = _count_ports(ws, port_col)
    if count > 0:
        label_port_ote(ws, count)
        return True
    return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    wb = load_workbook(INPUT_FILE)
    reordered = shifted = labeled = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if not naming_utils.is_tap(name, df=None):
            continue

        if reorder_sheath_blocks(ws):
            reordered += 1

        n = shift_port_columns(ws)
        if n:
            shifted += 1

        if label_enclosure(ws):
            labeled += 1

    wb.save(OUTPUT_FILE)
    print(f"✅ Tap processing complete: {OUTPUT_FILE}")
    print(f"   • sheets reordered : {reordered}")
    print(f"   • sheets shifted   : {shifted}")
    print(f"   • sheets labeled   : {labeled}")


if __name__ == "__main__":
    main()

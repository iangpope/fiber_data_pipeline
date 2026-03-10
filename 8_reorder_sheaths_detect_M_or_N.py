"""8_reorder_sheaths_detect_M_or_N.py

Reorder *tap* sheets so the sheath block that contains the yellow PORT rows
(i.e., the block with PORT labels in the PORT NAME column) appears first.

Fixes two issues in the older version:
  1) It only processed legacy sheet names (MIC...). This build also uses
     names like RC077E_FT_001.
  2) It tried to find "yellow section starts" using column A/B fills, which
     fails on many vendor formats (RC077E is a good example). We now detect
     blocks by changes in the sheath-name column (col B) and then rank blocks
     by whether they contain PORT rows.

Design notes:
  - We preserve full cell formatting (style, font, border, alignment, number
    format, hyperlink, comments) when swapping blocks.
  - We only touch sheets that are clearly taps: B3 contains "PORT OTE".
"""

from __future__ import annotations

from copy import copy
from typing import List, Tuple, Optional

from openpyxl import load_workbook


# File paths (relative to project root)
INPUT_FILE = "output/Combined_Formatted_Output_with_Addresses.xlsx"
OUTPUT_FILE = "output/Combined_Reordered_With_OTE.xlsx"


import re

legacy_tap_rx = re.compile(r"^MIC[A-Z]{2}\d{2}\d{4}$", re.IGNORECASE)


def is_tap_sheet(sheet_name: str, ws) -> bool:
    """Robust tap detection across naming conventions.

    NOTE: This runs *before* B3 is labeled (script 10), so we can't rely on B3.
    """
    name_up = sheet_name.upper()
    if "_FT_" in name_up or name_up.endswith("_FT"):
        return True
    if legacy_tap_rx.match(name_up):
        return True
    b3 = ws["B3"].value
    if isinstance(b3, str) and "PORT OTE" in b3.upper():
        return True
    # last-resort: PORT NAME header exists
    header_row = find_header_row(ws)
    return detect_port_col(ws, header_row) is not None


def find_header_row(ws) -> int:
    """Find the header row containing the word 'CONNECTION'. Defaults to 11."""
    for r in range(1, 40):
        for c in range(1, 30):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() == "CONNECTION":
                return r
    return 11


def detect_port_col(ws, header_row: int) -> Optional[int]:
    """Detect PORT NAME column by header text; fallback to scanning for PORT labels."""
    # Prefer explicit header "PORT NAME"
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().upper() == "PORT NAME":
            return c

    # Fallback: scan a few likely columns for PORT1/PORT2 values
    likely = [13, 14, 15, 16, 17]  # M..Q-ish
    for c in likely:
        if c > ws.max_column:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper().startswith("PORT"):
                return c
    return None


def first_data_row(ws, header_row: int) -> int:
    """Data starts immediately after header row in these cut sheets."""
    return header_row + 1


def build_sheath_blocks(ws, start_row: int) -> List[Tuple[int, int, str]]:
    """Return contiguous blocks by sheath-name changes in column B.

    Each tuple: (start_row, end_row, sheath_name)
    """
    SHEATH_COL = 2  # column B
    blocks: List[Tuple[int, int, str]] = []

    current_name = None
    block_start = None

    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(row=r, column=SHEATH_COL).value
        name = name.strip() if isinstance(name, str) else None

        # Stop if we hit a long blank tail (common) – but be conservative.
        if name is None and current_name is None:
            continue

        if current_name is None and name is not None:
            current_name = name
            block_start = r
            continue

        if name is None:
            # stay within the current block
            continue

        if current_name is not None and name != current_name:
            # close current block
            blocks.append((block_start, r - 1, current_name))
            current_name = name
            block_start = r

    if current_name is not None and block_start is not None:
        blocks.append((block_start, ws.max_row, current_name))

    # Filter out tiny/degenerate blocks
    blocks = [b for b in blocks if b[1] >= b[0]]
    return blocks


def block_has_ports(ws, block: Tuple[int, int, str], port_col: int) -> bool:
    start, end, _ = block
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=port_col).value
        if isinstance(v, str) and v.strip().upper().startswith("PORT"):
            return True
    return False


def snapshot_rows(ws, start: int, end: int):
    """Snapshot values + formatting for rows [start, end]."""
    max_col = ws.max_column
    rows = []
    row_dims = []

    for r in range(start, end + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_cells.append({
                "value": cell.value,
                "_style": copy(cell._style),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment),
                "comment": cell.comment,
                "hyperlink": cell.hyperlink,
            })
        rows.append(row_cells)

        # row dimension (height etc.)
        rd = ws.row_dimensions[r]
        row_dims.append({
            "height": rd.height,
            "hidden": rd.hidden,
            "outlineLevel": rd.outlineLevel,
        })

    return rows, row_dims


def write_rows(ws, start_row: int, rows, row_dims):
    max_col = ws.max_column
    for i, row_cells in enumerate(rows):
        r = start_row + i
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            snap = row_cells[c - 1]
            cell.value = snap["value"]
            cell._style = copy(snap["_style"])
            cell.number_format = snap["number_format"]
            cell.protection = copy(snap["protection"])
            cell.alignment = copy(snap["alignment"])
            cell.comment = snap["comment"]
            cell.hyperlink = snap["hyperlink"]

        # restore row dimensions
        rd = ws.row_dimensions[r]
        dim = row_dims[i]
        rd.height = dim["height"]
        rd.hidden = dim["hidden"]
        rd.outlineLevel = dim["outlineLevel"]


def reorder_blocks_in_place(ws, blocks, port_col: int):
    """If a later block contains PORT rows and earlier blocks do not, move it first."""
    if len(blocks) < 2:
        return

    scored = []
    for idx, b in enumerate(blocks):
        scored.append((0 if block_has_ports(ws, b, port_col) else 1, idx, b))

    # stable sort: PORT blocks first, then original order
    scored_sorted = sorted(scored, key=lambda t: (t[0], t[1]))

    # If order unchanged, do nothing
    if [t[2] for t in scored_sorted] == [t[2] for t in scored]:
        return

    # Snapshot the entire region covering all blocks, then write back in new order
    region_start = min(b[0] for b in blocks)
    region_end = max(b[1] for b in blocks)

    # Snapshot each original block
    block_snaps = []
    for _, _, b in scored:
        rows, dims = snapshot_rows(ws, b[0], b[1])
        block_snaps.append((rows, dims, b[1] - b[0] + 1))

    # Write blocks back in sorted order
    write_ptr = region_start
    for _, orig_idx, _b in scored_sorted:
        rows, dims, height = block_snaps[orig_idx]
        write_rows(ws, write_ptr, rows, dims)
        write_ptr += height

    # Clear any leftover rows if new ordering uses fewer rows (shouldn't happen)
    # but keep as a safety.
    for r in range(write_ptr, region_end + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None


def main():
    wb = load_workbook(INPUT_FILE)
    changed = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if not is_tap_sheet(name, ws):
            continue

        header_row = find_header_row(ws)
        port_col = detect_port_col(ws, header_row)
        if port_col is None:
            continue

        blocks = build_sheath_blocks(ws, first_data_row(ws, header_row))
        if len(blocks) < 2:
            continue

        before = [b[2] for b in blocks]
        reorder_blocks_in_place(ws, blocks, port_col)
        # recompute to check if changed
        blocks_after = build_sheath_blocks(ws, first_data_row(ws, header_row))
        after = [b[2] for b in blocks_after]
        if after != before:
            changed += 1

    wb.save(OUTPUT_FILE)
    print(f"✅ Reordered sheath blocks on {changed} tap sheets.")


if __name__ == "__main__":
    main()

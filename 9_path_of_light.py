"""
9_path_of_light.py -- PON Continuity Tracer: verify every tap PORT has a
complete optical path back to the OLT.

For each tap (FT) sheet and each PORT on that tap, this script traces the
signal path backward through the splice network:

  PORT on tap
    -> cable arriving at the tap (col B, SHEATHS section, PORT row)
    -> SE enclosure sheet (the start enclosure listed on that cable row)
    -> OPTICAL SPLITTERS sub-table: find the Out-N row whose sheath name
       matches the arriving cable, then find the Common row for the same
       splitter device (same device UUID)
    -> Common row sheath name = new cable; follow it to the next node
    -> Repeat until OLT sheet is reached (or error is detected)

Results are written to: output/Path_of_Light_Confirmation.xlsx
  - One row per (tap, port)
  - Green fill = traced successfully to OLT
  - Red fill   = trace broke; Status column explains where and why

Reads:  output/Asbuilt_Workbook_post12.xlsx
Writes: output/Path_of_Light_Confirmation.xlsx
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
INPUT_FILE  = Path("output") / "Asbuilt_Workbook_post12.xlsx"
OUTPUT_FILE = Path("output") / "Path_of_Light_Confirmation.xlsx"

# Maximum hops before declaring a circular or runaway trace.
MAX_HOPS = 40


# OLT token pattern: bare site ID like "RC73E" (no _FT_ or _SE_ suffix).
OLT_RX = re.compile(r"^[A-Z0-9]{2,10}E$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Report styling
# ---------------------------------------------------------------------------
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Calibri", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Workbook parsing helpers
# ---------------------------------------------------------------------------

def _v(ws, row: int, col: int) -> str:
    """Return the stripped string value of a cell, or empty string."""
    val = ws.cell(row, col).value
    return str(val).strip() if val is not None else ""


def _find_col(ws, header_row: int, label: str, max_col: int = 20) -> int | None:
    """Return the 1-based column index where the header row cell matches label."""
    label_up = label.strip().upper()
    for c in range(1, max_col + 1):
        if _v(ws, header_row, c).upper() == label_up:
            return c
    return None


def _find_sheaths_header(ws) -> int | None:
    """
    Return the row number of the SHEATHS data header row (the row that contains
    'CONNECTION' in any cell), scanning the first 120 rows.
    """
    for r in range(1, min(ws.max_row, 120) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            if _v(ws, r, c).upper() == "CONNECTION":
                return r
    return None


def _find_optical_splitters_row(ws) -> int | None:
    """Return the row number of the 'OPTICAL SPLITTERS' section header."""
    for r in range(1, min(ws.max_row, 1000) + 1):
        if "OPTICAL SPLITTERS" in _v(ws, r, 1).upper():
            return r
    return None


# ---------------------------------------------------------------------------
# Step 1: Extract PORT rows from a tap sheet
#
# Returns list of dicts: {port_label, cable_name, buffer, fiber}
#
# The PORT row in the tap sheet has:
#   col B  = cable name (the arriving cable from the SE)
#   col C  = start enclosure (the tap itself)
#   col D  = end enclosure  (the SE this cable comes from)
#   col G  = CONNECTION  (<---> means fused = this fiber is live)
#   col J  = PORT NAME (e.g. PORT1, PORT 1, PORT1)
# ---------------------------------------------------------------------------

def get_tap_ports(ws) -> list[dict]:
    """
    Scan the SHEATHS section of a tap sheet and return one entry per PORT row.

    A PORT row has `<--->` in the CONNECTION column and a PORT label in the
    right-side PORT NAME column (col J after the step-7 shift). Only rows
    where CONNECTION is `<--->` (live/fused) are considered — rows with `X`
    (cut/unused) are skipped since they carry no signal.
    """
    hdr = _find_sheaths_header(ws)
    if hdr is None:
        return []

    # Determine column indices from the header row.
    col_sheath_name   = _find_col(ws, hdr, "SHEATH NAME", max_col=16)
    col_start_enc     = _find_col(ws, hdr, "START ENCLOSURE", max_col=16)
    col_end_enc       = _find_col(ws, hdr, "END ENCLOSURE", max_col=16)
    col_buffer        = _find_col(ws, hdr, "BUFFER", max_col=16)
    col_fiber         = _find_col(ws, hdr, "FIBER", max_col=16)
    col_connection    = _find_col(ws, hdr, "CONNECTION", max_col=16)

    # Port NAME is always written to col J (10) by step 7.
    col_port_name = 10

    if not all([col_sheath_name, col_start_enc, col_end_enc,
                col_buffer, col_fiber, col_connection]):
        return []

    opt_row = _find_optical_splitters_row(ws)
    end_row = (opt_row - 1) if opt_row else ws.max_row

    ports = []
    for r in range(hdr + 1, end_row + 1):
        port_val = _v(ws, r, col_port_name)
        conn_val = _v(ws, r, col_connection)

        # A PORT row: port label present AND connection is fused (<--->).
        if port_val.upper().startswith("PORT") and "<" in conn_val:
            cable = _v(ws, r, col_sheath_name)
            # The tap is the START ENCLOSURE; the SE is the END ENCLOSURE
            # (from the tap's perspective its local side is START).
            start_enc = _v(ws, r, col_start_enc)
            end_enc   = _v(ws, r, col_end_enc)

            # The SE is whichever endpoint is not this tap.
            tap_name  = ws.title
            next_node = end_enc if start_enc == tap_name else start_enc

            ports.append({
                "port":      port_val,
                "cable":     cable,
                "next_node": next_node,
                "buffer":    _v(ws, r, col_buffer),
                "fiber":     _v(ws, r, col_fiber),
            })

    return ports


# ---------------------------------------------------------------------------
# Step 2: Build the OPTICAL SPLITTERS index for an SE sheet
#
# Returns: dict  cable_sheath_name -> common_cable_sheath_name
#
# For each Out-N row in the OPTICAL SPLITTERS table, look up the COMMON row
# for the same device (same DEVICE UUID). The SHEATH NAME on the COMMON row
# is the backhaul cable that carries the signal upstream.
#
# OPTICAL SPLITTERS columns (0-indexed from the sub-table header):
#   col A (1) = DEVICE UUID
#   col B (2) = DEVICE NAME
#   col C (3) = PORT NAME  (Common, Out-1, Out-2, ...)
#   col D (4) = CONNECTION
#   col G (7) = FIBER
#   col H (8) = BUFFER
#   col I (9) = END ENCLOSURE
#   col J (10)= START ENCLOSURE
#   col K (11)= SHEATH NAME
# ---------------------------------------------------------------------------

def build_splitter_map(ws) -> dict[str, str]:
    """
    Parse the OPTICAL SPLITTERS sub-table and return a dict mapping:
        output cable sheath name  ->  common (input) cable sheath name

    Handles both simple and cascaded/chained splitter configurations:
      - Simple:  Out-N row has a cable sheath name -> look up same device COMMON cable
      - Chained: Out-N row has no cable (empty sheath) but the right-side columns
                 show another device name in col E (right PORT NAME) and col F
                 (right DEVICE NAME). This means the Out-N port of one splitter
                 is wired directly to the COMMON port of another splitter without
                 an intermediate cable. In this case we follow the chain to find
                 the outermost COMMON cable.

    Sub-table column layout (1-based):
      col A (1)  = DEVICE UUID
      col B (2)  = DEVICE NAME
      col C (3)  = PORT NAME left  (Common, Out-1, Out-2, ...)
      col D (4)  = CONNECTION
      col E (5)  = PORT NAME right (for chained devices: e.g. 'Common')
      col F (6)  = DEVICE NAME right (for chained devices)
      col G (7)  = FIBER
      col H (8)  = BUFFER
      col I (9)  = END ENCLOSURE
      col J (10) = START ENCLOSURE
      col K (11) = SHEATH NAME
    """
    opt_row = _find_optical_splitters_row(ws)
    if opt_row is None:
        return {}

    hdr = opt_row + 1   # sub-table header row

    col_uuid        = _find_col(ws, hdr, "DEVICE UUID",  max_col=16)
    col_dev_name    = _find_col(ws, hdr, "DEVICE NAME",  max_col=16)
    col_port        = _find_col(ws, hdr, "PORT NAME",    max_col=16)
    col_conn        = _find_col(ws, hdr, "CONNECTION",   max_col=16)
    col_sheath      = _find_col(ws, hdr, "SHEATH NAME",  max_col=16)

    if not all([col_uuid, col_port, col_conn, col_sheath]):
        return {}

    # Right-side columns relative to CONNECTION (col D):
    #   col E = right PORT NAME  (chained: e.g. 'Common')
    #   col F = right DEVICE NAME (chained: the upstream splitter device name)
    col_r_port_name = col_conn + 1 if col_conn else None   # col E
    col_r_dev_name  = col_conn + 2 if col_conn else None   # col F


    # -----------------------------------------------------------------------
    # First pass: collect all OPTICAL SPLITTERS rows into a per-device dict.
    #
    # For each device (UUID) we record:
    #   common_cable     : cable sheath name on the COMMON port (str or None)
    #   common_from_dev  : device name that feeds this device's COMMON (chained case)
    #   outputs          : list of cable sheath names on active Out-N ports
    # -----------------------------------------------------------------------
    devices: dict[str, dict] = {}          # uuid -> device data
    name_to_uuid: dict[str, str] = {}      # device_name -> uuid

    for r in range(hdr + 1, ws.max_row + 1):
        uuid     = _v(ws, r, col_uuid)
        dev_name = _v(ws, r, col_dev_name)
        port     = _v(ws, r, col_port).upper()
        conn     = _v(ws, r, col_conn)
        sheath   = _v(ws, r, col_sheath)

        if not uuid:
            continue   # blank uuid = end of sub-table

        if uuid not in devices:
            devices[uuid] = {
                "name":           dev_name,
                "common_cable":   None,
                "common_from":    None,   # device name of upstream chained device
                "outputs":        [],
            }
            if dev_name:
                name_to_uuid[dev_name] = uuid

        if "COMMON" in port and "<" in conn:
            if sheath:
                # Normal case: COMMON port connects via a real cable.
                devices[uuid]["common_cable"] = sheath
            else:
                # Chained case: COMMON port is fed by another device's output.
                # The feeding device name appears in the right DEVICE NAME column.
                r_dev = _v(ws, r, col_r_dev_name) if col_r_dev_name else ""
                if r_dev:
                    devices[uuid]["common_from"] = r_dev

        elif port.startswith("OUT") and "<" in conn:
            if sheath:
                # Normal output: wire goes to a cable sheath.
                devices[uuid]["outputs"].append(sheath)
            # If no sheath, this Out-N feeds directly into another device —
            # we don't need to record it here because the downstream device's
            # COMMON row (above) will already have recorded that relationship.

    # -----------------------------------------------------------------------
    # Second pass: build output_cable -> upstream_common_cable mapping.
    #
    # For each device that has output cable sheaths:
    #   Walk UP the chain (via common_from) until we find a device with a
    #   real common_cable. That is the PON backhaul cable.
    # -----------------------------------------------------------------------
    def resolve_common(uuid: str, visited: set) -> str | None:
        """Recursively follow chained device links to find the upstream cable."""
        if uuid in visited:
            return None   # cycle guard
        visited.add(uuid)
        data = devices.get(uuid)
        if data is None:
            return None
        if data["common_cable"]:
            return data["common_cable"]
        # Chained: follow to the upstream device.
        upstream_name = data.get("common_from")
        if not upstream_name:
            return None
        upstream_uuid = name_to_uuid.get(upstream_name)
        if not upstream_uuid:
            return None
        return resolve_common(upstream_uuid, visited)

    output_to_common: dict[str, str] = {}
    for uuid, data in devices.items():
        upstream_cable = resolve_common(uuid, set())
        if not upstream_cable:
            continue
        for out_cable in data["outputs"]:
            output_to_common[out_cable] = upstream_cable

    return output_to_common



# ---------------------------------------------------------------------------
# Core trace function
# ---------------------------------------------------------------------------

def trace_port(wb, tap_name: str, port_info: dict) -> dict:
    """
    Trace one PORT's optical path backward from the tap to the OLT.

    Starting from the cable arriving at the tap PORT, follows the fiber through
    each intermediate SE enclosure by crossing through the OPTICAL SPLITTERS
    table until the OLT sheet is reached, a trace error occurs, or MAX_HOPS
    is exceeded (which would indicate a loop or corrupt design data).

    Returns a dict with:
        ok        : bool
        path      : list of node names visited (tap -> SE -> ... -> OLT)
        status    : descriptive result string
        fail_node : name of the node where the trace broke (if not ok)
    """
    path      = [tap_name]
    ok        = False
    fail_node = None
    status    = ""

    current_cable     = port_info["cable"]
    current_next_node = port_info["next_node"]

    if not current_cable:
        return {
            "ok": False, "path": path,
            "status": "No cable found on PORT row in tap sheet",
            "fail_node": tap_name,
        }

    for hop in range(MAX_HOPS):
        next_name = current_next_node.strip()

        if not next_name:
            status    = f"Cable '{current_cable}' has no destination node listed"
            fail_node = path[-1]
            break

        # If the next node is not in the workbook, it may be the OLT rack sheet.
        if next_name not in wb.sheetnames:
            # Check if it looks like a bare OLT token (reached the OLT rack).
            if OLT_RX.match(next_name):
                path.append(next_name)
                ok     = True
                status = "OK"
            else:
                status    = f"Sheet '{next_name}' not found in workbook"
                fail_node = path[-1]
            break

        ws = wb[next_name]
        path.append(next_name)

        # Check for bare OLT sheet (no FT/SE suffix — this is the OLT rack sheet).
        if OLT_RX.match(next_name):
            ok     = True
            status = "OK"
            break

        # This node is an SE enclosure. Build its splitter cable map.
        splitter_map = build_splitter_map(ws)

        if not splitter_map:
            # SE sheet has no OPTICAL SPLITTERS section — may be a distribution
            # or pass-through enclosure. In that case look for the cable on the
            # main SHEATHS section and follow the CONNECTION across to the other side.
            result = _trace_through_sheaths(wb, ws, current_cable, path)
            if result is None:
                status    = f"No splitter table and cable '{current_cable}' not found in SHEATHS at '{next_name}'"
                fail_node = next_name
                break
            current_cable, current_next_node = result
            continue

        # Look up this cable in the splitter output map.
        # NOTE: An SE enclosure can have both an OPTICAL SPLITTERS section AND
        # cables that simply pass through its SHEATHS (e.g. a feeder ring node
        # that hosts a splitter for some taps but acts as a pass-through for
        # other cable runs going further along the route). If the cable is not
        # found as a splitter output, fall back to the SHEATHS pass-through logic.
        if current_cable not in splitter_map:
            result = _trace_through_sheaths(wb, ws, current_cable, path)
            if result is None:
                status    = f"Cable '{current_cable}' not found as splitter output or SHEATHS pass-through at '{next_name}'"
                fail_node = next_name
                break
            current_cable, current_next_node = result
            continue


        # Found it — get the upstream backhaul cable via the COMMON port.
        common_cable = splitter_map[current_cable]

        # Now find where the COMMON cable goes (its other endpoint).
        next_node_from_common = _find_cable_far_end(ws, common_cable, next_name)
        if not next_node_from_common:
            status    = f"Common cable '{common_cable}' found at '{next_name}' but far-end node cannot be determined"
            fail_node = next_name
            break

        current_cable     = common_cable
        current_next_node = next_node_from_common

    else:
        # Exceeded MAX_HOPS without reaching OLT.
        status    = f"Trace exceeded {MAX_HOPS} hops — possible loop in design data"
        fail_node = path[-1]

    return {"ok": ok, "path": path, "status": status, "fail_node": fail_node}


def _find_cable_far_end(ws, cable_name: str, current_node: str) -> str | None:
    """
    Given a cable name and the current node (SE enclosure), find the cable's
    far-end node by scanning the SHEATHS section for the cable name and reading
    the endpoint that is NOT the current node.

    Returns the far-end node name string, or None if not found.
    """
    hdr = _find_sheaths_header(ws)
    if hdr is None:
        return None

    col_sheath    = _find_col(ws, hdr, "SHEATH NAME",     max_col=16)
    col_start_enc = _find_col(ws, hdr, "START ENCLOSURE", max_col=16)
    col_end_enc   = _find_col(ws, hdr, "END ENCLOSURE",   max_col=16)

    if not all([col_sheath, col_start_enc, col_end_enc]):
        return None

    opt_row = _find_optical_splitters_row(ws)
    end_row = (opt_row - 1) if opt_row else ws.max_row

    for r in range(hdr + 1, end_row + 1):
        if _v(ws, r, col_sheath) == cable_name:
            start = _v(ws, r, col_start_enc)
            end   = _v(ws, r, col_end_enc)
            # Return whichever endpoint is not the current node.
            return end if start == current_node else start

    return None


def _trace_through_sheaths(wb, ws, cable_name: str, path: list) -> tuple | None:
    """
    For a distribution/pass-through SE enclosure (no OPTICAL SPLITTERS table),
    find the given cable in the SHEATHS section, cross through the CONNECTION
    to the fused fiber on the other side, and return the (next_cable, next_node)
    to continue the trace.

    Returns (cable_name, next_node) or None if the cable cannot be found.
    """
    hdr = _find_sheaths_header(ws)
    if hdr is None:
        return None

    col_sheath       = _find_col(ws, hdr, "SHEATH NAME",      max_col=16)
    col_start_enc    = _find_col(ws, hdr, "START ENCLOSURE",  max_col=16)
    col_end_enc      = _find_col(ws, hdr, "END ENCLOSURE",    max_col=16)
    col_connection   = _find_col(ws, hdr, "CONNECTION",       max_col=16)
    # Right-side columns mirror the left-side layout; col L (12) = right SHEATH NAME
    col_r_sheath     = 12   # right-side SHEATH NAME (post step-5 column layout)
    col_r_start_enc  = 10   # right-side START ENCLOSURE
    col_r_end_enc    = 11   # right-side END ENCLOSURE (sometimes swapped; check both)

    if not all([col_sheath, col_start_enc, col_end_enc, col_connection]):
        return None

    opt_row = _find_optical_splitters_row(ws)
    end_row = (opt_row - 1) if opt_row else ws.max_row

    current_node = ws.title

    for r in range(hdr + 1, end_row + 1):
        sheath = _v(ws, r, col_sheath)
        conn   = _v(ws, r, col_connection)

        if sheath != cable_name:
            continue
        if "<" not in conn:
            continue  # X = unused fiber; skip

        # Crossed the connection — read the right-side cable.
        r_sheath = _v(ws, r, col_r_sheath)
        r_start  = _v(ws, r, col_r_start_enc)
        r_end    = _v(ws, r, col_r_end_enc)

        if not r_sheath:
            return None

        next_node = r_end if r_start == current_node else r_start
        return (r_sheath, next_node)

    return None


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def _write_report(results: list[dict], output_path: Path) -> None:
    """
    Write the Path of Light Confirmation report to an Excel file.

    Each row corresponds to one (tap, port) trace. Green = reached OLT.
    Red = trace failed. The Status column contains 'OK' or a description
    of the failure point.
    """
    wb_out = Workbook()
    ws     = wb_out.active
    ws.title = "Path of Light"

    # -- Header row --
    headers = [
        "Tap", "Port", "Status", "Hops", "Path",
        "Cable into Tap", "Failed At Node",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = CELL_BORDER

    ws.row_dimensions[1].height = 20

    # -- Data rows --
    for row_idx, r in enumerate(results, start=2):
        fill  = GREEN_FILL if r["ok"] else RED_FILL
        path_str = " -> ".join(r["path"])

        values = [
            r["tap"],
            r["port"],
            r["status"],
            len(r["path"]) - 1,   # number of hops (tap doesn't count)
            path_str,
            r["cable"],
            r.get("fail_node", ""),
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill      = fill
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c > 2 else CENTER
            cell.border    = CELL_BORDER

        ws.row_dimensions[row_idx].height = 16

    # -- Column widths --
    widths = [18, 10, 55, 6, 80, 45, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w

    # -- Summary header below the table --
    total = len(results)
    ok    = sum(1 for r in results if r["ok"])
    bad   = total - ok

    ws.freeze_panes = "A2"   # freeze the header row for easy scrolling

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(str(output_path))

    print(f"Path of Light Confirmation: {output_path}")
    print(f"  Total ports traced : {total}")
    print(f"  Passed (OLT found) : {ok}")
    print(f"  Failed             : {bad}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(data_dir: str = "data", output_dir: str = "output") -> None:
    global INPUT_FILE, OUTPUT_FILE
    INPUT_FILE  = Path(output_dir) / "Asbuilt_Workbook_post12.xlsx"
    OUTPUT_FILE = Path(output_dir) / "Path_of_Light_Confirmation.xlsx"
    if not INPUT_FILE.exists():
        raise SystemExit(f"Input file not found: {INPUT_FILE}")

    print(f"Loading: {INPUT_FILE}")
    wb = load_workbook(str(INPUT_FILE))

    # Identify all tap and SE sheets.
    tap_sheets = [s for s in wb.sheetnames if "_FT_" in s.upper()]
    print(f"Found {len(tap_sheets)} tap sheets to trace.")

    results: list[dict] = []

    for tap_name in tap_sheets:
        ws   = wb[tap_name]
        ports = get_tap_ports(ws)

        if not ports:
            # Tap has no PORT rows — flag it as a problem.
            results.append({
                "tap":       tap_name,
                "port":      "—",
                "cable":     "",
                "ok":        False,
                "path":      [tap_name],
                "status":    "No PORT rows found in tap sheet (design may be missing port assignments)",
                "fail_node": tap_name,
            })
            continue

        for port_info in ports:
            trace = trace_port(wb, tap_name, port_info)
            results.append({
                "tap":    tap_name,
                "port":   port_info["port"],
                "cable":  port_info["cable"],
                **trace,
            })

    _write_report(results, OUTPUT_FILE)


if __name__ == "__main__":
    main()

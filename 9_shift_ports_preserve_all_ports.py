import re
from copy import copy
from openpyxl import load_workbook

from openpyxl.styles import PatternFill

import naming_utils

# File paths
INPUT_FILE = "output/Combined_Reordered_With_OTE.xlsx"
OUTPUT_FILE = "output/Combined_Final_Shifted.xlsx"

PORT_RX = re.compile(r"^\s*PORT\s*\d+\s*$", re.IGNORECASE)

def _norm(s):
    return str(s).strip() if s is not None else ""

def is_port_label(value) -> bool:
    s = _norm(value).upper().replace(" ", "")
    # Accept PORT1 / PORT 1 / PORT01
    return s.startswith("PORT") and len(s) > 4 and s[4].isdigit()

def find_header_row_and_connection_col(sheet, max_scan_rows=60, max_scan_cols=40):
    for r in range(1, min(sheet.max_row, max_scan_rows) + 1):
        for c in range(1, min(sheet.max_column, max_scan_cols) + 1):
            v = sheet.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "CONNECTION":
                return r, c
    return None, None

def find_port_name_col(sheet, header_row=None):
    # Prefer header row detection (most reliable)
    if header_row:
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(header_row, c).value
            if isinstance(v, str) and v.strip().upper() == "PORT NAME":
                return c

    # Fallback: find any cell that literally says PORT NAME
    for r in range(1, min(sheet.max_row, 80) + 1):
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "PORT NAME":
                return c
    return None

def copy_cell(src, dst):
    dst.value = src.value
    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.font = copy(src.font)
    dst.protection = copy(src.protection)
    dst.comment = src.comment

def clear_cell(cell):
    cell.value = None
    cell._style = copy(cell._style)  # keep object valid
    cell.fill = PatternFill(fill_type=None)
    cell.font = copy(cell.font)
    cell.border = copy(cell.border)
    cell.alignment = copy(cell.alignment)
    cell.number_format = cell.number_format
    cell.protection = copy(cell.protection)
    cell.comment = None

def shift_ports_and_clear(sheet):
    """
    Goal:
      - On tap sheets, move PORT NAME / PORT WAVELENGTH / DEVICE NAME left into columns J/K/L (10/11/12)
      - Then clear everything from column M (13) to the end of the sheet (removes SHEATH UUID + PORT/DEVICE UUID cols)

    Important:
      - Works whether PORT NAME starts in column M or N (or elsewhere), so taps with 2 sheaths don't get skipped.
    """
    header_row, _ = find_header_row_and_connection_col(sheet)
    port_name_col = find_port_name_col(sheet, header_row=header_row)

    if not port_name_col:
        print(f"  ⚠️  No PORT NAME column found in {sheet.title}")
        return 0

    max_row = sheet.max_row
    max_col = sheet.max_column

    target_start_col = 10  # J
    shifted_rows = 0

    # Shift only the PORT rows
    for r in range((header_row or 1) + 1, max_row + 1):
        if is_port_label(sheet.cell(r, port_name_col).value):
            shifted_rows += 1
            for i in range(3):  # PORT NAME, PORT WAVELENGTH, DEVICE NAME
                src = sheet.cell(r, port_name_col + i)
                dst = sheet.cell(r, target_start_col + i)
                copy_cell(src, dst)

    # If we moved anything, clear col M+ for the whole sheet
    if shifted_rows:
        for r in range(1, max_row + 1):
            for c in range(13, max_col + 1):  # M onward
                cell = sheet.cell(r, c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                # Keep fonts/borders/alignment as-is for left region; right region can be blank

    return shifted_rows

def process_file(input_path, output_path):
    wb = load_workbook(input_path)

    total_shifted_rows = 0
    processed_sheets = 0

    for sheet_name in wb.sheetnames:
        if naming_utils.is_tap(sheet_name):
            processed_sheets += 1
            sheet = wb[sheet_name]
            print(f"Processing {sheet_name}...")
            shifted = shift_ports_and_clear(sheet)
            total_shifted_rows += shifted

    wb.save(output_path)
    print(f"✅ Done. Processed {processed_sheets} tap sheets; shifted {total_shifted_rows} PORT rows total.")

if __name__ == "__main__":
    process_file(INPUT_FILE, OUTPUT_FILE)

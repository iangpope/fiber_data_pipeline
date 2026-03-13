"""
config.py -- Shared constants for the Fiber Data Pipeline.

This is the single source of truth for all colors, fill objects, directory
paths, and skip-sheet lists used across every step in the pipeline. Importing
from here instead of redefining values in each script ensures consistency and
makes future changes (e.g. adjusting a color) a one-line edit in one place.
"""

import os
from openpyxl.styles import PatternFill


# ---------------------------------------------------------------------------
# Directory paths
# All input files live under DATA_DIR; all generated workbooks go to OUTPUT_DIR.
# ---------------------------------------------------------------------------
DATA_DIR   = "data"       # raw inputs: KMZ, cut sheet Excel
OUTPUT_DIR = "output"     # all intermediate and final output workbooks


# ---------------------------------------------------------------------------
# Sheet names to always skip when iterating a workbook.
# These are administrative or reference sheets, not splice location data.
# Stored lowercase so callers compare with name.lower().
# ---------------------------------------------------------------------------
SKIP_SHEETS = {"index", "legend", "notes", "sheet1", "tap report"}


# ---------------------------------------------------------------------------
# Color palette -- 6-character hex strings (no alpha prefix).
#
# Directional colors are assigned based on the compass bearing of each fiber
# cable as it leaves a splice enclosure. The OLT and MST entries have fixed
# colors that are applied regardless of bearing. Splitter/MUX colors are used
# in the OPTICAL SPLITTERS sub-table and the right-side port section.
# ---------------------------------------------------------------------------
COLOR = {
    # Directional sheath colors (applied to the main sheath rows on each sheet)
    "NORTH":  "FFA500",   # orange  -- cable runs northbound
    "SOUTH":  "8B4513",   # brown   -- cable runs southbound
    "EAST":   "008000",   # green   -- cable runs eastbound
    "WEST":   "708090",   # slate   -- cable runs westbound
    "MST":    "FF0000",   # red     -- main sheath terminal (MST tap enclosure)
    "OLT":    "C5D9B5",   # olive   -- optical line terminal connection

    # Splice and connection type markers
    "FUSION": "FFFF00",   # yellow  -- fusion splice midpoint row marker
    "YELLOW": "FFFF00",   # alias for FUSION used in some reverse lookups

    # Splitter and MUX port type colors (used in the OPTICAL SPLITTERS section)
    "COMMON": "ADD8E6",   # light blue  -- 1xN splitter common (input) port
    "1X32":   "FFB6C1",   # light pink  -- 1x32 splitter output port
    "1X2":    "DB7093",   # dark pink   -- 1x2 splitter output port
    "MUX":    "FFDAB9",   # peach       -- MUX channel port
    "DEMUX":  "FFA07A",   # salmon      -- DEMUX channel port
}


# Reverse lookup: hex value -> human-readable direction/type label.
# Used by step 4 (format_top_section) when building the direction key bar.
COLOR_TO_DIRECTION = {
    COLOR["NORTH"]: "North",
    COLOR["SOUTH"]: "South",
    COLOR["EAST"]:  "East",
    COLOR["WEST"]:  "West",
    COLOR["OLT"]:   "OLT",
    COLOR["MST"]:   "MST",
}


# ---------------------------------------------------------------------------
# Pre-built PatternFill cache.
#
# openpyxl PatternFill construction is called thousands of times across large
# splice sheets. Caching fills by hex string here avoids redundant object
# creation inside cell-level loops.
# ---------------------------------------------------------------------------
def _fill(hex6: str) -> PatternFill:
    """Return a solid PatternFill for the given 6-character hex color string."""
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

# Pre-populate the cache with every color the pipeline uses.
FILLS = {hex6: _fill(hex6) for hex6 in {
    "FFA500", "8B4513", "008000", "708090", "FF0000", "C5D9B5",
    "FFFF00", "ADD8E6", "FFB6C1", "DB7093", "FFDAB9", "FFA07A",
    "7FFF00", "FFF9DB",
}}

EMPTY_FILL = PatternFill()  # represents "no fill" / clear cell background


# ---------------------------------------------------------------------------
# Connection column value strings
#
# The raw Magellan export uses verbose markers in the CONNECTION column.
# Step 5 normalizes them to the compact arrow form used in the final workbook.
# Defining them here as constants means a single edit propagates everywhere.
# ---------------------------------------------------------------------------
CONN_RAW_FUSION     = "<- FUSION ->"      # raw Magellan: active fusion splice
CONN_RAW_CONTINUOUS = "<- CONTINUOUS ->"  # raw Magellan: continuous pass-through
CONN_FUSED          = "< --- >"           # normalized: connected / spliced fiber
CONN_UNUSED         = "X"                 # unused fiber (same in raw and normalized)


# ---------------------------------------------------------------------------
# Public helper: resolve a color and return its cached PatternFill.
# ---------------------------------------------------------------------------
def get_fill(key_or_hex: str) -> PatternFill:
    """
    Return a cached PatternFill for a COLOR dict key (e.g. 'NORTH') or a
    raw 6-character hex string (e.g. 'FFA500').

    If the color is not already in the cache, it is created on the fly and
    added so future calls are served from the cache.
    """
    # Resolve a named key to its hex value; fall back to treating the input
    # directly as a hex string. Take the last 6 chars to strip any FF alpha prefix.
    hex6 = COLOR.get(key_or_hex.upper(), key_or_hex).upper()[-6:]
    if hex6 not in FILLS:
        FILLS[hex6] = _fill(hex6)
    return FILLS[hex6]

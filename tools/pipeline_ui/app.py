"""
app.py -- Pipeline Web UI Flask application.

Runs the full fiber data pipeline from the browser with:
  - Dual file upload (KMZ + cut sheet; optional HAF + tap template)
  - Real-time log streaming via Server-Sent Events (SSE)
  - Inline checkpoint review page with Leaflet map overlay (after step 2)
  - Download links for all output files
"""

from __future__ import annotations

import io
import json
import os
import queue
import shutil
import sys
import tempfile
import threading
import uuid
import zipfile
from pathlib import Path

import re
import openpyxl
from openpyxl.styles import PatternFill
from flask import (
    Flask, Response, redirect, render_template,
    request, send_file, session, url_for, flash, jsonify
)

# ---------------------------------------------------------------------------
# Path setup — pipeline root must be on sys.path for pipeline_runner import
# ---------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent          # tools/pipeline_ui/
_ROOT = _HERE.parent.parent                      # Fiber Data Pipeline/
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# TIA-598 standard color order (buffer tubes and individual fibers).
# config is at project root; import after sys.path is updated above.
from config import FIBER_COLORS as _TIA_COLORS

import pipeline_runner
from map_builder import build_geojson


def _safe_save_workbook(wb, fpath: Path) -> None:
    """
    Save `wb` to `fpath` safely:
      1. Serialize to BytesIO so any openpyxl exception is raised before
         touching the on-disk file.
      2. Validate the BytesIO result by decompressing every ZIP entry
         (catches truncated XML that would fail at load time).
      3. Write atomically via a sibling .tmp file then rename.

    Raises RuntimeError if the serialized bytes fail validation.
    """
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    # Decompress every entry in the ZIP to catch truncated XML.
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as _zf:
            for _name in _zf.namelist():
                _zf.read(_name)
    except Exception as exc:
        raise RuntimeError(
            f"Workbook serialization produced an unreadable file for "
            f"{fpath.name}: {exc}"
        ) from exc

    # Atomic write: write to .tmp, then rename over the target.
    tmp = fpath.with_suffix(".tmp")
    tmp.write_bytes(data)
    tmp.replace(fpath)


def _rename_cable_size(cable: str, new_size: str) -> str:
    """Return the cable name with its CT suffix replaced by new_size."""
    m = re.search(r'_(\d{2,3}CT)$', cable, re.IGNORECASE)
    if m:
        return cable[:m.start()] + '_' + new_size.upper()
    m = re.match(r'^(\d{2,3}CT)(\s)', cable, re.IGNORECASE)
    if m:
        return new_size.upper() + cable[m.end(1):]
    return cable


def _adjust_sheath_row_count(wb, cable: str, new_size: str) -> int:
    """
    Adjust the number of fiber rows for `cable` in every worksheet of `wb`
    to match the fiber count implied by `new_size` (e.g. '96CT' -> 96 rows).

    New rows are generated with:
    - The same SHEATH UUID as the existing block (cols A and P)
    - BUFFER / FIBER continuing the TIA-598 sequence from the last existing row
    - CONNECTION = connection symbol (copied from existing rows) when the
      right-side cable supports that fiber position; 'X' with blank right-side
      columns when it does not.  A 96CT cable spliced to a 48CT cable gets
      '<- FUSION ->' on fibers 1-48 and 'X' / blank on fibers 49-96.
    - Right-side columns (FIBER, BUFFER, END ENC, START ENC, SHEATH NAME, UUID)
      filled only for connected rows; cleared to None for 'X' rows.

    Returns the total number of rows added or removed across all sheets.
    """
    m = re.match(r'^(\d+)CT$', new_size, re.IGNORECASE)
    if not m:
        return 0
    target_count = int(m.group(1))
    rows_changed = 0

    for ws in wb.worksheets:
        # Locate the SHEATHS header row (row containing 'SHEATH UUID')
        hdr_row = None
        for r in range(1, min(ws.max_row, 120) + 1):
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and 'SHEATH UUID' in v.upper():
                    hdr_row = r
                    break
            if hdr_row:
                break
        if hdr_row is None:
            continue

        # Locate columns by header label (AGENTS.md defaults in parentheses)
        def _col(label, after=0):
            for c in range(1, ws.max_column + 1):
                if c <= after:
                    continue
                v = ws.cell(hdr_row, c).value
                if isinstance(v, str) and v.strip().upper() == label.upper():
                    return c
            return None

        sheath_col      = _col('SHEATH NAME')                      or 2
        start_enc_col   = _col('START ENC')                        or 3
        end_enc_col     = _col('END ENC')                          or 4
        buffer_col      = _col('BUFFER')                           or 5
        fiber_col       = _col('FIBER')                            or 6
        conn_col        = _col('CONNECTION')                       or 10
        r_fiber_col     = _col('FIBER',       after=conn_col)      or 11
        r_buffer_col    = _col('BUFFER',      after=conn_col)      or 12
        r_end_enc_col   = _col('END ENC',     after=conn_col)      or 13
        r_start_enc_col = _col('START ENC',   after=conn_col)      or 14
        r_sheath_col    = _col('SHEATH NAME', after=conn_col)      or 15
        r_uuid_col      = _col('SHEATH UUID', after=conn_col)      or 16

        # Stop before the OPTICAL SPLITTERS section if present
        splitter_row = None
        for r in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and 'OPTICAL SPLITTERS' in v.upper():
                splitter_row = r
                break
        end_row = (splitter_row - 1) if splitter_row else ws.max_row

        # Collect all real fiber rows for this cable (sub-circuit rows have None in col A)
        block_rows = []
        for r in range(hdr_row + 1, end_row + 1):
            v = ws.cell(r, sheath_col).value
            if isinstance(v, str) and v.strip() == cable:
                block_rows.append(r)

        if not block_rows:
            continue

        current_count = len(block_rows)
        last_row_idx  = block_rows[-1]
        max_col       = ws.max_column

        # All rows in a sheath block share the same UUID — read from the first row
        sheath_uuid = ws.cell(block_rows[0], 1).value

        if current_count == target_count:
            # No row count change, but check for residual X rows left over from a prior
            # resize where the right-side cable was smaller than it is now.
            # Read the right-side cable from the LAST connected row so multi-partner
            # blocks (e.g. South: MST rows 1-12, West rows 13-48, East rows 49-96)
            # use the most-recent partner for the tail section.
            r_cable_refresh     = None
            r_uuid_refresh      = None
            r_enc_end_refresh   = None
            r_enc_start_refresh = None
            conn_sym_refresh    = "<- FUSION ->"
            for br in reversed(block_rows):
                rv = ws.cell(br, r_sheath_col).value
                if rv and str(rv).strip():
                    r_cable_refresh = str(rv).strip()
                    r_uuid_refresh  = ws.cell(br, r_uuid_col).value
                    break
            if r_cable_refresh is None:
                continue
            rm = re.search(r'_(\d+)CT\b', r_cable_refresh, re.IGNORECASE)
            if not rm:
                rm = re.match(r'^(\d+)CT\b', r_cable_refresh, re.IGNORECASE)
            r_ct_refresh = int(rm.group(1)) if rm else 0
            if r_ct_refresh == 0:
                continue
            for br in block_rows:
                v = str(ws.cell(br, conn_col).value or "").strip()
                if v and v != "X":
                    conn_sym_refresh = v
                    break
            for br in reversed(block_rows):
                v = ws.cell(br, r_end_enc_col).value
                if v is not None:
                    r_enc_end_refresh   = v
                    r_enc_start_refresh = ws.cell(br, r_start_enc_col).value
                    break
            for pos_idx, br in enumerate(block_rows):
                if pos_idx + 1 > r_ct_refresh:
                    break
                cv    = str(ws.cell(br, conn_col).value or "").strip()
                rname = ws.cell(br, r_sheath_col).value
                if cv != "X" or rname is not None:
                    continue
                buf_c = str(ws.cell(br, buffer_col).value or "").strip()
                fib_c = str(ws.cell(br, fiber_col).value  or "").strip()
                ws.cell(br, conn_col).value         = conn_sym_refresh
                ws.cell(br, r_fiber_col).value      = fib_c
                ws.cell(br, r_buffer_col).value     = buf_c
                ws.cell(br, r_end_enc_col).value    = r_enc_end_refresh
                ws.cell(br, r_start_enc_col).value  = r_enc_start_refresh
                ws.cell(br, r_sheath_col).value     = r_cable_refresh
                ws.cell(br, r_uuid_col).value       = r_uuid_refresh
                rows_changed += 1
            continue

        if current_count < target_count:
            add_count = target_count - current_count

            # Detect the connection symbol used in this workbook.
            # Magellan raw format is '<- FUSION ->'; step 5 normalises to '<--->'.
            # Read from existing connected rows so we match whatever is already there.
            conn_symbol = "<- FUSION ->"
            for br in block_rows:
                v = str(ws.cell(br, conn_col).value or "").strip()
                if v and v != "X":
                    conn_symbol = v
                    break

            # Determine how many fibers the right-side cable can accept.
            # Read from the LAST connected row so that multi-partner blocks use the
            # most-recent right-side cable (e.g. in a South block that transitions
            # MST → West → East, new rows appended at the East tail connect to East).
            r_cable_name      = None
            r_sheath_uuid_val = None
            for br in reversed(block_rows):
                rv = ws.cell(br, r_sheath_col).value
                if rv and str(rv).strip():
                    r_cable_name      = str(rv).strip()
                    r_sheath_uuid_val = ws.cell(br, r_uuid_col).value
                    break

            r_cable_ct = 0
            if r_cable_name:
                rm = re.search(r'_(\d+)CT\b', r_cable_name, re.IGNORECASE)
                if not rm:
                    rm = re.match(r'^(\d+)CT\b', r_cable_name, re.IGNORECASE)
                if rm:
                    r_cable_ct = int(rm.group(1))
            # Fallback: count existing connected (non-X) rows in the block
            if r_cable_ct == 0:
                r_cable_ct = sum(
                    1 for br in block_rows
                    if str(ws.cell(br, conn_col).value or "").strip() not in ("X", "")
                )

            # Left-side enclosure comes from the last row (always populated, even on X rows).
            template_start_enc = ws.cell(last_row_idx, start_enc_col).value
            template_end_enc   = ws.cell(last_row_idx, end_enc_col).value

            # Right-side enclosure: scan backward for the last CONNECTED row, because
            # last_row_idx may be an X row (blank right side) from a prior resize.
            template_r_end_enc   = None
            template_r_start_enc = None
            for br in reversed(block_rows):
                v = ws.cell(br, r_end_enc_col).value
                if v is not None:
                    template_r_end_enc   = v
                    template_r_start_enc = ws.cell(br, r_start_enc_col).value
                    break

            # Find starting position in TIA-598 sequence from last existing row
            last_buf = str(ws.cell(last_row_idx, buffer_col).value or "").strip().upper()
            last_fib = str(ws.cell(last_row_idx, fiber_col).value  or "").strip().upper()
            buf_idx  = _TIA_COLORS.index(last_buf) if last_buf in _TIA_COLORS else 0
            fib_idx  = _TIA_COLORS.index(last_fib) if last_fib in _TIA_COLORS else 11
            fib_idx += 1
            if fib_idx >= 12:
                fib_idx = 0
                buf_idx += 1

            # Shift existing rows downward manually instead of using ws.insert_rows().
            # openpyxl's insert_rows rewrites internal XML references (merged cells,
            # conditional formatting, named ranges) which corrupts complex Magellan
            # workbooks when saved.  Copying values cell-by-cell avoids that entirely.
            _ceil = ws.max_row
            for _r in range(_ceil, last_row_idx, -1):
                for _c in range(1, max_col + 1):
                    ws.cell(_r + add_count, _c).value = ws.cell(_r, _c).value
            for _r in range(last_row_idx + 1, last_row_idx + add_count + 1):
                for _c in range(1, max_col + 1):
                    ws.cell(_r, _c).value = None

            for i in range(add_count):
                new_r          = last_row_idx + 1 + i
                fiber_position = current_count + i + 1   # 1-based position in block

                buf_code = _TIA_COLORS[min(buf_idx, 11)]
                fib_code = _TIA_COLORS[fib_idx]

                # Left-side columns (always populated for every new row)
                ws.cell(new_r, 1).value                = sheath_uuid
                ws.cell(new_r, sheath_col).value       = cable
                ws.cell(new_r, start_enc_col).value    = template_start_enc
                ws.cell(new_r, end_enc_col).value      = template_end_enc
                ws.cell(new_r, buffer_col).value       = buf_code
                ws.cell(new_r, fiber_col).value        = fib_code

                if r_cable_ct > 0 and fiber_position <= r_cable_ct:
                    # Right-side cable has a fiber at this position — show connection
                    ws.cell(new_r, conn_col).value        = conn_symbol
                    ws.cell(new_r, r_fiber_col).value     = fib_code
                    ws.cell(new_r, r_buffer_col).value    = buf_code
                    ws.cell(new_r, r_end_enc_col).value   = template_r_end_enc
                    ws.cell(new_r, r_start_enc_col).value = template_r_start_enc
                    ws.cell(new_r, r_sheath_col).value    = r_cable_name
                    ws.cell(new_r, r_uuid_col).value      = r_sheath_uuid_val
                else:
                    # Right-side cable doesn't reach this fiber — X and blank right side
                    ws.cell(new_r, conn_col).value = "X"
                    for c in range(conn_col + 1, max_col + 1):
                        ws.cell(new_r, c).value = None

                fib_idx += 1
                if fib_idx >= 12:
                    fib_idx = 0
                    buf_idx += 1

            rows_changed += add_count

        else:
            # Remove excess rows from the end of the block
            remove_count = current_count - target_count
            ws.delete_rows(last_row_idx - remove_count + 1, amount=remove_count)
            rows_changed += remove_count

    return rows_changed


def _heal_partner_x_rows(wb, resized_cable: str, new_size: str) -> int:
    """
    After resizing ``resized_cable`` to ``new_size``, scan every worksheet
    for cable blocks whose *last* connected right-side partner is
    ``resized_cable`` and that have trailing X rows (blank right side).

    This handles the order-of-resize problem: when cable B is resized
    before cable A, B's new rows become X because A was still small.
    Once A is resized here, those X rows in B are healed automatically.

    Only heals rows whose 1-based position within the block is ≤ the new
    CT count of the resized cable, and only where CONNECTION == 'X' and
    the right-side SHEATH NAME is blank — i.e. rows that were set to X
    specifically because the partner didn't reach them yet.
    """
    m = re.match(r'^(\d+)CT$', new_size, re.IGNORECASE)
    if not m:
        return 0
    new_ct = int(m.group(1))
    healed = 0

    for ws in wb.worksheets:
        hdr_row = None
        for r in range(1, min(ws.max_row, 120) + 1):
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and 'SHEATH UUID' in v.upper():
                    hdr_row = r
                    break
            if hdr_row:
                break
        if hdr_row is None:
            continue

        def _col(label, after=0):
            for c in range(1, ws.max_column + 1):
                if c <= after:
                    continue
                v = ws.cell(hdr_row, c).value
                if isinstance(v, str) and v.strip().upper() == label.upper():
                    return c
            return None

        sheath_col      = _col('SHEATH NAME')                      or 2
        buffer_col      = _col('BUFFER')                           or 5
        fiber_col       = _col('FIBER')                            or 6
        conn_col        = _col('CONNECTION')                       or 10
        r_fiber_col     = _col('FIBER',       after=conn_col)      or 11
        r_buffer_col    = _col('BUFFER',      after=conn_col)      or 12
        r_end_enc_col   = _col('END ENC',     after=conn_col)      or 13
        r_start_enc_col = _col('START ENC',   after=conn_col)      or 14
        r_sheath_col    = _col('SHEATH NAME', after=conn_col)      or 15
        r_uuid_col      = _col('SHEATH UUID', after=conn_col)      or 16

        splitter_row = None
        for r in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and 'OPTICAL SPLITTERS' in v.upper():
                splitter_row = r
                break
        end_row = (splitter_row - 1) if splitter_row else ws.max_row

        # Group rows into cable blocks by left-side sheath name.
        current_name = None
        current_rows = []
        blocks: list[tuple[str, list[int]]] = []
        for r in range(hdr_row + 1, end_row + 1):
            v = ws.cell(r, sheath_col).value
            if isinstance(v, str) and v.strip():
                name = v.strip()
                if name != current_name:
                    if current_name and current_rows:
                        blocks.append((current_name, current_rows))
                    current_name = name
                    current_rows = [r]
                else:
                    current_rows.append(r)
        if current_name and current_rows:
            blocks.append((current_name, current_rows))

        for blk_name, brows in blocks:
            if blk_name == resized_cable:
                continue  # skip the cable we just resized

            # Find the right-side cable of the LAST connected row.
            last_r_cable = None
            last_r_uuid  = None
            for br in reversed(brows):
                rv = ws.cell(br, r_sheath_col).value
                if rv and str(rv).strip():
                    last_r_cable = str(rv).strip()
                    last_r_uuid  = ws.cell(br, r_uuid_col).value
                    break

            # Only heal blocks whose last partner is the cable we just resized.
            if last_r_cable != resized_cable:
                continue

            # Get right-side enclosure values and connection symbol.
            r_enc_end = r_enc_start = None
            for br in reversed(brows):
                v = ws.cell(br, r_end_enc_col).value
                if v is not None:
                    r_enc_end   = v
                    r_enc_start = ws.cell(br, r_start_enc_col).value
                    break

            conn_sym = "<- FUSION ->"
            for br in brows:
                v = str(ws.cell(br, conn_col).value or "").strip()
                if v and v != "X":
                    conn_sym = v
                    break

            # Heal trailing X rows (blank right side) up to new_ct capacity.
            for pos_idx, br in enumerate(brows):
                if pos_idx + 1 > new_ct:
                    break
                cv    = str(ws.cell(br, conn_col).value or "").strip()
                rname = ws.cell(br, r_sheath_col).value
                if cv != "X" or rname is not None:
                    continue
                buf_c = str(ws.cell(br, buffer_col).value or "").strip()
                fib_c = str(ws.cell(br, fiber_col).value  or "").strip()
                ws.cell(br, conn_col).value          = conn_sym
                ws.cell(br, r_fiber_col).value       = fib_c
                ws.cell(br, r_buffer_col).value      = buf_c
                ws.cell(br, r_end_enc_col).value     = r_enc_end
                ws.cell(br, r_start_enc_col).value   = r_enc_start
                ws.cell(br, r_sheath_col).value      = resized_cable
                ws.cell(br, r_uuid_col).value        = last_r_uuid
                healed += 1

    return healed


# ---------------------------------------------------------------------------
# In-memory job registry
# Each job: { queue, thread, checkpoint_event, checkpoint_continue,
#             job_dir, status, outputs }
# ---------------------------------------------------------------------------
_JOBS: dict[str, dict] = {}


def _make_job_id() -> str:
    return uuid.uuid4().hex[:12]


# ---------------------------------------------------------------------------
# Pipeline background worker
# ---------------------------------------------------------------------------

def _run_job(job_id: str, data_dir: str, output_dir: str) -> None:
    job = _JOBS[job_id]
    q   = job["queue"]

    def log_cb(step: int, line: str) -> None:
        q.put({"type": "log", "step": step, "line": line})

    def checkpoint_cb(file_path: str) -> bool:
        """Pause the pipeline thread and wait for the browser to respond."""
        q.put({"type": "checkpoint", "path": file_path})
        job["checkpoint_event"].wait()          # block until user clicks Continue/Stop
        return job["checkpoint_continue"]       # True = continue, False = abort

    try:
        result = pipeline_runner.run_pipeline(
            data_dir=data_dir,
            output_dir=output_dir,
            start=1,
            stop=10,
            checkpoint_callback=checkpoint_cb,
            log_callback=log_cb,
        )
        job["status"] = result["status"]
        # Collect output files
        outputs = {}
        for fname in ["Asbuilt_Workbook_post12.xlsx", "Path_of_Light_Confirmation.xlsx"]:
            fpath = Path(output_dir) / fname
            if fpath.exists():
                outputs[fname] = str(fpath)
        # Tap report is named "{OLT} Tap Report.xlsx" (OLT name varies by project)
        for tap_path in sorted(Path(output_dir).glob("* Tap Report.xlsx")):
            outputs[tap_path.name] = str(tap_path)
            break
        job["outputs"] = outputs
        q.put({"type": "done", "status": result["status"],
               "error": result.get("error"), "outputs": list(outputs.keys())})
    except Exception as exc:
        job["status"] = "failed"
        q.put({"type": "done", "status": "failed", "error": str(exc), "outputs": []})


# ---------------------------------------------------------------------------
# App factory
# ---------------------------------------------------------------------------

def create_app() -> Flask:
    app = Flask(__name__)
    app.secret_key = os.urandom(32)
    app.config["MAX_CONTENT_LENGTH"] = 128 * 1024 * 1024  # 128 MB

    # -----------------------------------------------------------------------
    # Landing / upload page
    # -----------------------------------------------------------------------
    @app.route("/", methods=["GET"])
    def index():
        return render_template("index.html")

    # -----------------------------------------------------------------------
    # Start a pipeline run
    # -----------------------------------------------------------------------
    @app.route("/run", methods=["POST"])
    def start_run():
        kmz_file   = request.files.get("kmz")
        sheet_file = request.files.get("cutsheet")

        if not kmz_file or not kmz_file.filename.lower().endswith(".kmz"):
            flash("Please upload a .kmz file.")
            return redirect(url_for("index"))
        if not sheet_file or not sheet_file.filename.lower().endswith(".xlsx"):
            flash("Please upload the cut sheet (.xlsx).")
            return redirect(url_for("index"))

        # Create an isolated temp job directory.
        job_id  = _make_job_id()
        job_dir = Path(tempfile.mkdtemp(prefix=f"pipeline_{job_id}_"))
        data_dir   = job_dir / "data"
        output_dir = job_dir / "output"
        data_dir.mkdir(); output_dir.mkdir()

        # Save required files.
        kmz_file.save(data_dir / kmz_file.filename)
        sheet_file.save(data_dir / sheet_file.filename)

        # Save optional files.
        haf_file = request.files.get("haf")
        if haf_file and haf_file.filename:
            haf_file.save(data_dir / haf_file.filename)

        # Auto-copy Tap Report Template from pipeline data/ folder if present.
        _template_src = _ROOT / "data" / "Tap_Report_Template.xlsx"
        if _template_src.exists():
            shutil.copy(str(_template_src), str(data_dir / "Tap_Report_Template.xlsx"))

        # Register job.
        _JOBS[job_id] = {
            "queue":              queue.Queue(),
            "thread":             None,
            "checkpoint_event":   threading.Event(),
            "checkpoint_continue": True,
            "job_dir":            str(job_dir),
            "status":             "running",
            "outputs":            {},
        }

        # Launch background thread.
        t = threading.Thread(
            target=_run_job,
            args=(job_id, str(data_dir), str(output_dir)),
            daemon=True,
        )
        t.start()
        _JOBS[job_id]["thread"] = t

        return redirect(url_for("run_page", job_id=job_id))

    # -----------------------------------------------------------------------
    # Live run page
    # -----------------------------------------------------------------------
    @app.route("/run/<job_id>")
    def run_page(job_id):
        if job_id not in _JOBS:
            flash("Job not found.")
            return redirect(url_for("index"))
        return render_template("run.html", job_id=job_id)

    # -----------------------------------------------------------------------
    # SSE endpoint — streams log lines to the browser
    # -----------------------------------------------------------------------
    @app.route("/stream/<job_id>")
    def stream(job_id):
        if job_id not in _JOBS:
            return Response("data: {}\n\n", mimetype="text/event-stream")

        job = _JOBS[job_id]

        def generate():
            while True:
                try:
                    msg = job["queue"].get(timeout=30)
                except queue.Empty:
                    yield "data: {\"type\": \"heartbeat\"}\n\n"
                    continue

                yield f"data: {json.dumps(msg)}\n\n"

                if msg["type"] == "checkpoint":
                    # Pause the stream and wait — the browser will redirect to /checkpoint/
                    # The pipeline thread is also waiting. We just stop sending new events.
                    break

                if msg["type"] == "done":
                    break

        return Response(generate(), mimetype="text/event-stream",
                        headers={"Cache-Control": "no-cache",
                                 "X-Accel-Buffering": "no"})

    # -----------------------------------------------------------------------
    # Checkpoint page
    # -----------------------------------------------------------------------
    @app.route("/checkpoint/<job_id>", methods=["GET"])
    def checkpoint_page(job_id):
        if job_id not in _JOBS:
            flash("Job not found.")
            return redirect(url_for("index"))
        job = _JOBS[job_id]
        # Look for the Colored Connections Table in the output dir.
        job_dir    = Path(job["job_dir"])
        conn_table = job_dir / "output" / "Colored_Connections_Table.xlsx"
        return render_template("checkpoint.html", job_id=job_id,
                               has_conn_table=conn_table.exists())

    @app.route("/checkpoint/<job_id>/continue", methods=["POST"])
    def checkpoint_continue(job_id):
        if job_id in _JOBS:
            _JOBS[job_id]["checkpoint_continue"] = True
            _JOBS[job_id]["checkpoint_event"].set()
        return redirect(url_for("run_page", job_id=job_id))

    @app.route("/checkpoint/<job_id>/stop", methods=["POST"])
    def checkpoint_stop(job_id):
        if job_id in _JOBS:
            _JOBS[job_id]["checkpoint_continue"] = False
            _JOBS[job_id]["checkpoint_event"].set()
        return redirect(url_for("complete_page", job_id=job_id))

    # -----------------------------------------------------------------------
    # GeoJSON API — cable network map data for the checkpoint Leaflet map
    # -----------------------------------------------------------------------
    @app.route("/geojson/<job_id>")
    def geojson(job_id):
        if job_id not in _JOBS:
            return jsonify({"type": "FeatureCollection", "features": []})
        try:
            data = build_geojson(_JOBS[job_id]["job_dir"])
        except Exception as exc:
            data = {"type": "FeatureCollection", "features": [],
                    "error": str(exc)}
        return jsonify(data)

    # -----------------------------------------------------------------------
    # Direction override — update cable fill in the Colored Connections Table
    # -----------------------------------------------------------------------
    _DIR_TO_HEX = {
        "NORTH": "FFA500", "SOUTH": "8B4513", "EAST": "008000",
        "WEST":  "708090", "OLT":   "C5D9B5", "MST":  "FF0000",
    }
    _DIR_LABELS = {
        "NORTH": "North", "SOUTH": "South", "EAST": "East",
        "WEST":  "West",  "OLT":   "OLT",   "MST":  "MST",
    }

    @app.route("/override/<job_id>", methods=["POST"])
    def override_direction(job_id):
        if job_id not in _JOBS:
            return jsonify({"ok": False, "error": "Job not found"}), 404
        data      = request.get_json(force=True) or {}
        cable     = str(data.get("cable", "")).strip()
        direction = str(data.get("direction", "")).upper()
        if not cable or direction not in _DIR_TO_HEX:
            return jsonify({"ok": False, "error": "Invalid cable or direction"}), 400

        hex6      = _DIR_TO_HEX[direction]
        new_fill  = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
        css_color = f"#{hex6}"

        colored_path = Path(_JOBS[job_id]["job_dir"]) / "output" / "Colored_Connections_Table.xlsx"
        if not colored_path.exists():
            return jsonify({"ok": False, "error": "Connections table not found"}), 404

        wb = openpyxl.load_workbook(str(colored_path))
        ws = wb.active
        changed = 0
        for row in ws.iter_rows(min_row=2):
            for cell in row[3:]:   # skip Location, Lat, Lon columns
                if str(cell.value or "").strip() == cable:
                    cell.fill = new_fill
                    changed  += 1
        _safe_save_workbook(wb, colored_path)

        return jsonify({"ok": True, "changed": changed,
                        "color": css_color, "direction": _DIR_LABELS[direction]})

    # -----------------------------------------------------------------------
    # Cable size rename — updates CT suffix in both xlsx tables
    # -----------------------------------------------------------------------
    @app.route("/resize/<job_id>", methods=["POST"])
    def resize_cable(job_id):
        if job_id not in _JOBS:
            return jsonify({"ok": False, "error": "Job not found"}), 404
        data     = request.get_json(force=True) or {}
        cable    = str(data.get("cable", "")).strip()
        new_size = str(data.get("size",  "")).strip().upper()
        if not cable or not re.match(r'^\d{2,3}CT$', new_size):
            return jsonify({"ok": False, "error": "Invalid cable or size"}), 400

        new_cable = _rename_cable_size(cable, new_size)
        if new_cable == cable:
            # Distinguish "already at this size" (ok, no-op) from
            # "CT pattern not found in name" (actual failure).
            already = bool(
                re.search(r'_' + re.escape(new_size) + r'$', cable, re.IGNORECASE)
                or re.match(r'^' + re.escape(new_size) + r'\b', cable, re.IGNORECASE)
            )
            if already:
                return jsonify({"ok": True, "old_cable": cable, "new_cable": cable, "changed": 0})
            return jsonify({
                "ok": False,
                "error": (
                    f"Cannot resize '{cable}': no recognized CT count found in the cable name. "
                    "Only cables with a trailing _NNCT suffix (e.g. _48CT) or leading NNCT "
                    "prefix (legacy format) can be resized here."
                ),
            }), 400

        job_dir = Path(_JOBS[job_id]["job_dir"])
        changed_total   = 0
        cut_sheet_found = False   # tracks whether cut_sheet.xlsx existed
        cut_sheet_changed = 0     # rename hits in cut_sheet.xlsx specifically
        for fpath in [
            job_dir / "data"   / "Connections_Table.xlsx",
            job_dir / "data"   / "cut_sheet.xlsx",
            job_dir / "output" / "Colored_Connections_Table.xlsx",
        ]:
            if not fpath.exists():
                continue
            if fpath.name == "cut_sheet.xlsx":
                cut_sheet_found = True
            wb = openpyxl.load_workbook(str(fpath), keep_links=False)
            changed = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value or "").strip() == cable:
                            cell.value = new_cable
                            changed += 1
            if fpath.name == "cut_sheet.xlsx":
                cut_sheet_changed = changed
            # For the cut sheet, also adjust the per-sheath row count and
            # heal any partner cable blocks that have trailing X rows from a
            # previous resize where the partner was still smaller.
            row_adj = 0
            if fpath.name == "cut_sheet.xlsx":
                row_adj  = _adjust_sheath_row_count(wb, new_cable, new_size)
                row_adj += _heal_partner_x_rows(wb, new_cable, new_size)
                changed += row_adj
            if changed:
                try:
                    _safe_save_workbook(wb, fpath)
                except RuntimeError as save_err:
                    if row_adj and fpath.name == "cut_sheet.xlsx":
                        # Row adjustment may have triggered the corruption.
                        # Fall back to rename-only: reload and apply just the rename.
                        try:
                            wb2 = openpyxl.load_workbook(str(fpath), keep_links=False)
                            rename_count = 0
                            for ws2 in wb2.worksheets:
                                for row2 in ws2.iter_rows():
                                    for cell2 in row2:
                                        if str(cell2.value or "").strip() == cable:
                                            cell2.value = new_cable
                                            rename_count += 1
                            if rename_count:
                                _safe_save_workbook(wb2, fpath)
                                changed_total += rename_count
                        except Exception:
                            pass
                        return jsonify({
                            "ok": False,
                            "error": (
                                "Row count adjustment failed to save cleanly "
                                f"({save_err}). Cable was renamed but fiber rows "
                                "were not added — proceed with caution."
                            ),
                        }), 500
                    return jsonify({"ok": False, "error": str(save_err)}), 500
                changed_total += changed

        # Warn if the cut sheet existed but the cable name wasn't found there —
        # this means the name in the connections table differs from the cut sheet,
        # so the pipeline will colorize using an unmatched name.
        if cut_sheet_found and cut_sheet_changed == 0 and changed_total > 0:
            return jsonify({
                "ok": False,
                "error": (
                    f"'{cable}' was found in the Connections Table but not in the "
                    "cut sheet — the name may differ slightly between files. "
                    "The rename was NOT applied to cut_sheet.xlsx. Do not continue "
                    "the pipeline until this is resolved."
                ),
            }), 400

        return jsonify({"ok": True, "old_cable": cable, "new_cable": new_cable,
                        "changed": changed_total})

    # -----------------------------------------------------------------------
    # Completion page
    # -----------------------------------------------------------------------
    @app.route("/complete/<job_id>")
    def complete_page(job_id):
        if job_id not in _JOBS:
            flash("Job not found.")
            return redirect(url_for("index"))
        job = _JOBS[job_id]
        return render_template("complete.html", job_id=job_id,
                               status=job["status"],
                               outputs=list(job["outputs"].keys()))

    # -----------------------------------------------------------------------
    # File downloads
    # -----------------------------------------------------------------------
    @app.route("/download/<job_id>/<filename>")
    def download(job_id, filename):
        if job_id not in _JOBS:
            flash("Job not found.")
            return redirect(url_for("index"))
        job     = _JOBS[job_id]
        outputs = job.get("outputs", {})
        if filename not in outputs:
            flash(f"{filename} not found.")
            return redirect(url_for("complete_page", job_id=job_id))
        return send_file(outputs[filename], as_attachment=True,
                         download_name=filename)

    @app.route("/download/<job_id>/connections")
    def download_connections(job_id):
        if job_id not in _JOBS:
            return redirect(url_for("index"))
        fpath = Path(_JOBS[job_id]["job_dir"]) / "output" / "Colored_Connections_Table.xlsx"
        if not fpath.exists():
            flash("Connections table not available yet.")
            return redirect(url_for("checkpoint_page", job_id=job_id))
        return send_file(str(fpath), as_attachment=True,
                         download_name="Colored_Connections_Table.xlsx")

    return app

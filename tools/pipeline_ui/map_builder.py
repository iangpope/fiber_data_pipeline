"""
map_builder.py -- Build GeoJSON from pipeline step 1+2 outputs for the
checkpoint map view.

Combines:
  1. data/Connections_Table.xlsx        -- location names + GPS coords + cable names
  2. output/Colored_Connections_Table.xlsx -- cable-to-color assignments (cell fill)
  3. data/<project>.kmz                 -- cable line geometries

Outputs a single GeoJSON FeatureCollection with:
  - One MultiLineString feature per cable (all segments, neutral color)
  - One Point feature per location (markers, with per-location cable colors)
"""

from __future__ import annotations

import json
import os
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from typing import Optional

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Pipeline root on sys.path for config import
# ---------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent          # tools/pipeline_ui/
_ROOT = _HERE.parent.parent                      # Fiber Data Pipeline/
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from config import COLOR


# ---------------------------------------------------------------------------
# Color mapping: openpyxl fill hex → human label + CSS hex
# ---------------------------------------------------------------------------

_FILL_TO_INFO: dict[str, dict] = {
    COLOR["NORTH"]: {"label": "North",   "css": f"#{COLOR['NORTH']}"},
    COLOR["SOUTH"]: {"label": "South",   "css": f"#{COLOR['SOUTH']}"},
    COLOR["EAST"]:  {"label": "East",    "css": f"#{COLOR['EAST']}"},
    COLOR["WEST"]:  {"label": "West",    "css": f"#{COLOR['WEST']}"},
    COLOR["OLT"]:   {"label": "OLT",     "css": f"#{COLOR['OLT']}"},
    COLOR["MST"]:   {"label": "MST",     "css": f"#{COLOR['MST']}"},
}

_DEFAULT_COLOR = {"label": "Unknown", "css": "#888888"}


def _parse_cable_endpoints(cable_name: str):
    """Return (a_end, b_end) location names from a cable name, or (None, None)."""
    t = str(cable_name or "").strip()
    if "_TO_" in t:
        a, b = t.split("_TO_", 1)
        b = re.sub(r"_\d{2,3}CT\b", "", b, flags=re.IGNORECASE)
        return a.strip(), b.strip()
    m = re.match(r"^\s*\d{2,3}CT\s+(\S+)\s+TO\s+(\S+)\s*$", t, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, None


def _snap_cable_segments(
    segments: list[list[tuple[float, float]]],
    a_coord: tuple[float, float],
    b_coord: tuple[float, float],
) -> list[list[tuple[float, float]]]:
    """
    Snap a cable's endpoint coordinates to the known CT splice locations.

    fiberCable KMZ endpoints often drift up to ~300 m from the actual
    Connections_Table GPS coordinates (GPS inaccuracy, data updates).
    This function replaces the first point of the first segment and the
    last point of the last segment with the known CT coordinates, keeping
    all intermediate waypoints intact.
    """
    if not segments:
        return segments

    # Determine which CT endpoint is closer to the start of segment[0]
    start = segments[0][0]
    da = abs(start[0] - a_coord[0]) + abs(start[1] - a_coord[1])
    db = abs(start[0] - b_coord[0]) + abs(start[1] - b_coord[1])
    near_start, near_end = (a_coord, b_coord) if da <= db else (b_coord, a_coord)

    snapped = [list(seg) for seg in segments]  # shallow copy each seg list
    snapped[0][0]   = near_start
    snapped[-1][-1] = near_end
    return snapped


def _cable_short_name(cable_name: str, loc_name: str) -> str:
    """Return compact 'other-end (size)' label for this cable at this location."""
    if '_TO_' in cable_name:
        end_a, rest = cable_name.split('_TO_', 1)
        m = re.search(r'_(\d{2,3}CT)$', rest, re.IGNORECASE)
        size  = m.group(1) if m else ''
        end_b = rest[:m.start()] if m else rest
        other = end_b if end_a == loc_name else end_a
        return f"{other} ({size})" if size else other
    m = re.match(r'^(\d{2,3}CT)\s+(\S+)\s+TO\s+(\S+)$', cable_name, re.IGNORECASE)
    if m:
        size, ea, eb = m.group(1), m.group(2), m.group(3)
        return f"{eb if ea == loc_name else ea} ({size})"
    return cable_name


def _cable_size(cable_name: str) -> str:
    """Extract the fiber-count suffix from a cable name (e.g. '48CT')."""
    m = re.search(r'(\d{2,3}CT)', cable_name, re.IGNORECASE)
    return m.group(1).upper() if m else ''


def _hex_from_fill(cell) -> Optional[str]:
    """Extract the 6-char hex color from an openpyxl cell's PatternFill."""
    try:
        fgcolor = cell.fill.fgColor
        raw = fgcolor.rgb if fgcolor.type == "rgb" else None
        if raw and len(raw) >= 6:
            return raw[-6:].upper()
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# KMZ geometry helpers
#
# Magellan KMZ structure:
#   fiber/fiberCable/<cable name folder>/<Placemarks>   — named cables (2-pt straight lines)
#   undergroundSupport/supportCable/<GUID folder>/<Placemarks>  — actual conduit routes
#   aerialSupport/supportCable/<GUID folder>/<Placemarks>       — aerial cable routes
#
# The fiberCable layer only stores start/end points (straight lines), while
# supportCable stores the full detailed conduit path with many small segments.
# Both are needed: supportCable fills the map with real routes;
# fiberCable provides the named cable associations for directional coloring.
# ---------------------------------------------------------------------------

def _strip_ns(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def _get_name(el) -> str:
    n = next((c for c in el if _strip_ns(c.tag) == "name"), None)
    return (n.text or "").strip() if n is not None else ""


def _collect_linestring_segments(folder) -> list[list[tuple[float, float]]]:
    """Collect all LineString segments from all Placemarks in a folder."""
    segments = []
    for pm in folder:
        if _strip_ns(pm.tag) != "Placemark":
            continue
        ls = next((c for c in pm.iter() if _strip_ns(c.tag) == "LineString"), None)
        if ls is None:
            continue
        coords_el = next((c for c in ls.iter() if _strip_ns(c.tag) == "coordinates"), None)
        if coords_el is None or not coords_el.text:
            continue
        points = []
        for tok in coords_el.text.split():
            parts = tok.split(",")
            if len(parts) >= 2:
                try:
                    points.append((float(parts[1]), float(parts[0])))  # lat, lon
                except ValueError:
                    pass
        if points:
            segments.append(points)
    return segments


def _parse_kmz_cables(kmz_path: str) -> dict[str, list[list[tuple[float, float]]]]:
    """
    Return {cable_name: [[seg1_points], ...]} from fiber/fiberCable.
    Cable names are the folder names (e.g. 'RC73E_SE_001_TO_RC73E_FT_071_48CT').
    """
    cable_segments: dict[str, list] = {}

    with zipfile.ZipFile(kmz_path, "r") as kmz:
        kml_name = next((n for n in kmz.namelist() if n.endswith(".kml")), None)
        if not kml_name:
            return cable_segments
        root = ET.fromstring(kmz.read(kml_name))

    fiber_cable_folder = None
    for folder in root.iter():
        if _strip_ns(folder.tag) == "Folder" and _get_name(folder) == "fiberCable":
            fiber_cable_folder = folder
            break

    if fiber_cable_folder is None:
        return cable_segments

    for guid_folder in fiber_cable_folder:
        if _strip_ns(guid_folder.tag) != "Folder":
            continue
        cable_name = _get_name(guid_folder)
        if not cable_name:
            continue
        segments = _collect_linestring_segments(guid_folder)
        if segments:
            cable_segments[cable_name] = segments

    return cable_segments


def _parse_kmz_infrastructure(kmz_path: str) -> list[list[tuple[float, float]]]:
    """
    Return all LineString segments from undergroundSupport/supportCable and
    aerialSupport/supportCable as a flat list of segments.

    These segments form the actual physical conduit routes that match what
    Google Earth renders — far denser than the straight-line fiberCable layer.
    """
    all_segments: list = []

    with zipfile.ZipFile(kmz_path, "r") as kmz:
        kml_name = next((n for n in kmz.namelist() if n.endswith(".kml")), None)
        if not kml_name:
            return all_segments
        root = ET.fromstring(kmz.read(kml_name))

    doc = next((c for c in root if _strip_ns(c.tag) == "Document"), root)

    for top_folder in doc:
        if _strip_ns(top_folder.tag) != "Folder":
            continue
        top_name = _get_name(top_folder)
        if top_name not in ("undergroundSupport", "aerialSupport"):
            continue
        for sub in top_folder:
            if _strip_ns(sub.tag) != "Folder" or _get_name(sub) != "supportCable":
                continue
            # Each child is a GUID folder containing Placemarks
            for guid_folder in sub:
                if _strip_ns(guid_folder.tag) != "Folder":
                    continue
                all_segments.extend(_collect_linestring_segments(guid_folder))

    return all_segments


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------

def build_geojson(job_dir: str) -> dict:
    """
    Build a GeoJSON FeatureCollection from the outputs of pipeline steps 1 and 2.

    Parameters
    ----------
    job_dir : str
        Path to the job's temp directory (contains data/ and output/ sub-dirs).

    Returns
    -------
    dict -- GeoJSON FeatureCollection ready for json.dumps().
            Returns an empty FeatureCollection on any error.
    """
    data_dir   = Path(job_dir) / "data"
    output_dir = Path(job_dir) / "output"

    conn_table_path    = data_dir   / "Connections_Table.xlsx"
    colored_table_path = output_dir / "Colored_Connections_Table.xlsx"

    if not conn_table_path.exists() or not colored_table_path.exists():
        return {"type": "FeatureCollection", "features": []}

    # Find KMZ file.
    kmz_path = next(
        (str(data_dir / f) for f in os.listdir(data_dir) if f.lower().endswith(".kmz")),
        None,
    )

    # Load direction confidence data written by step 2 (absent on CLI-only runs).
    cable_confidence: dict[str, dict] = {}
    conf_summary: dict = {}
    _conf_path = output_dir / "direction_confidence.json"
    if _conf_path.exists():
        with open(_conf_path) as _f:
            _cdata = json.load(_f)
            cable_confidence = _cdata.get("cables", {})
            conf_summary     = _cdata.get("summary", {})

    # -----------------------------------------------------------------------
    # 1. Read cable → color from the Colored Connections Table.
    #
    # The same cable name can appear at multiple locations with DIFFERENT
    # colors (step 2 assigns direction relative to each location; the
    # deduplication pass may also shift a color at one location vs another).
    # Build both a per-location lookup AND a global fallback.
    # -----------------------------------------------------------------------
    cable_info: dict[str, dict] = {}            # cable_name → {label, css}  (global, last-seen)
    loc_cable_info: dict[str, dict] = {}        # location_name → {cable_name → {label, css}}

    wb_colored = openpyxl.load_workbook(colored_table_path)
    ws = wb_colored.active

    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, max_col=max_col):
        loc_val = str(row[0].value or "").strip()   # column A = Location
        for cell in row[3:]:   # skip Location, Lat, Lon columns
            val = str(cell.value or "").strip()
            if not val:
                continue
            hex6 = _hex_from_fill(cell)
            info = _FILL_TO_INFO.get(hex6, _DEFAULT_COLOR) if hex6 else _DEFAULT_COLOR
            cable_info[val] = info          # global (last-seen wins; used for cable line features)
            if loc_val:
                loc_cable_info.setdefault(loc_val, {})[val] = info

    # -----------------------------------------------------------------------
    # 2. Read location nodes from Connections Table.
    # -----------------------------------------------------------------------
    df = pd.read_excel(conn_table_path)
    location_data: dict[str, dict] = {}  # name → {lat, lon, cables:[...]}

    for _, row in df.iterrows():
        loc = str(row.get("Location", "")).strip()
        try:
            lat = float(row.get("Latitude", 0))
            lon = float(row.get("Longitude", 0))
        except (ValueError, TypeError):
            continue
        if not loc or lat == 0:
            continue

        cables = []
        loc_colors = loc_cable_info.get(str(loc), {})
        for col in df.columns:
            if str(col).startswith("Connection"):
                val = str(row.get(col, "")).strip()
                if val and val.lower() not in ("nan", "none", ""):
                    # Prefer the color from this location's own row in the Excel
                    # (direction is relative per-location; global cable_info can
                    # be wrong because a later location's row may overwrite it).
                    info = loc_colors.get(val) or cable_info.get(val, _DEFAULT_COLOR)
                    cables.append({
                        "name":       val,
                        "color":      info["css"],
                        "label":      info["label"],
                        "short_name": _cable_short_name(val, loc),
                        "size":       _cable_size(val),
                    })

        location_data[loc] = {"lat": lat, "lon": lon, "cables": cables}

    # -----------------------------------------------------------------------
    # 3. Parse KMZ geometries and snap cable endpoints to CT coordinates.
    # -----------------------------------------------------------------------
    kmz_cables = _parse_kmz_cables(kmz_path) if kmz_path else {}
    infra_segments = _parse_kmz_infrastructure(kmz_path) if kmz_path else []

    # Build a flat lat/lon lookup from location_data for snapping
    loc_coords = {name: (d["lat"], d["lon"]) for name, d in location_data.items()}

    # Snap each named cable's start/end to CT coordinates so cable lines
    # visually connect the location markers they reference in their name.
    for cable_name, segs in list(kmz_cables.items()):
        a_end, b_end = _parse_cable_endpoints(cable_name)
        if not a_end or not b_end:
            continue
        a_coord = loc_coords.get(a_end)
        b_coord = loc_coords.get(b_end)
        if a_coord and b_coord:
            kmz_cables[cable_name] = _snap_cable_segments(segs, a_coord, b_coord)

    # -----------------------------------------------------------------------
    # 4. Build GeoJSON features.
    # -----------------------------------------------------------------------
    features = []

    # --- Infrastructure background lines (conduit routes from supportCable) ---
    # Grouped into one MultiLineString feature to keep the GeoJSON compact.
    if infra_segments:
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "MultiLineString",
                "coordinates": [[[lon, lat] for lat, lon in seg] for seg in infra_segments],
            },
            "properties": {"layer": "infrastructure"},
        })

    # --- Cable MultiLineString features ---
    # Color is stored in properties for node-click highlighting and override
    # tracking, but lines are drawn neutral on the map until a node is clicked.
    seen_cables: set[str] = set()
    for cable, info in cable_info.items():
        if cable in seen_cables:
            continue
        seen_cables.add(cable)

        segments = kmz_cables.get(cable)
        if not segments:
            continue

        # GeoJSON coordinates are [lon, lat]
        coords_geojson = [
            [[lon, lat] for lat, lon in seg]
            for seg in segments
        ]

        _conf = cable_confidence.get(cable, {})
        features.append({
            "type": "Feature",
            "geometry": {
                "type":        "MultiLineString",
                "coordinates": coords_geojson,
            },
            "properties": {
                "cable":      cable,
                "color":      info["css"],
                "direction":  info["label"],
                "bearing":    _conf.get("bearing"),
                "confidence": _conf.get("confidence", "ok"),
            },
        })

    _missing_cables = {c for c, d in cable_confidence.items() if d.get("confidence") == "missing"}

    # --- Location point features ---
    for loc, data in location_data.items():
        missing_here = [c["name"] for c in data["cables"] if c["name"] in _missing_cables]
        features.append({
            "type": "Feature",
            "geometry": {
                "type":        "Point",
                "coordinates": [data["lon"], data["lat"]],
            },
            "properties": {
                "name":           loc,
                "cables":         data["cables"],
                "missing_cables": missing_here,
            },
        })

    return {"type": "FeatureCollection", "features": features, "summary": conf_summary}

"""
resize_logic.py -- Pure Python core for the cable resize tool.

Provides three public functions:
    scan_sheaths(wb)                      -- scans a completed asbuilt workbook
    resize_sheath(wb, record, new_ct)     -- resizes one sheath block
    apply_all_resizes(wb, resize_map)     -- applies multiple resizes safely

No Flask dependency; fully testable standalone.
"""

from __future__ import annotations

import re
import sys
import os
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Walk up from tools/resize_cables/ to the pipeline root so we can import
# config and naming_utils without the user having to set PYTHONPATH.
# ---------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent          # tools/resize_cables/
_ROOT = _HERE.parent.parent                      # Fiber Data Pipeline/
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from config import CT_BREAKDOWN, VALID_CT_SIZES, FIBER_COLORS
from naming_utils import (
    is_location_sheet,
    find_header_row,
    find_col_by_header,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Regex to extract a cable count from a sheath name like "RC73E_048CT_FT001"
_CT_RX = re.compile(r'(\d+)\s*CT', re.IGNORECASE)

# CONNECTION column values that indicate a live splice.
_LIVE_CONN = {"<--->", "<-->", "<- FUSION ->", "<- CONTINUOUS ->"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_ct(sheath_name: str) -> Optional[int]:
    """Extract the cable count from a sheath name (e.g. '48CT' -> 48)."""
    if not sheath_name:
        return None
    m = _CT_RX.search(str(sheath_name))
    if m:
        val = int(m.group(1))
        # Round to nearest valid CT size in case names use e.g. "24ct"
        return val if val in CT_BREAKDOWN else None
    return None


def _replace_ct_in_name(name: str, new_ct: int) -> str:
    """Replace the CT token in a sheath name (e.g. '48CT' -> '96CT')."""
    return _CT_RX.sub(f"{new_ct}CT", str(name), count=1)


def _copy_row_format(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """Copy cell fills and fonts from src_row to dst_row."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)


def _make_fill(hex6: str) -> PatternFill:
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")


# TIA-598 two-letter code -> ARGB hex color used for cell background
# (These are light pastel versions for readability when coloring fiber rows.)
_FIBER_HEX = {
    "BL": "ADD8E6",  # light blue
    "OR": "FFD580",  # light orange
    "GR": "90EE90",  # light green
    "BR": "D2B48C",  # light brown/tan
    "SL": "B0C4DE",  # light slate
    "WH": "FFFFFF",  # white
    "RD": "FFB6C1",  # light red/pink
    "BK": "D3D3D3",  # light gray
    "YE": "FFFF99",  # light yellow
    "VI": "EE82EE",  # violet
    "PI": "FFB6C1",  # pink
    "AQ": "7FFFD4",  # aqua
}


# ---------------------------------------------------------------------------
# Public: scan_sheaths
# ---------------------------------------------------------------------------

def scan_sheaths(wb) -> list[dict]:
    """
    Scan every location sheet in a completed asbuilt workbook and return
    a list of sheath block records.

    Each record is a dict:
        sheet         str   -- worksheet name
        sheath_uuid   str   -- col A UUID (first row of the block)
        sheath_name   str   -- col B sheath name
        current_ct    int   -- detected cable count (None if unknown)
        start_row     int   -- first data row of the block
        end_row       int   -- last data row of the block
        has_live      bool  -- True if any row has an active connection
        live_rows     list  -- rows with live connections in this block
    """
    records = []

    for sheet_name in wb.sheetnames:
        if not is_location_sheet(sheet_name):
            continue

        ws = wb[sheet_name]

        # Locate the SHEATHS header row dynamically.
        hdr_row = find_header_row(ws, "SHEATH UUID")
        if hdr_row is None:
            continue

        col_uuid = find_col_by_header(ws, "SHEATH UUID", hdr_row)
        col_name = find_col_by_header(ws, "SHEATH NAME", hdr_row)
        col_conn = find_col_by_header(ws, "CONNECTION",  hdr_row)

        if not all([col_uuid, col_name, col_conn]):
            continue

        # Walk rows from the first data row to the end of SHEATHS content.
        current_uuid  = None
        current_name  = None
        block_start   = None
        block_end     = None
        block_live    = []

        def _flush():
            nonlocal current_uuid, current_name, block_start, block_end, block_live
            if current_uuid is None or block_start is None:
                return
            ct = _parse_ct(current_name)
            has_live = bool(block_live)
            records.append({
                "sheet":       sheet_name,
                "sheath_uuid": current_uuid,
                "sheath_name": current_name,
                "current_ct":  ct,
                "start_row":   block_start,
                "end_row":     block_end,
                "has_live":    has_live,
                "live_rows":   list(block_live),
            })
            current_uuid = current_name = block_start = block_end = None
            block_live = []

        max_r = ws.max_row
        for r in range(hdr_row + 1, max_r + 1):
            uuid_val = ws.cell(r, col_uuid).value
            name_val = ws.cell(r, col_name).value
            conn_val = str(ws.cell(r, col_conn).value or "").strip()

            # A non-empty UUID signals the start of a new sheath block.
            if uuid_val and str(uuid_val).strip() not in ("", "None"):
                _flush()
                current_uuid  = str(uuid_val).strip()
                current_name  = str(name_val).strip() if name_val else ""
                block_start   = r
                block_end     = r
                if conn_val.upper() in {c.upper() for c in _LIVE_CONN}:
                    block_live.append(r)
            else:
                # Continuation row for the current sheath block.
                if block_start is not None and (uuid_val is not None or name_val is not None or conn_val):
                    block_end = r
                    if conn_val.upper() in {c.upper() for c in _LIVE_CONN}:
                        block_live.append(r)
                else:
                    # Blank row or OPTICAL SPLITTERS separator — stop scanning SHEATHS.
                    if block_start is not None and not uuid_val and not name_val and not conn_val:
                        _flush()
                        break

        _flush()

    return records


# ---------------------------------------------------------------------------
# Public: resize_sheath
# ---------------------------------------------------------------------------

def resize_sheath(wb, record: dict, new_ct: int) -> Optional[str]:
    """
    Resize a single sheath block in-place.

    Expands or shrinks the fiber rows for the given sheath to match new_ct.
    Refuses (returns an error string) if shrinking would delete rows that
    contain live connections (<--->).

    Returns None on success, or an error message string on failure.
    """
    if new_ct == record["current_ct"]:
        return None  # no change needed

    if new_ct not in CT_BREAKDOWN:
        return f"Invalid cable count: {new_ct}. Valid sizes: {VALID_CT_SIZES}"

    ws     = wb[record["sheet"]]
    start  = record["start_row"]
    end    = record["end_row"]

    current_ct  = record["current_ct"] or (end - start + 1)
    current_rows = end - start + 1
    new_rows     = new_ct

    # Locate important columns.
    hdr_row  = find_header_row(ws, "SHEATH UUID")
    col_uuid = find_col_by_header(ws, "SHEATH UUID", hdr_row)
    col_name = find_col_by_header(ws, "SHEATH NAME", hdr_row)
    col_buf  = find_col_by_header(ws, "BUFFER",      hdr_row)
    col_fib  = find_col_by_header(ws, "FIBER",       hdr_row)
    col_conn = find_col_by_header(ws, "CONNECTION",  hdr_row)

    # Try to find the right-side mirror columns (they share the same header
    # text; find_col_by_header returns the first match, so we look for all).
    from naming_utils import find_all_cols_by_header
    all_buf_cols = find_all_cols_by_header(ws, "BUFFER",  hdr_row)
    all_fib_cols = find_all_cols_by_header(ws, "FIBER",   hdr_row)
    col_buf_r = all_buf_cols[1]  if len(all_buf_cols) > 1 else None
    col_fib_r = all_fib_cols[1]  if len(all_fib_cols) > 1 else None

    # ------------------------------------------------------------------
    # Safety check for shrinking.
    # ------------------------------------------------------------------
    if new_ct < current_ct:
        rows_to_keep  = new_ct
        rows_to_cut   = current_ct - new_ct
        delete_start  = start + rows_to_keep
        delete_end    = delete_start + rows_to_cut - 1

        blocked = [
            r for r in record["live_rows"]
            if delete_start <= r <= delete_end
        ]
        if blocked:
            return (
                f"Cannot shrink '{record['sheath_name']}' from {current_ct}CT to {new_ct}CT: "
                f"{len(blocked)} live splice row(s) would be deleted "
                f"(rows {blocked[0]}–{blocked[-1]})."
            )

    # ------------------------------------------------------------------
    # Build the full TIA-598 fiber row list for new_ct.
    # ------------------------------------------------------------------
    fiber_rows = []   # list of (buffer_code, fiber_code)
    for buf_code, fibers in CT_BREAKDOWN[new_ct]:
        for fib_code in fibers:
            fiber_rows.append((buf_code, fib_code))

    # ------------------------------------------------------------------
    # Expand: insert blank rows then fill with color data.
    # ------------------------------------------------------------------
    if new_rows > current_rows:
        added = new_rows - current_rows
        ws.insert_rows(end + 1, amount=added)
        # Copy format from the last existing row.
        max_col = ws.max_column
        for offset in range(added):
            _copy_row_format(ws, end, end + 1 + offset, max_col)

    # ------------------------------------------------------------------
    # Shrink: delete extra rows from the bottom.
    # ------------------------------------------------------------------
    elif new_rows < current_rows:
        removed = current_rows - new_rows
        ws.delete_rows(start + new_rows, amount=removed)

    # ------------------------------------------------------------------
    # Write TIA-598 BUFFER and FIBER values for all new_ct rows.
    # ------------------------------------------------------------------
    for offset, (buf_code, fib_code) in enumerate(fiber_rows):
        r = start + offset

        # Leave UUID and sheath name only on the first row.
        if offset > 0:
            if col_uuid: ws.cell(r, col_uuid).value = None
            if col_name: ws.cell(r, col_name).value = None

        if col_buf:
            ws.cell(r, col_buf).value = buf_code
            ws.cell(r, col_buf).fill  = _make_fill(_FIBER_HEX.get(buf_code, "FFFFFF"))

        if col_fib:
            ws.cell(r, col_fib).value = fib_code
            ws.cell(r, col_fib).fill  = _make_fill(_FIBER_HEX.get(fib_code, "FFFFFF"))

        # Right-side mirror (same TIA-598 sequence in reverse reading order).
        if col_buf_r:
            ws.cell(r, col_buf_r).value = buf_code
            ws.cell(r, col_buf_r).fill  = _make_fill(_FIBER_HEX.get(buf_code, "FFFFFF"))

        if col_fib_r:
            ws.cell(r, col_fib_r).value = fib_code
            ws.cell(r, col_fib_r).fill  = _make_fill(_FIBER_HEX.get(fib_code, "FFFFFF"))

    # ------------------------------------------------------------------
    # Update the sheath name (CT token) in col B and right-side mirror.
    # ------------------------------------------------------------------
    old_name = record["sheath_name"]
    new_name = _replace_ct_in_name(old_name, new_ct)
    if col_name:
        ws.cell(start, col_name).value = new_name

    # Right-side sheath name mirror column (look for second "SHEATH NAME" col).
    from naming_utils import find_all_cols_by_header
    all_name_cols = find_all_cols_by_header(ws, "SHEATH NAME", hdr_row)
    if len(all_name_cols) > 1:
        ws.cell(start, all_name_cols[1]).value = new_name

    return None  # success


# ---------------------------------------------------------------------------
# Public: apply_all_resizes
# ---------------------------------------------------------------------------

def apply_all_resizes(wb, resize_map: dict) -> list[str]:
    """
    Apply multiple sheath resizes from a {sheath_uuid: new_ct} map.

    Resizes are applied per-sheet in bottom-up row order so that row
    insertions/deletions on earlier blocks don't corrupt the indices of
    later blocks within the same sheet.

    Returns a list of error message strings (empty list = all succeeded).
    """
    # Re-scan to get fresh row positions (the caller's records may be stale
    # if the workbook was already modified; re-scanning ensures accuracy).
    records = scan_sheaths(wb)

    # Filter to only the sheaths that need changing.
    to_resize = [
        r for r in records
        if r["sheath_uuid"] in resize_map
        and resize_map[r["sheath_uuid"]] != r["current_ct"]
    ]

    # Sort: process bottom-most rows first within each sheet to avoid index drift.
    to_resize.sort(key=lambda r: (r["sheet"], -r["start_row"]))

    errors = []
    for record in to_resize:
        new_ct = resize_map[record["sheath_uuid"]]
        err = resize_sheath(wb, record, new_ct)
        if err:
            errors.append(err)

    return errors
